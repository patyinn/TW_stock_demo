import time
from tkinter import Tk, Button, Label, StringVar, W,E,N,S, Frame, BooleanVar, Checkbutton, CENTER, NO, BOTTOM
from tkinter import ttk, scrolledtext, WORD, INSERT, filedialog

import numpy as np
import matplotlib.pyplot as plt
from matplotlib.pylab import mpl
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import matplotlib.ticker as mticker

import TW_ScrapperModule as scpr
import os
from datetime import datetime
from openpyxl import load_workbook

# https://www.delftstack.com/zh-tw/howto/python-tkinter/how-to-switch-frames-in-tkinter/
class SockApp(Tk):
    def __init__(self):
        Tk.__init__(self)
        self.title("TW STOCK")
        self.configure(background="light yellow")
        self.geometry("960x680")
        # self.resizable(height=False, width=False)
        self._frame = None
        self.switch_frame(StartPage)

    def switch_frame(self, frame_class):
        new_frame = frame_class(self)
        if self._frame is not None:
            self._frame.destroy()
        self._frame = new_frame
        self._frame.pack()

class StartPage(Frame):
    def __init__(self, master):
        Frame.__init__(self, master)

        # 設置資料庫位置
        self.db_path_lbl = Label(self, text="資料庫路徑: ", background="pink", font=("TkDefaultFont", 16))
        self.db_path_lbl.grid(row=0, column=0, sticky=W + E + N + S)
        self.db_path_text = StringVar()
        # if scpr.get_path_sql("db"):
        #     self.db_path = scpr.get_path_sql("db")
        # else:
        #     self.db_path = os.path.join("data", "data.db")
        self.db_path = os.path.join("data", "data.db")
        scpr.sqlite_path = self.db_path
        self.db_path_text.set(self.db_path)
        self.db_path_Entry = ttk.Entry(self, width=30, textvariable=self.db_path_text)
        self.db_path_Entry.grid(row=0, column=1, columnspan=3, sticky=W + E + N + S)
        self.db_path_btn = Button(self, text='請選擇檔案', command=self.getdbpath)
        self.db_path_btn.grid(row=0, column=4, sticky=W + E + N + S)

        Button(self, text="Go to Monthly Report Scrapper", command=lambda: master.switch_frame(M_Scrapper)).grid(row=1, column=1)
        Button(self, text="Go to Seasonal Report Scrapper", command=lambda: master.switch_frame(S_Scrapper)).grid(row=2, column=1)
        Button(self, text="Go to Price Scrapper", command=lambda: master.switch_frame(Price_Scrapper)).grid(row=3, column=1)
        Button(self, text="Go to Financial Statement Analysis", command=lambda: master.switch_frame(Page_FSAnalysis)).grid(row=4, column=1)
        Button(self, text="Go to Select Stock App", command=lambda: master.switch_frame(Page_SelectStock)).grid(row=5, column=1)
        Button(self, text="Go to Select Stock Analysis App", command=lambda: master.switch_frame(Page_StockAnalysis)).grid(row=6, column=1)

    # 取得樣板檔案位置
    def getdbpath(self):
        # 獲取文件全路徑
        filename = filedialog.askopenfilename(title='Select Template',
                                              filetypes=[('.DB', 'db')],
                                              initialdir=os.path.dirname(self.db_path))

        scpr.sqlite_path = filename
        scpr.del_path_sql("db", self.db_path_Entry.get())
        self.db_path_Entry.delete(0, 'end')
        self.db_path_Entry.insert(0, filename)
        scpr.save_path_sql(filename)

class M_Scrapper(Frame):
    def __init__(self, master):
        Frame.__init__(self, master)
        Frame.configure(self, bg='pink')

        # 選擇要爬取的資料型態
        self.title_label = Label(self, text="月報爬取: ", background="pink", font=("TkDefaultFont", 16))
        self.title_label.grid(row=0, column=0, columnspan=2, sticky=W)
        self.Frdate_label = Label(self, text="From: ", background="pink", font=("TkDefaultFont", 16))
        self.Frdate_label.grid(row=1, column=0, sticky=W)
        self.Frdate_combo = ttk.Combobox(self, postcommand=lambda: self.Frdate_combo.configure(
                                             values=scpr.date_func(table="monthly_revenue", type="F")))
        self.Frdate_combo.grid(row=1, column=1, sticky=W)

        self.Todate_label = Label(self, text="To: ", background="pink", font=("TkDefaultFont", 16))
        self.Todate_label.grid(row=1, column=2, sticky=W)
        self.Todate_combo = ttk.Combobox(self, postcommand=lambda: self.Todate_combo.configure(
                                             values=scpr.date_func(table="monthly_revenue", type="T")))
        self.Todate_combo.grid(row=1, column=3, sticky=W)

        self.Execution_btn = Button(self, text="Execute", comman=self.execute_func)
        self.Execution_btn.grid(row=1, column=4, sticky=W)

        # 顯示更新動作進度
        self.scrolltxt = scrolledtext.ScrolledText(self, wrap=WORD, height=16, width=40)
        self.scrolltxt.grid(row=2, column=0, columnspan=5, sticky=W+E+N+S, padx=20, pady=30)

        # 返回主頁面、更新、清除、離開程式
        # Label(self, text="Page one", font=('Helvetica', 18, "bold")).pack(side="top", fill="x", pady=5)
        self.back_btn = Button(self, text="Go back", command=lambda: master.switch_frame(StartPage))
        self.back_btn.grid(row=3, column=0, sticky=W)
        self.update_btn = Button(self, text="Update message", command=self.update_func)
        self.update_btn.grid(row=3, column=1, sticky=E)
        self.clear_btn = Button(self, text="Clear message", command=self.clear_func)
        self.clear_btn.grid(row=3, column=3, sticky=W)
        self.exit_btn = Button(self, text="Exit Application", command=self.quit)
        self.exit_btn.grid(row=3, column=4, sticky=W)

    # 顯示作業進度
    def update_func(self):
        To = scpr.date_func(table="monthly_revenue", type="T")
        From = scpr.date_func(table="monthly_revenue", type="F")

        self.Frdate_combo['values'] = From
        self.Todate_combo['values'] = To

    # 顯示執行項目
    def execute_func(self):
        from_Date = self.Frdate_combo.get()
        from_Date = str(from_Date.replace(" ", "-"))
        from_Date = datetime.strptime(from_Date, '%Y-%m-%d')

        to_Date = self.Todate_combo.get()
        to_Date = str(to_Date.replace(" ", "-"))
        to_Date = datetime.strptime(to_Date, '%Y-%m-%d')

        cmd = (from_Date, to_Date)

        self.scrolltxt.insert(INSERT, "正在爬取從 {} 至 {} 周期間的 月營收\n".format(cmd[0], cmd[1]))
        self.update()
        self.after(1000)
        self.Frdate_combo.delete(0, "end")
        self.Todate_combo.delete(0, "end")
        scpr.exec_func("M", cmd[0], cmd[1])
        self.scrolltxt.insert(INSERT, "完成爬取從 {} 至 {} 周期間的 月營收\n".format(cmd[0], cmd[1]))
        self.update()
        self.after(1000)

    # 清除顯示
    def clear_func(self):
        self.scrolltxt.delete(1.0, "end")
        self.Frdate_combo.delete(0, "end")
        self.Todate_combo.delete(0, "end")

class S_Scrapper(Frame):
    def __init__(self, master):
        Frame.__init__(self, master)
        Frame.configure(self, bg='pink')

        # 選擇要爬取的資料型態
        self.title_label = Label(self, text="季報爬取 (時間須包含季報發表的日期): ", background="pink", font=("TkDefaultFont", 16))
        self.title_label.grid(row=0, column=0, columnspan=4, sticky=W)
        self.Frdate_label = Label(self, text="From: ", background="pink", font=("TkDefaultFont", 16))
        self.Frdate_label.grid(row=1, column=0, sticky=W)
        self.Frdate_combo = ttk.Combobox(self, postcommand=lambda: self.Frdate_combo.configure(
                                                            values=scpr.date_func(table="balance_sheet", type="F")))
        self.Frdate_combo.grid(row=1, column=1, sticky=W)

        self.Todate_label = Label(self, text="To: ", background="pink", font=("TkDefaultFont", 16))
        self.Todate_label.grid(row=1, column=2, sticky=W)
        self.Todate_combo = ttk.Combobox(self, postcommand=lambda: self.Todate_combo.configure(
                                                            values=scpr.date_func(table="balance_sheet", type="T")))
        self.Todate_combo.grid(row=1, column=3, sticky=W)

        self.Execution_btn = Button(self, text="Execute", comman=self.execute_func)
        self.Execution_btn.grid(row=1, column=4, sticky=W)

        # 顯示更新動作進度
        self.scrolltxt = scrolledtext.ScrolledText(self, wrap=WORD, height=16, width=40)
        self.scrolltxt.grid(row=2, column=0, columnspan=5, sticky=W+E+N+S, padx=20, pady=30)

        # 返回主頁面、更新、清除、離開程式
        # Label(self, text="Page one", font=('Helvetica', 18, "bold")).pack(side="top", fill="x", pady=5)
        self.back_btn = Button(self, text="Go back", command=lambda: master.switch_frame(StartPage))
        self.back_btn.grid(row=3, column=0, sticky=W)
        self.update_btn = Button(self, text="Update message", command=self.update_func)
        self.update_btn.grid(row=3, column=1, sticky=E)
        self.clear_btn = Button(self, text="Clear message", command=self.clear_func)
        self.clear_btn.grid(row=3, column=3, sticky=W)
        self.exit_btn = Button(self, text="Exit Application", command=self.quit)
        self.exit_btn.grid(row=3, column=4, sticky=W)

    # 顯示作業進度
    def update_func(self):
        To = scpr.date_func(table="balance_sheet", type="T")
        From = scpr.date_func(table="balance_sheet", type="F")

        self.Frdate_combo['values'] = From
        self.Todate_combo['values'] = To

    # 顯示執行項目
    def execute_func(self):
        from_Date = self.Frdate_combo.get()
        from_Date = str(from_Date.replace(" ", "-"))
        from_Date = datetime.strptime(from_Date, '%Y-%m-%d')

        to_Date = self.Todate_combo.get()
        to_Date = str(to_Date.replace(" ", "-"))
        to_Date = datetime.strptime(to_Date, '%Y-%m-%d')

        cmd = (from_Date, to_Date)
        self.scrolltxt.insert(INSERT, "正在爬取從 {} 至 {} 周期間的 季財報\n".format(cmd[0], cmd[1]))
        self.update()
        self.after(1000)
        self.Frdate_combo.delete(0, "end")
        self.Todate_combo.delete(0, "end")
        scpr.exec_func("S", cmd[0], cmd[1])
        self.scrolltxt.insert(INSERT, "完成爬取從 {} 至 {} 周期間的 季財報\n".format(cmd[0], cmd[1]))
        self.update()
        self.after(1000)

    # 清除顯示
    def clear_func(self):
        self.scrolltxt.delete(1.0, "end")
        self.Frdate_combo.delete(0, "end")
        self.Todate_combo.delete(0, "end")

class Price_Scrapper(Frame):
    def __init__(self, master):
        Frame.__init__(self, master)
        Frame.configure(self, bg='pink')

        # 選擇要爬取的資料型態
        self.title_label = Label(self, text="價位爬取: ", background="pink", font=("TkDefaultFont", 16))
        self.title_label.grid(row=0, column=0, columnspan=2, sticky=W)
        self.Frdate_label = Label(self, text="From: ", background="pink", font=("TkDefaultFont", 16))
        self.Frdate_label.grid(row=1, column=0, sticky=W)
        self.Frdate_combo = ttk.Combobox(self, postcommand=lambda: self.Frdate_combo.configure(
                                             values=scpr.date_func(table="price", type="F")))
        self.Frdate_combo.grid(row=1, column=1, sticky=W)

        self.Todate_label = Label(self, text="To: ", background="pink", font=("TkDefaultFont", 16))
        self.Todate_label.grid(row=1, column=2, sticky=W)
        self.Todate_combo = ttk.Combobox(self, postcommand=lambda: self.Todate_combo.configure(
                                             values=scpr.date_func(table="price", type="T")))
        self.Todate_combo.grid(row=1, column=3, sticky=W)

        self.Execution_btn = Button(self, text="Execute", comman=self.execute_func)
        self.Execution_btn.grid(row=1, column=4, sticky=W)

        # 顯示更新動作進度
        self.scrolltxt = scrolledtext.ScrolledText(self, wrap=WORD, height=16, width=40)
        self.scrolltxt.grid(row=2, column=0, columnspan=5, sticky=W+E+N+S, padx=20, pady=30)

        # 返回主頁面、更新、清除、離開程式
        # Label(self, text="Page one", font=('Helvetica', 18, "bold")).pack(side="top", fill="x", pady=5)
        self.back_btn = Button(self, text="Go back", command=lambda: master.switch_frame(StartPage))
        self.back_btn.grid(row=3, column=0, sticky=W)
        self.update_btn = Button(self, text="Update message", command=self.update_func)
        self.update_btn.grid(row=3, column=1, sticky=E)
        self.clear_btn = Button(self, text="Clear message", command=self.clear_func)
        self.clear_btn.grid(row=3, column=3, sticky=W)
        self.exit_btn = Button(self, text="Exit Application", command=self.quit)
        self.exit_btn.grid(row=3, column=4, sticky=W)

    # 顯示作業進度
    def update_func(self):
        To = scpr.date_func(table="price", type="T")
        From = scpr.date_func(table="price", type="F")

        self.Frdate_combo['values'] = From
        self.Todate_combo['values'] = To

    # 顯示執行項目
    def execute_func(self):
        from_Date = self.Frdate_combo.get()
        from_Date = str(from_Date.replace(" ", "-"))
        from_Date = datetime.strptime(from_Date, '%Y-%m-%d')

        to_Date = self.Todate_combo.get()
        to_Date = str(to_Date.replace(" ", "-"))
        to_Date = datetime.strptime(to_Date, '%Y-%m-%d')

        cmd = (from_Date, to_Date)
        self.scrolltxt.insert(INSERT, "正在爬取從 {} 至 {} 周期間的 價位\n".format(cmd[0], cmd[1]))
        self.update()
        self.after(1000)
        self.Frdate_combo.delete(0, "end")
        self.Todate_combo.delete(0, "end")
        scpr.exec_func("P", cmd[0], cmd[1])
        self.scrolltxt.insert(INSERT, "完成爬取從 {} 至 {} 周期間的 價位\n".format(cmd[0], cmd[1]))
        self.update()
        self.after(1000)

    # 清除顯示
    def clear_func(self):
        self.scrolltxt.delete(1.0, "end")
        self.Frdate_combo.delete(0, "end")
        self.Todate_combo.delete(0, "end")

class Page_FSAnalysis(Frame):
    def __init__(self, master):
        Frame.__init__(self, master)
        Frame.configure(self, bg='pink')

        # 設置選取樣板的資料夾及檔案按鈕，並取得路徑
        self.tplt_path_lbl = Label(self, text="樣板路徑: ", background="pink", font=("TkDefaultFont", 16))
        self.tplt_path_lbl.grid(row=0, column=0, sticky=W + E + N + S)
        self.tplt_path_text = StringVar()
        if not scpr.get_path_sql("file"):
            self.tplt_path = os.path.abspath('')
        else:
            self.tplt_path = scpr.get_path_sql("file")[-1]
        self.tplt_path_text.set(self.tplt_path)
        self.tplt_path_combo = ttk.Combobox(self, width=70, textvariable=self.tplt_path_text,
                                            postcommand=lambda: self.tplt_path_combo.configure(values=scpr.get_path_sql("file")))
        self.tplt_path_combo.grid(row=0, column=1, columnspan=3, sticky=W + E + N + S)
        self.tplt_path_btn = Button(self, text='請選擇檔案', command=self.gettpltpath)
        self.tplt_path_btn.grid(row=0, column=4, sticky=W + E + N + S)
        self.del_tplt_path_btn = Button(self, text='刪除紀錄', command=self.del_tplt)
        self.del_tplt_path_btn.grid(row=0, column=5, sticky=W + E + N + S)

        # 設置選取要更新的資料夾與檔案按鈕，並取得路徑
        self.path_lbl = Label(self, text="資料夾路徑: ", background="pink", font=("TkDefaultFont", 16))
        self.path_lbl.grid(row=1, column=0, sticky=W + E + N + S)
        self.path_text = StringVar()
        if not scpr.get_path_sql("directory"):
            self.path = os.path.abspath('')
        else:
            self.path = scpr.get_path_sql("directory")[-1]
        self.path_text.set(self.path)
        self.path_combo = ttk.Combobox(self, width=70, textvariable=self.path_text,
                                       postcommand=lambda: self.path_combo.configure(values=scpr.get_path_sql("directory")))
        self.path_combo.grid(row=1, column=1, columnspan=3, sticky=W + E + N + S)
        self.path_btn = Button(self, text='請選擇檔案', command=self.getpath)
        self.path_btn.grid(row=1, column=4, sticky=W + E + N + S)
        self.del_path_btn = Button(self, text='刪除紀錄', command=self.del_path)
        self.del_path_btn.grid(row=1, column=5, sticky=W + E + N + S)

        # 欲更新財報分析excel編號設定、執行的項目
        self.symbol_label = Label(self, text="Symbol: ", background="pink", font=("TkDefaultFont", 16))
        self.symbol_label.grid(row=2, column=0, sticky=W)
        self.symbol_text = StringVar()
        self.symbol_list = self.getfileid()[0]
        self.symbol_list.insert(0, "all")
        self.symbol_combo = ttk.Combobox(self, textvariable=self.symbol_text, values=self.symbol_list,
                                         postcommand=self.update_func)
        self.symbol_combo.grid(row=2, column=1, sticky=W)

        # 選擇要執行的項目
        self.exec_label = Label(self, text="執行項目: ", background="pink", font=("TkDefaultFont", 16))
        self.exec_label.grid(row=3, column=0, sticky=W)
        self.exec_combo = ttk.Combobox(self, values=["all", "更新月報", "更新季報", "更新PER與今日價位", "更新股東占比"])
        self.exec_combo.grid(row=3, column=1, sticky=W)
        self.Execution_btn = Button(self, text="Execute", comman=self.execute_func)
        self.Execution_btn.grid(row=3, column=2, sticky=W)

        # 顯示更新動作進度
        self.scrolltxt = scrolledtext.ScrolledText(self, wrap=WORD, height=16, width=40)
        self.scrolltxt.grid(row=4, column=0, columnspan=5, sticky=W+E+N+S, padx=20, pady=30)

        # 返回主頁面、更新、清除、離開程式
        self.back_btn = Button(self, text="Go back", command=lambda: master.switch_frame(StartPage))
        self.back_btn.grid(row=5, column=0, columnspan=2, sticky=W)
        self.update_btn = Button(self, text="Update message", command=self.update_func)
        self.update_btn.grid(row=5, column=1, columnspan=2, sticky=W)
        self.clear_btn = Button(self, text="Clear message", command=self.clear_func)
        self.clear_btn.grid(row=5, column=2, columnspan=2, sticky=W)
        self.exit_btn = Button(self, text="Exit Application", command=self.quit)
        self.exit_btn.grid(row=5, column=3, columnspan=2, sticky=W)

    # 刪除已儲存的樣板路徑
    def del_tplt(self):
        if self.tplt_path_combo.get():
            path = self.tplt_path_combo.get()
            type = "file"
            scpr.del_path_sql(type, path)

    # 刪除已儲存的資料夾路徑
    def del_path(self):
        if self.path_combo.get():
            path = self.path_combo.get()
            type = "directory"
            scpr.del_path_sql(type, path)

    # 更新股票代號
    def update_func(self):
        self.path = self.path_combo.get()
        symbol = self.getfileid()[0]
        symbol.insert(0, "all")
        self.symbol_combo['values'] = symbol

    # 顯示執行項目
    def execute_func(self):
        exec = self.exec_combo.get()
        symbol = self.symbol_text.get()
        list_id, list_dict = self.getfileid()

        if symbol == "all":
            Stock_ID_list = list_id
        else:
            Stock_ID_list = str(symbol).replace(" ", ",").split(",")
            Stock_ID_list = [i for i in Stock_ID_list if i.isdigit()]

        for id in Stock_ID_list:
            if id not in list_id:
                folder_path = os.path.join(self.path_text.get(), "自選新增")
                self.SaveExcel(id, folder=folder_path)
                File_path = os.path.join(folder_path, "O_" + id + "_財報分析.xlsx")
                time.sleep(3)
            else:
                File_path = list_dict[id]

            FSA = scpr.FinancialAnalysis(File_path)

            try:
                if exec == "all":
                    FSA.Update_Monthly_report(id, path=File_path)
                    self.scrolltxt.insert(INSERT, "完成更新 {} 的 月報\n".format(id))
                    self.update()
                    self.after(1000)
                    FSA.Update_Season_report(id, path=File_path)
                    self.scrolltxt.insert(INSERT, "完成更新 {} 的 季報\n".format(id))
                    self.update()
                    self.after(1000)
                    FSA.Update_CashFlow(id, path=File_path)
                    self.scrolltxt.insert(INSERT, "完成更新 {} 的 現金流量表\n".format(id))
                    self.update()
                    self.after(1000)
                    FSA.Update_PRICEToday(id, path=File_path)
                    self.scrolltxt.insert(INSERT, "完成更新 {} 的 價位\n".format(id))
                    self.update()
                    self.after(1000)
                    FSA.Update_PER(id, path=File_path)
                    self.scrolltxt.insert(INSERT, "完成更新 {} 的 本益比\n".format(id))
                    self.update()
                    self.after(1000)
                    FSA.Update_Directors_and_supervisors(id, path=File_path)
                    self.scrolltxt.insert(INSERT, "完成更新 {} 的 股東占比\n".format(id))
                    self.update()
                    self.after(1000)
                elif exec == "更新月報":
                    FSA.Update_Monthly_report(id, path=File_path)
                    self.scrolltxt.insert(INSERT, "完成更新 {} 的 月報\n".format(id))
                    self.update()
                    self.after(1000)
                    FSA.Update_PRICEToday(id, path=File_path)
                    self.scrolltxt.insert(INSERT, "完成更新 {} 的 價位\n".format(id))
                    self.update()
                    self.after(1000)
                    FSA.Update_PER(id, path=File_path)
                    self.scrolltxt.insert(INSERT, "完成更新 {} 的 本益比\n".format(id))
                    self.update()
                    self.after(1000)
                elif exec == "更新季報":
                    FSA.Update_Season_report(id, path=File_path)
                    self.scrolltxt.insert(INSERT, "完成更新 {} 的 季報\n".format(id))
                    self.update()
                    self.after(1000)
                    FSA.Update_CashFlow(id, path=File_path)
                    self.scrolltxt.insert(INSERT, "完成更新 {} 的 現金流量表\n".format(id))
                    self.update()
                    self.after(1000)
                    FSA.Update_PRICEToday(id, path=File_path)
                    self.scrolltxt.insert(INSERT, "完成更新 {} 的 價位\n".format(id))
                    self.update()
                    self.after(1000)
                    FSA.Update_PER(id, path=File_path)
                    self.scrolltxt.insert(INSERT, "完成更新 {} 的 本益比\n".format(id))
                    self.update()
                    self.after(1000)
                elif exec == "更新PER與今日價位":
                    FSA.Update_PRICEToday(id, path=File_path)
                    self.scrolltxt.insert(INSERT, "完成更新 {} 的 價位\n".format(id))
                    self.update()
                    self.after(1000)
                    FSA.Update_PER(id, path=File_path)
                    self.scrolltxt.insert(INSERT, "完成更新 {} 的 本益比\n".format(id))
                    self.update()
                    self.after(1000)
                elif exec == "更新股東占比":
                    FSA.Update_PRICEToday(id, path=File_path)
                    self.scrolltxt.insert(INSERT, "完成更新 {} 的 價位\n".format(id))
                    self.update()
                    self.after(1000)
                    FSA.Update_PER(id, path=File_path)
                    self.scrolltxt.insert(INSERT, "完成更新 {} 的 本益比\n".format(id))
                    self.update()
                    self.after(1000)
                    FSA.Update_Directors_and_supervisors(id, path=File_path)
                    self.scrolltxt.insert(INSERT, "完成更新 {} 的 股東占比\n".format(id))
                    self.update()
                    self.after(1000)
                else:
                    self.scrolltxt.insert(INSERT, "輸入有誤\n")
                self.scrolltxt.insert(INSERT, "完成\n")

            except Exception as e:
                self.scrolltxt.insert(INSERT, "{}發生問題，問題原因: {}\n".format(id, e))
                print("{}發生問題，問題原因: {}\n".format(id, e))


        path = self.tplt_path_text.get()
        directory = self.path_text.get()
        scpr.save_path_sql(path)
        scpr.save_path_sql(directory)

    # 清除顯示
    def clear_func(self):
        self.scrolltxt.delete(1.0, "end")
        self.symbol_combo.delete(0, "end")
        self.exec_combo.delete(0, "end")

    # 取得樣板檔案位置
    def gettpltpath(self):
        # 獲取文件全路徑
        filename = filedialog.askopenfilename(title='Select Template',
                                              filetypes=[('.XLSX', 'xlsx'), ('All Files', '*')],
                                              initialdir=os.path.dirname(self.tplt_path))
        self.tplt_path_combo.delete(0, 'end')
        self.tplt_path_combo.insert(0, filename)

    # 欲更新檔案位置
    def getpath(self):
        directory = filedialog.askdirectory(title='Select directory',
                                            initialdir=self.path)
        self.path_combo.delete(0, 'end')
        self.path_combo.insert(0, directory)

    # 取得現有檔案的代號
    def getfileid(self):
        target_folder = self.path
        file = scpr.show_folder_content(target_folder, prefix="O_", postfix=".xlsx")
        index = []
        dictionary = {}
        for num in file[0:]:
            idx = ''.join([x for x in num if x.isdigit()])
            dictionary[idx] = num
            index.append(idx)
        return index, dictionary

    # 以樣板儲存新檔案
    def SaveExcel(self, ID, folder):
        path = self.tplt_path_text.get()
        wb = load_workbook(path)
        new_path = os.path.join(folder, "O_" + ID + "_財報分析.xlsx")
        wb.save(new_path)

class Page_SelectStock(Frame):
    def __init__(self, master):
        Frame.__init__(self, master)
        Frame.configure(self, bg='pink')
        self.selected_stock = []
        self.chk_list = []
        self.chkvar_list = []
        self.content_list = []

        # 設置選取樣板的資料夾及檔案按鈕，並取得路徑
        self.tplt_path_lbl = Label(self, text="樣板路徑: ", background="pink", font=("TkDefaultFont", 16))
        self.tplt_path_lbl.grid(row=0, column=0, sticky=W + E + N + S)
        self.tplt_path_text = StringVar()
        if not scpr.get_path_sql("file"):
            self.tplt_path = os.path.abspath('')
        else:
            self.tplt_path = scpr.get_path_sql("file")[-1]
        self.tplt_path_text.set(self.tplt_path)
        self.tplt_path_combo = ttk.Combobox(self, width=30, textvariable=self.tplt_path_text,
                                            postcommand=lambda: self.tplt_path_combo.configure(values=scpr.get_path_sql("file")))
        self.tplt_path_combo.grid(row=0, column=1, columnspan=5, sticky=W + E + N + S)
        self.tplt_path_btn = Button(self, text='請選擇檔案', command=self.gettpltpath)
        self.tplt_path_btn.grid(row=0, column=6, sticky=W + E + N + S)
        self.del_tplt_path_btn = Button(self, text='刪除紀錄', command=self.del_tplt)
        self.del_tplt_path_btn.grid(row=0, column=7, sticky=W + E + N + S)

        # 設置選取要更新的資料夾與檔案按鈕，並取得路徑
        self.path_lbl = Label(self, text="資料夾路徑: ", background="pink", font=("TkDefaultFont", 16))
        self.path_lbl.grid(row=1, column=0, sticky=W + E + N + S)
        self.path_text = StringVar()

        if not scpr.get_path_sql("SSdirectory"):
            self.path = os.path.abspath('')
        else:
            self.path = scpr.get_path_sql("SSdirectory")[-1]
        self.path_text.set(self.path)
        self.path_combo = ttk.Combobox(self, width=30, textvariable=self.path_text,
                                       postcommand=lambda: self.path_combo.configure(values=scpr.get_path_sql("SSdirectory")))
        self.path_combo.grid(row=1, column=1, columnspan=5, sticky=W + E + N + S)
        self.path_btn = Button(self, text='請選擇檔案', command=self.getpath)
        self.path_btn.grid(row=1, column=6, sticky=W + E + N + S)
        self.del_path_btn = Button(self, text='刪除紀錄', command=self.del_path)
        self.del_path_btn.grid(row=1, column=7, sticky=W + E + N + S)

        # 選取欲使用的條件以及其設定值
        Label(self, text="選股條件:", bg="red", font=("TkDefaultFont", 14)).grid(row=2, column=0, columnspan=6, sticky=W+E)

        self.start_label = Label(self, text="選股日期:", bg="pink", font=("TkDefaultFont", 12))
        self.start_label.grid(row=3, column=0, sticky=W+E)
        self.start_var1 = StringVar()
        self.start = ttk.Entry(self, textvariable=self.start_var1, font=("TkDefaultFont", 12))
        self.start.grid(row=3, column=1, columnspan=3, sticky=W)
        cache_start = scpr.get_cache_sql("選股日期:")
        if not cache_start.empty:
            self.start.insert(0, cache_start["entry"][0])

        self.entry_var1 = StringVar()
        self.entry1 = ttk.Entry(self, textvariable=self.entry_var1, width=15, font=("TkDefaultFont", 12))
        self.entry1.grid(row=4, column=2, sticky=W)
        self.combo1 = ttk.Combobox(self, width=4, values=[">", ">=", "=", "<", "<="])
        self.combo1.grid(row=4, column=1, sticky=W)
        self.chkvar1 = BooleanVar()
        self.chk1 = Checkbutton(self, variable=self.chkvar1, bg="pink", text="市值", font=("TkDefaultFont", 12))
        self.chk1.grid(row=4, column=0, sticky=E)
        cache1 = scpr.get_cache_sql("市值")
        if not cache1.empty:
            self.chkvar1.set(scpr.get_cache_sql("市值", bool=True))
            self.entry1.insert(0, cache1["entry"][0]) if self.chkvar1.get() else self.entry1.delete(0, "end")
            self.combo1.insert(0, cache1["combo"][0]) if self.chkvar1.get() else self.entry1.delete(0, "end")

        self.entry_var2 = StringVar()
        self.entry2 = ttk.Entry(self, textvariable=self.entry_var2, width=15, font=("TkDefaultFont", 12))
        self.entry2.grid(row=4, column=5, sticky=W)
        self.combo2 = ttk.Combobox(self, width=4, values=[">", ">=", "=", "<", "<="])
        self.combo2.grid(row=4, column=4, sticky=W)
        self.chkvar2 = BooleanVar()
        self.chk2 = Checkbutton(self, variable=self.chkvar2, bg="pink", text="三年自由現金流", font=("TkDefaultFont", 12))
        self.chk2.grid(row=4, column=3, sticky=E)
        cache2 = scpr.get_cache_sql("三年自由現金流")
        if not cache2.empty:
            self.chkvar2.set(scpr.get_cache_sql("三年自由現金流", bool=True))
            self.entry2.insert(0, cache2["entry"][0]) if self.chkvar2.get() else self.entry2.delete(0, "end")
            self.combo2.insert(0, cache2["combo"][0]) if self.chkvar2.get() else self.entry2.delete(0, "end")


        self.entry_var3 = StringVar()
        self.entry3 = ttk.Entry(self, textvariable=self.entry_var3, width=15, font=("TkDefaultFont", 12))
        self.entry3.grid(row=5, column=2, sticky=W)
        self.combo3 = ttk.Combobox(self, width=4, values=[">", ">=", "=", "<", "<="])
        self.combo3.grid(row=5, column=1, sticky=W)
        self.chkvar3 = BooleanVar()
        self.chk3 = Checkbutton(self, variable=self.chkvar3, bg="pink", text="股東權益報酬率", font=("TkDefaultFont", 12))
        self.chk3.grid(row=5, column=0, sticky=E)
        cache3 = scpr.get_cache_sql("股東權益報酬率")
        if not cache3.empty:
            self.chkvar3.set(scpr.get_cache_sql("股東權益報酬率", bool=True))
            self.entry3.insert(0, cache3["entry"][0]) if self.chkvar3.get() else self.entry3.delete(0, "end")
            self.combo3.insert(0, cache3["combo"][0]) if self.chkvar3.get() else self.entry3.delete(0, "end")

        self.entry_var4 = StringVar()
        self.entry4 = ttk.Entry(self, textvariable=self.entry_var4, width=15, font=("TkDefaultFont", 12))
        self.entry4.grid(row=5, column=5, sticky=W)
        self.combo4 = ttk.Combobox(self, width=4, values=[">", ">=", "=", "<", "<="])
        self.combo4.grid(row=5, column=4, sticky=W)
        self.chkvar4 = BooleanVar()
        self.chk4 = Checkbutton(self, variable=self.chkvar4, bg="pink", text="營業利益年成長率", font=("TkDefaultFont", 12))
        self.chk4.grid(row=5, column=3, sticky=E)
        cache4 = scpr.get_cache_sql("營業利益年成長率")
        if not cache4.empty:
            self.chkvar4.set(scpr.get_cache_sql("營業利益年成長率", bool=True))
            self.entry4.insert(0, cache4["entry"][0]) if self.chkvar4.get() else self.entry4.delete(0, "end")
            self.combo4.insert(0, cache4["combo"][0]) if self.chkvar4.get() else self.entry4.delete(0, "end")

        self.entry_var5 = StringVar()
        self.entry5 = ttk.Entry(self, textvariable=self.entry_var5, width=15, font=("TkDefaultFont", 12))
        self.entry5.grid(row=6, column=2, sticky=W)
        self.combo5 = ttk.Combobox(self, width=4, values=[">", ">=", "=", "<", "<="])
        self.combo5.grid(row=6, column=1, sticky=W)
        self.chkvar5 = BooleanVar()
        self.chk5 = Checkbutton(self, variable=self.chkvar5, bg="pink", text="八季營益率變化", font=("TkDefaultFont", 12))
        self.chk5.grid(row=6, column=0, sticky=E)
        cache5 = scpr.get_cache_sql("八季營益率變化")
        if not cache5.empty:
            self.chkvar5.set(scpr.get_cache_sql("八季營益率變化", bool=True))
            self.entry5.insert(0, cache5["entry"][0]) if self.chkvar5.get() else self.entry5.delete(0, "end")
            self.combo5.insert(0, cache5["combo"][0]) if self.chkvar5.get() else self.entry5.delete(0, "end")

        self.entry_var6 = StringVar()
        self.entry6 = ttk.Entry(self, textvariable=self.entry_var6, width=15, font=("TkDefaultFont", 12))
        self.entry6.grid(row=6, column=5, sticky=W)
        self.combo6 = ttk.Combobox(self, width=4, values=[">", ">=", "=", "<", "<="])
        self.combo6.grid(row=6, column=4, sticky=W)
        self.chkvar6 = BooleanVar()
        self.chk6 = Checkbutton(self, variable=self.chkvar6, bg="pink", text="市值營收比", font=("TkDefaultFont", 12))
        self.chk6.grid(row=6, column=3, sticky=E)
        cache6 = scpr.get_cache_sql("市值營收比")
        if not cache6.empty:
            self.chkvar6.set(scpr.get_cache_sql("市值營收比", bool=True))
            self.entry6.insert(0, cache6["entry"][0]) if self.chkvar6.get() else self.entry6.delete(0, "end")
            self.combo6.insert(0, cache6["combo"][0]) if self.chkvar6.get() else self.entry6.delete(0, "end")

        self.entry_var7 = StringVar()
        self.entry7 = ttk.Entry(self, textvariable=self.entry_var7, width=15, font=("TkDefaultFont", 12))
        self.entry7.grid(row=7, column=2, sticky=W)
        self.combo7 = ttk.Combobox(self, width=4, values=[">", ">=", "=", "<", "<="])
        self.combo7.grid(row=7, column=1, sticky=W)
        self.chkvar7 = BooleanVar()
        self.chk7 = Checkbutton(self, variable=self.chkvar7, bg="pink", text="短期營收年增", font=("TkDefaultFont", 12))
        self.chk7.grid(row=7, column=0, sticky=E)
        cache7 = scpr.get_cache_sql("短期營收年增")
        if not cache7.empty:
            self.chkvar7.set(scpr.get_cache_sql("短期營收年增", bool=True))
            self.entry7.insert(0, cache7["entry"][0]) if self.chkvar7.get() else self.entry7.delete(0, "end")
            self.combo7.insert(0, cache7["combo"][0]) if self.chkvar7.get() else self.entry7.delete(0, "end")

        self.entry_var8 = StringVar()
        self.entry8 = ttk.Entry(self, textvariable=self.entry_var8, width=15, font=("TkDefaultFont", 12))
        self.entry8.grid(row=7, column=5, sticky=W)
        self.combo8 = ttk.Combobox(self, width=4, values=[">", ">=", "=", "<", "<="])
        self.combo8.grid(row=7, column=4, sticky=W)
        self.chkvar8 = BooleanVar()
        self.chk8 = Checkbutton(self, variable=self.chkvar8, bg="pink", text="短期營收年增", font=("TkDefaultFont", 12))
        self.chk8.grid(row=7, column=3, sticky=E)
        cache8 = scpr.get_cache_sql("短期營收年增2")
        if not cache8.empty:
            self.chkvar8.set(scpr.get_cache_sql("短期營收年增2", bool=True))
            self.entry8.insert(0, cache8["entry"][0]) if self.chkvar8.get() else self.entry8.delete(0, "end")
            self.combo8.insert(0, cache8["combo"][0]) if self.chkvar8.get() else self.entry8.delete(0, "end")

        self.entry_var9 = StringVar()
        self.entry9 = ttk.Entry(self, textvariable=self.entry_var9, width=15, font=("TkDefaultFont", 12))
        self.entry9.grid(row=8, column=2, sticky=W)
        self.combo9 = ttk.Combobox(self, width=4, values=[">", ">=", "=", "<", "<="])
        self.combo9.grid(row=8, column=1, sticky=W)
        self.chkvar9 = BooleanVar()
        self.chk9 = Checkbutton(self, variable=self.chkvar9, bg="pink", text="短期淨利年增", font=("TkDefaultFont", 12))
        self.chk9.grid(row=8, column=0, sticky=E)
        cache9 = scpr.get_cache_sql("短期淨利年增")
        if not cache9.empty:
            self.chkvar9.set(scpr.get_cache_sql("短期淨利年增", bool=True))
            self.entry9.insert(0, cache9["entry"][0]) if self.chkvar9.get() else self.entry9.delete(0, "end")
            self.combo9.insert(0, cache9["combo"][0]) if self.chkvar9.get() else self.entry9.delete(0, "end")

        self.entry_var10 = StringVar()
        self.entry10 = ttk.Entry(self, textvariable=self.entry_var10, width=15, font=("TkDefaultFont", 12))
        self.entry10.grid(row=8, column=5, sticky=W)
        self.combo10 = ttk.Combobox(self, width=4, values=[">", ">=", "=", "<", "<="])
        self.combo10.grid(row=8, column=4, sticky=W)
        self.chkvar10 = BooleanVar()
        self.chk10 = Checkbutton(self, variable=self.chkvar10, bg="pink", text="存貨周轉變化率", font=("TkDefaultFont", 12))
        self.chk10.grid(row=8, column=3, sticky=E)
        cache10 = scpr.get_cache_sql("存貨周轉變化率")
        if not cache10.empty:
            self.chkvar10.set(scpr.get_cache_sql("存貨周轉變化率", bool=True))
            self.entry10.insert(0, cache10["entry"][0]) if self.chkvar10.get() else self.entry10.delete(0, "end")
            self.combo10.insert(0, cache10["combo"][0]) if self.chkvar10.get() else self.entry10.delete(0, "end")

        self.entry_var11 = StringVar()
        self.entry11 = ttk.Entry(self, textvariable=self.entry_var11, width=15, font=("TkDefaultFont", 12))
        self.entry11.grid(row=9, column=2, sticky=W)
        self.combo11 = ttk.Combobox(self, width=4, values=[">", ">=", "=", "<", "<="])
        self.combo11.grid(row=9, column=1, sticky=W)
        self.chkvar11 = BooleanVar()
        self.chk11 = Checkbutton(self, variable=self.chkvar11, bg="pink", text="rsv", font=("TkDefaultFont", 12))
        self.chk11.grid(row=9, column=0, sticky=E)
        cache11 = scpr.get_cache_sql("rsv")
        if not cache11.empty:
            self.chkvar11.set(scpr.get_cache_sql("rsv", bool=True))
            self.entry11.insert(0, cache11["entry"][0]) if self.chkvar11.get() else self.entry11.delete(0, "end")
            self.combo11.insert(0, cache11["combo"][0]) if self.chkvar11.get() else self.entry11.delete(0, "end")


        # 選擇要執行的項目
        self.Execution_btn = Button(self, text="Execute", command=self.execute_func)
        self.Execution_btn.grid(row=9, column=6, sticky=W+E)
        self.Save_btn = Button(self, text="Save excel", command=self.save_excel)
        self.Save_btn.grid(row=9, column=7, sticky=W+E)

        # 回測設置
        Label(self, text="回測設定:", bg="red", font=("TkDefaultFont", 14)).grid(row=10, column=0, columnspan=6, sticky=W+E)

        # 回測的起始時間
        self.end_label = Label(self, text="回測起始日期:", bg="pink", font=("TkDefaultFont", 12))
        self.end_label.grid(row=11, column=0, sticky=W+E)
        self.end_var1 = StringVar()
        self.end = ttk.Entry(self, textvariable=self.end_var1, font=("TkDefaultFont", 12))
        self.end.grid(row=11, column=1, columnspan=2, sticky=W)
        cache_end = scpr.get_cache_sql("回測起始日期:")
        if not cache_end.empty:
            self.end.insert(0, cache_end["entry"][0])

        # 多少週期更新一次
        self.period_label = Label(self, text="週期天數:", bg="pink", font=("TkDefaultFont", 12))
        self.period_label.grid(row=11, column=3, sticky=W+E)
        self.period_var1 = StringVar()
        self.period = ttk.Entry(self, textvariable=self.period_var1, font=("TkDefaultFont", 12))
        self.period.grid(row=11, column=4, columnspan=2, sticky=W)
        cache_period = scpr.get_cache_sql("週期天數:")
        if not cache_period.empty:
            self.period.insert(0, cache_period["entry"][0])

        # 是否停利
        self.sp_var = StringVar()
        self.sp_entry = ttk.Entry(self, textvariable=self.sp_var, font=("TkDefaultFont", 12))
        self.sp_entry.grid(row=12, column=1, columnspan=2, sticky=W)
        self.sp_chkvar = BooleanVar()
        self.sp_chk = Checkbutton(self, variable=self.sp_chkvar, bg="pink", text="停利", font=("TkDefaultFont", 12))
        self.sp_chk.grid(row=12, column=0, sticky=W+E)
        sp_cache = scpr.get_cache_sql("停利")
        if not sp_cache.empty:
            self.sp_chkvar.set(scpr.get_cache_sql("停利", bool=True))
            self.sp_entry.insert(0, sp_cache["entry"][0]) if self.sp_chkvar.get() else self.sp_entry.delete(0, "end")

        # 是否停損
        self.sl_var = StringVar()
        self.sl_entry = ttk.Entry(self, textvariable=self.sl_var, font=("TkDefaultFont", 12))
        self.sl_entry.grid(row=12, column=4, columnspan=2, sticky=W)
        self.sl_chkvar = BooleanVar()
        self.sl_chk = Checkbutton(self, variable=self.sl_chkvar, bg="pink", text="停損", font=("TkDefaultFont", 12))
        self.sl_chk.grid(row=12, column=3, sticky=W+E)
        sl_cache = scpr.get_cache_sql("停損")
        if not sl_cache.empty:
            self.sl_chkvar.set(scpr.get_cache_sql("停損", bool=True))
            self.sl_entry.insert(0, sl_cache["entry"][0]) if self.sl_chkvar.get() else self.sl_entry.delete(0, "end")

        # 執行回測
        self.backtest_btn = Button(self, text='執行回測', command=self.backtest_func)
        self.backtest_btn.grid(row=12, column=6, sticky=W+E)

        # 顯示更新動作進度
        self.scrolltxt = scrolledtext.ScrolledText(self, wrap=WORD, height=10, width=20)
        self.scrolltxt.grid(row=13, column=0, columnspan=6, sticky=W+E+N+S, padx=20, pady=30)

        # 返回主頁面、更新、清除、離開程式
        self.back_btn = Button(self, text="Go back", command=lambda: master.switch_frame(StartPage))
        self.back_btn.grid(row=18, column=0, columnspan=2, sticky=W)
        self.update_btn = Button(self, text="Update message", command=self.update_func)
        self.update_btn.grid(row=18, column=2, columnspan=2, sticky=W)
        self.clear_btn = Button(self, text="Clear message", command=self.clear_func)
        self.clear_btn.grid(row=18, column=4, columnspan=2, sticky=W)
        self.exit_btn = Button(self, text="Exit Application", command=self.quit)
        self.exit_btn.grid(row=18, column=6, columnspan=2, sticky=W)

    # 取得樣板檔案位置
    def gettpltpath(self):
        # 獲取文件全路徑
        filename = filedialog.askopenfilename(title='Select Template',
                                              filetypes=[('.XLSX', 'xlsx'), ('All Files', '*')],
                                              initialdir=os.path.dirname(self.tplt_path))
        self.tplt_path_combo.delete(0, 'end')
        self.tplt_path_combo.insert(0, filename)

    # 刪除已儲存的樣板路徑
    def del_tplt(self):
        if self.tplt_path_combo.get():
            path = self.tplt_path_combo.get()
            type = "file"
            scpr.del_path_sql(type, path)

    # 欲更新檔案位置
    def getpath(self):
        directory = filedialog.askdirectory(title='Select directory',
                                            initialdir=self.path)
        self.path_combo.delete(0, 'end')
        self.path_combo.insert(0, directory)

    # 刪除已儲存的資料夾路徑
    def del_path(self):
        if self.path_combo.get():
            path = self.path_combo.get()
            type = "SSdirectory"
            scpr.del_path_sql(type, path)

    # 選定後，自動帶入上次執行成功的條件
    def save_cache(self):
        self.chk_list = [
            self.chk1.cget("text"),
            self.chk2.cget("text"),
            self.chk3.cget("text"),
            self.chk4.cget("text"),
            self.chk5.cget("text"),
            self.chk6.cget("text"),
            self.chk7.cget("text"),
            self.chk8.cget("text") + "2",
            self.chk9.cget("text"),
            self.chk10.cget("text"),
            self.chk11.cget("text"),
            self.start_label.cget("text"),
            self.end_label.cget("text"),
            self.period_label.cget("text"),
            self.sp_chk.cget("text"),
            self.sl_chk.cget("text")
        ]
        self.chkvar_list = [
            self.chkvar1.get(),
            self.chkvar2.get(),
            self.chkvar3.get(),
            self.chkvar4.get(),
            self.chkvar5.get(),
            self.chkvar6.get(),
            self.chkvar7.get(),
            self.chkvar8.get(),
            self.chkvar9.get(),
            self.chkvar10.get(),
            self.chkvar11.get(),
            False,
            False,
            False,
            self.sp_chkvar.get(),
            self.sl_chkvar.get()
        ]
        self.combo_list = [
            self.combo1.get(),
            self.combo2.get(),
            self.combo3.get(),
            self.combo4.get(),
            self.combo5.get(),
            self.combo6.get(),
            self.combo7.get(),
            self.combo8.get(),
            self.combo9.get(),
            self.combo10.get(),
            self.combo11.get(),
            None,
            None,
            None,
            None,
            None
        ]
        self.entry_list = [
            self.entry1.get(),
            self.entry2.get(),
            self.entry3.get(),
            self.entry4.get(),
            self.entry5.get(),
            self.entry6.get(),
            self.entry7.get(),
            self.entry8.get(),
            self.entry9.get(),
            self.entry10.get(),
            self.entry11.get(),
            self.start.get(),
            self.end.get(),
            self.period.get(),
            self.sp_entry.get(),
            self.sl_entry.get()
        ]
        self.content_list = [
            (self.chk1.cget("text") + ' ' + self.combo1.get() + ' ' + self.entry1.get()),
            (self.chk2.cget("text") + ' ' + self.combo2.get() + ' ' + self.entry2.get()),
            (self.chk3.cget("text") + ' ' + self.combo3.get() + ' ' + self.entry3.get()),
            (self.chk4.cget("text") + ' ' + self.combo4.get() + ' ' + self.entry4.get()),
            (self.chk5.cget("text") + ' ' + self.combo5.get() + ' ' + self.entry5.get()),
            (self.chk6.cget("text") + ' ' + self.combo6.get() + ' ' + self.entry6.get()),
            (self.chk7.cget("text") + ' ' + self.combo7.get() + ' ' + self.entry7.get()),
            (self.chk8.cget("text") + ' ' + self.combo8.get() + ' ' + self.entry8.get()),
            (self.chk9.cget("text") + ' ' + self.combo9.get() + ' ' + self.entry9.get()),
            (self.chk10.cget("text") + ' ' + self.combo10.get() + ' ' + self.entry10.get()),
            (self.chk11.cget("text") + ' ' + self.combo11.get() + ' ' + self.entry11.get()),
            None,
            None,
            None,
            None,
            None
        ]
        scpr.save_cache_sql((self.chk_list, self.chkvar_list, self.content_list, self.combo_list, self.entry_list))

    # 顯示執行項目
    def execute_func(self):
        self.save_cache()
        path = self.tplt_path_text.get()
        directory = self.path_text.get()
        scpr.save_path_sql(path)
        scpr.save_path_sql(directory, source="SS")
        self.update()
        self.after(1000)

        date = datetime.strptime(self.start.get(), "%Y-%m-%d")
        bool_list = self.chkvar_list[:11]
        entry_list = self.content_list[:11]
        # entry_list = list(filter(None, entry_list))
        select_func = scpr.SelectStock()
        result = select_func.mystrategy(date=date, exec=entry_list, bool=bool_list)
        self.selected_stock = list(result.index)

        self.update()
        self.after(1000)
        self.scrolltxt.insert(INSERT, "符合選擇條件的股票有: {}\n\n".format(self.selected_stock))

    # 回測功能
    def backtest_func(self):
        self.save_cache()
        self.update()
        self.after(1000)

        start = datetime.strptime(self.end.get(), "%Y-%m-%d")
        end = datetime.strptime(self.start.get(), "%Y-%m-%d")
        period = int(self.period.get())
        bool_list = self.chkvar_list[:11]
        exection = self.content_list[:11]
        # exection = list(filter(None, exection))

        if self.sp_chkvar.get():
            sp = float(self.sp_entry.get())
        else:
            sp = None
        if self.sl_chkvar.get():
            sl = float(self.sl_entry.get())
        else:
            sl = None

        backtest_func = scpr.SelectStock()
        profit, record, max, min, process = backtest_func.backtest(self, start, end, period, exection, bool_list,
                                                                   stop_loss=sl, stop_profit=sp)
        self.update()
        self.after(1000)

        for txt in process:
            self.scrolltxt.insert(INSERT, txt+"\n")
            self.update()
            self.after(1000)
        self.scrolltxt.insert(INSERT, '\n')
        self.scrolltxt.insert(INSERT, '每次換手最大報酬 : %.2f ％\n' % max)
        self.scrolltxt.insert(INSERT, '每次換手最少報酬 : %.2f ％\n\n' % min)
        self.scrolltxt.insert(INSERT, '交易利潤 :\n {}\n\n'.format(profit))
        self.scrolltxt.insert(INSERT, '交易紀錄 :\n {}\n\n'.format(record))
        self.scrolltxt.insert(INSERT, "完成\n")
    # 顯示作業進度
    def update_func(self):
        cache1 = scpr.get_cache_sql("市值")
        self.chkvar1.set(scpr.get_cache_sql("市值", bool=True))
        self.entry1.insert(0, cache1["entry"][0]) if self.chkvar1.get() else self.entry1.delete(0, "end")
        self.combo1.insert(0, cache1["combo"][0]) if self.chkvar1.get() else self.entry1.delete(0, "end")

        cache2 = scpr.get_cache_sql("三年自由現金流")
        self.chkvar2.set(scpr.get_cache_sql("三年自由現金流", bool=True))
        self.entry2.insert(0, cache2["entry"][0]) if self.chkvar2.get() else self.entry2.delete(0, "end")
        self.combo2.insert(0, cache2["combo"][0]) if self.chkvar2.get() else self.entry2.delete(0, "end")

        cache3 = scpr.get_cache_sql("股東權益報酬率")
        self.chkvar3.set(scpr.get_cache_sql("股東權益報酬率", bool=True))
        self.entry3.insert(0, cache3["entry"][0]) if self.chkvar3.get() else self.entry3.delete(0, "end")
        self.combo3.insert(0, cache3["combo"][0]) if self.chkvar3.get() else self.entry3.delete(0, "end")

        cache4 = scpr.get_cache_sql("營業利益年成長率")
        self.chkvar4.set(scpr.get_cache_sql("營業利益年成長率", bool=True))
        self.entry4.insert(0, cache4["entry"][0]) if self.chkvar4.get() else self.entry4.delete(0, "end")
        self.combo4.insert(0, cache4["combo"][0]) if self.chkvar4.get() else self.entry4.delete(0, "end")

        cache5 = scpr.get_cache_sql("八季營益率衰退")
        self.chkvar5.set(scpr.get_cache_sql("八季營益率衰退", bool=True))
        self.entry5.insert(0, cache5["entry"][0]) if self.chkvar5.get() else self.entry5.delete(0, "end")
        self.entry5.insert(0, cache5["combo"][0]) if self.chkvar5.get() else self.entry5.delete(0, "end")

        cache6 = scpr.get_cache_sql("市值營收比")
        self.chkvar6.set(scpr.get_cache_sql("市值營收比", bool=True))
        self.entry6.insert(0, cache6["entry"][0]) if self.chkvar6.get() else self.entry6.delete(0, "end")
        self.entry6.insert(0, cache6["combo"][0]) if self.chkvar6.get() else self.entry6.delete(0, "end")

        cache7 = scpr.get_cache_sql("短期營收年增")
        self.chkvar7.set(scpr.get_cache_sql("短期營收年增", bool=True))
        self.entry7.insert(0, cache7["entry"][0]) if self.chkvar7.get() else self.entry7.delete(0, "end")
        self.entry7.insert(0, cache7["combo"][0]) if self.chkvar7.get() else self.entry7.delete(0, "end")

        cache8 = scpr.get_cache_sql("短期營收年增2")
        self.chkvar8.set(scpr.get_cache_sql("短期營收年增2", bool=True))
        self.entry8.insert(0, cache8["entry"][0]) if self.chkvar8.get() else self.entry8.delete(0, "end")
        self.entry8.insert(0, cache8["combo"][0]) if self.chkvar8.get() else self.entry8.delete(0, "end")

        cache9 = scpr.get_cache_sql("短期淨利年增")
        self.chkvar9.set(scpr.get_cache_sql("短期淨利年增", bool=True))
        self.entry9.insert(0, cache9["entry"][0]) if self.chkvar9.get() else self.entry9.delete(0, "end")
        self.entry9.insert(0, cache9["combo"][0]) if self.chkvar9.get() else self.entry9.delete(0, "end")

        cache10 = scpr.get_cache_sql("存貨周轉變化率")
        self.chkvar10.set(scpr.get_cache_sql("存貨周轉變化率", bool=True))
        self.entry10.insert(0, cache10["entry"][0]) if self.chkvar10.get() else self.entry10.delete(0, "end")
        self.entry10.insert(0, cache10["combo"][0]) if self.chkvar10.get() else self.entry10.delete(0, "end")

        cache11 = scpr.get_cache_sql("rsv")
        self.chkvar11.set(scpr.get_cache_sql("rsv", bool=True))
        self.entry11.insert(0, cache11["entry"][0]) if self.chkvar11.get() else self.entry11.delete(0, "end")
        self.entry11.insert(0, cache11["combo"][0]) if self.chkvar11.get() else self.entry11.delete(0, "end")

        self.start.insert(0, scpr.get_cache_sql("選股日期:")["entry"][0])
        self.end.insert(0, scpr.get_cache_sql("回測起始日期:")["entry"][0])
        self.period.insert(0, scpr.get_cache_sql("週期天數:")["entry"][0])
        self.sp_chkvar.set(scpr.get_cache_sql("停利", bool=True))
        self.sl_chkvar.set(scpr.get_cache_sql("停損", bool=True))
        self.sp_entry.insert(0, scpr.get_cache_sql("停利")["entry"][0])
        self.sl_entry.insert(0, scpr.get_cache_sql("停損")["entry"][0])

    # 清除顯示
    def clear_func(self):
        self.scrolltxt.delete(1.0, "end")
        self.chkvar1.set(0)
        self.chkvar2.set(0)
        self.chkvar3.set(0)
        self.chkvar4.set(0)
        self.chkvar5.set(0)
        self.chkvar6.set(0)
        self.chkvar7.set(0)
        self.chkvar8.set(0)
        self.chkvar9.set(0)
        self.chkvar10.set(0)
        self.chkvar11.set(0)
        self.sp_chkvar.set(0)
        self.sl_chkvar.set(0)
        self.entry1.delete(0, "end")
        self.entry2.delete(0, "end")
        self.entry3.delete(0, "end")
        self.entry4.delete(0, "end")
        self.entry5.delete(0, "end")
        self.entry6.delete(0, "end")
        self.entry7.delete(0, "end")
        self.entry8.delete(0, "end")
        self.entry9.delete(0, "end")
        self.entry10.delete(0, "end")
        self.entry11.delete(0, "end")
        self.start.delete(0, "end")
        self.end.delete(0, "end")
        self.period.delete(0, "end")
        self.sp_entry.delete(0, "end")
        self.sl_entry.delete(0, "end")

    # 取得現有檔案的代號
    def getfileid(self):
        target_folder = self.path
        file = scpr.show_folder_content(target_folder, prefix="O_", postfix=".xlsx")
        index = []
        dictionary = {}
        for num in file[0:]:
            idx = ''.join([x for x in num if x.isdigit()])
            dictionary[idx] = num
            index.append(idx)
        return index, dictionary

    # 以樣板儲存新檔案
    def SaveExcel(self, ID, folder):
        path = self.tplt_path_text.get()
        wb = load_workbook(path)
        new_path = os.path.join(folder, "O_" + ID + "_財報分析.xlsx")
        wb.save(new_path)

    def save_excel(self):
        Exist = self.getfileid()[0]
        folder_path = os.path.join(self.path, "選股結果")
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        self.scrolltxt.insert(INSERT, "\n")
        for id in self.selected_stock:
            if id in Exist:
                self.scrolltxt.insert(INSERT, "{}已存在\n\n".format(id))
                self.update()
                self.after(1000)
            else:
                self.scrolltxt.insert(INSERT, "新增{}\n\n".format(id))
                self.SaveExcel(id, folder=folder_path)
                self.update()
                self.after(1000)

class Page_StockAnalysis(Frame):
    def __init__(self, master):
        Frame.__init__(self, master)
        Frame.configure(self, bg='pink')
        self.data_getter = scpr.TW_FinanceGet()
        self.prev_id = ""

        self.StockID_label = Label(self, text="分析股票代號: ", background="pink", font=("TkDefaultFont", 16))
        self.StockID_label.grid(row=0, column=0, columnspan=3, sticky=W)
        self.StockID_combo = ttk.Combobox(self, postcommand="", values=["2330", "0050"])
        self.StockID_combo.current(0)
        self.StockID_combo.grid(row=0, column=3, columnspan=3, sticky=W)

        self.back_btn = Button(self, text="Go back", command=lambda: master.switch_frame(StartPage))
        self.back_btn.grid(row=1, column=0, sticky=W)
        self.Mreport_btn = Button(self, text="月財報", command=lambda: [self.initial_data(),
                                                                     self.show_table(self.month_df),
                                                                     self.createWidget(self.month_fig)])
        self.Mreport_btn.grid(row=1, column=1, sticky=W)
        self.Sreport_btn = Button(self, text="季財報", command=lambda: [self.initial_data(),
                                                                     self.show_table(self.season_df),
                                                                     self.createWidget(self.season_fig, x=0, y=4, xs=5)
                                                                     ])
        self.Sreport_btn.grid(row=1, column=2, sticky=W)
        self.Cash_btn = Button(self, text="現金流", command=lambda: [self.initial_data(),
                                                                     self.show_table(self.cash_df)
                                                                     ])
        self.Cash_btn.grid(row=1, column=3, sticky=W)
        self.Price_btn = Button(self, text="價位分析", command=lambda: [self.initial_data(),
                                                                      self.show_table(self.est_price),
                                                                      self.createWidget(self.month_fig)
                                                                     ])
        self.Price_btn.grid(row=1, column=4, sticky=W)
        self.exit_btn = Button(self, text="Exit", command=self.quit)
        self.exit_btn.grid(row=1, column=5, sticky=W)

    def initial_data(self):
        id = self.StockID_combo.get()
        if self.prev_id != id:
            # 月財報
            month_setting = {
                "title": "股價/月營收年增",
                "main": ["股價"],
                "sub": ["月營收年增率3個月移動平均", "月營收年增率12個月移動平均"],
                "xlabel": ["日期"],
                "ylabel": ["價位", "增幅(%)"],
            }
            self.month_df = self.data_getter.Month_data(id)
            fig, setting = self.data_getter.module_data_to_draw(id, month_setting)
            self.month_fig = self.draw_figure(fig, setting)

            # 季財報
            self.season_df = self.data_getter.Season_data(id)
            self.season_fig = self.draw_figures()

            # 現金流
            self.cash_df = self.data_getter.Cash_data(id)

            # 預估股價
            self.est_price = self.data_getter.Price_estimation(id)

            # 記錄此次分析股票代號
            self.prev_id = self.StockID_combo.get()

    def createWidget(self, figure, x=7, y=2, xs=1, ys=1, s=W+E+N+S, tool=True):
        self.canvas = FigureCanvasTkAgg(figure, self)
        self.canvas.draw()
        self.canvas.get_tk_widget().grid(row=y, column=x, sticky=s, rowspan=ys, columnspan=xs)

        # 把matplotlib繪製圖形的導航工具欄顯示到tkinter視窗上
        if tool:
            toolbar = NavigationToolbar2Tk(self.canvas, self, pack_toolbar=False)
            toolbar.grid(row=y+1, column=x, sticky=W+E)
            # self.canvas._tkcanvas.grid(row=y-1, column=x, sticky=s)

    def draw_figure(self, df, setting):
        """建立繪圖物件"""

        # 設定中文顯示字型
        mpl.rcParams['font.sans-serif'] = ['Microsoft JhengHei']  # 中文顯示
        mpl.rcParams['axes.unicode_minus'] = False  # 負號顯示

        # 建立繪圖物件f figsize的單位是英寸 畫素 = 英寸*解析度
        figure = plt.figure(num=1, figsize=(10, 10), dpi=90, facecolor="pink", edgecolor='green', frameon=True)

        # 建立一副子圖
        fig, ax1 = plt.subplots(figsize=(4, 4), constrained_layout=False)  # 三個引數，依次是：行，列，當前索引
        plt.subplots_adjust(left=0.1, right=0.9, bottom=0.15, top=0.9)

        secondary_y = False
        x = df.index
        color_list = ["red", "green", "blue", "pink", "orange"]
        for col_name, color in zip(df.columns, color_list):
            axis = col_name.split("*")
            if axis[0] == "m":
                ax1.plot(x, df[col_name], color=color, label=axis[1])
            else:
                if not secondary_y:
                    ax2 = ax1.twinx()
                ax2.plot(x, df[col_name], color=color, label=axis[1], linestyle="--")
                secondary_y = True

        ax1.set_title(setting.get("title", ""), loc='center', pad=5, fontsize='large', color='red')  # 設定標題
        # 定義legend 重新定義了一次label
        line, label = ax1.get_legend_handles_labels()

        # ,fontsize='xx-large'
        ax1.set_xlabel(setting.get("xlabel", [""])[0])  # 確定座標軸標題
        ax1.xaxis.set_label_coords(0, -0.05)
        tick_spacing = x.size / 10  # x軸密集度
        ax1.xaxis.set_major_locator(mticker.MultipleLocator(tick_spacing))
        ax1.tick_params('x', labelrotation=70)

        ax1.set_ylabel(setting.get("ylabel", [""])[0], rotation=0)
        ax1.yaxis.set_label_coords(0, 1.02)

        if secondary_y:
            ax2.set_ylabel(setting.get("ylabel", ["", ""])[1], rotation=0)
            ax2.yaxis.set_label_coords(1, 1.05)
            line2, label2 = ax2.get_legend_handles_labels()
            line += line2
            label += label2

        ax1.grid(which='major', axis='x', color='gray', linestyle='-', linewidth=0.5, alpha=0.2)  # 設定網格
        ax1.invert_xaxis()
        ax1.legend(line, label, loc=0)

        return fig

    def draw_figures(self):
        """建立繪圖物件"""
        # 設定中文顯示字型
        mpl.rcParams['font.sans-serif'] = ['Microsoft JhengHei']  # 中文顯示
        mpl.rcParams['axes.unicode_minus'] = False  # 負號顯示

        figure = plt.figure(num=8, figsize=(5, 2), dpi=80, facecolor="gold", edgecolor='green', frameon=True)
        draw_df = self.season_df.set_index("項目").loc[[
            "營業利益率", "應收帳款週轉率", "存貨周轉率", "存貨營收比",
            "股東權益報酬率(年預估)", "稅後淨利率(累計)", "總資產週轉率(次/年)", "權益係數"]]

        nrows, ncols = 2, 4
        fig, axes = plt.subplots(nrows=nrows, ncols=ncols)
        count = 0
        for r in range(nrows):
            for c in range(ncols):
                df = draw_df.iloc[count]
                df.plot(ax=axes[r, c])
                axes[r, c].set_title(df.name, loc='center', color='red', pad=5)
                axes[r, c].invert_xaxis()
                count += 1

        # plt.subplots_adjust(left=None, bottom=None, right=None, top=None, wspace=None, hspace=None)
        # fig.subplots_adjust(left=None, bottom=None, right=None, top=None, wspace=None, hspace=None)

        return fig

    def show_table(self, df):
        def fixed_map(option):
            return [elm for elm in style.map('Treeview', query_opt=option) if
                    elm[:2] != ('!disabled', '!selected')]
        style = ttk.Style()
        style.map('Treeview', foreground=fixed_map('foreground'), background=fixed_map('background'))

        self.data_table = ttk.Treeview(self, columns=("Tags"), height=15)
        self.data_table.grid(row=2, column=0, columnspan=5, sticky=W+E)

        vsb = ttk.Scrollbar(self, orient="vertical", command=self.data_table.yview)
        vsb.grid(column=6, row=2, rowspan=2, sticky=N+S)
        hsb = ttk.Scrollbar(self, orient="horizontal", command=self.data_table.xview)
        hsb.grid(column=0, row=3, columnspan=5, sticky=W+E)
        self.data_table.configure(yscrollcommand=vsb.set)
        self.data_table.configure(xscrollcommand=hsb.set)

        self.data_table.column("#0", width=0, stretch=NO)
        self.data_table.heading("#0", text="", anchor=CENTER)

        df_col = df.columns.values
        df_row = df.index.values
        counter = len(df_col)
        self.data_table['columns'] = tuple(df_col)

        # 建立欄位名
        for n in range(counter):
            title = df_col[n]
            self.data_table.column(title, width=55, stretch=NO, anchor=CENTER)
            self.data_table.heading(title, text=title, anchor=CENTER)

        # 建立數值至表格中
        self.data_table.tag_configure('highlight', background='#DD99FF')
        for m in range(len(df_row)):
            value = tuple(df.iloc[m].replace(['NaN', 'nan', np.nan], "").tolist())
            if value[0][0] == "*":
                self.data_table.insert(parent='', index='end', text='', values=value, tag='highlight', open=False)
            else:
                self.data_table.insert(parent='', index='end', text='', values=value)

root = SockApp()

root.mainloop()
