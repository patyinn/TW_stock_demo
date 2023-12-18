import os
import re
import math
import json
import sqlite3
import datetime
import operator
import requests
import asyncio
import threading
import numpy as np
import pandas as pd

from io import StringIO
from dateutil.relativedelta import relativedelta
from matplotlib import pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side

from finlab.crawler_module import Crawler, CrawlerConnection
from finlab.data_module import RetrieveDataModule


class SystemProcessor:
    lock = threading.Lock()
    col_keys = ["path", "condition", "analysis"]

    def __init__(self, sys_path):
        self.sys_path = sys_path
        self._check_file_existed()

    def _check_file_existed(self):
        if not os.path.exists(os.path.dirname(self.sys_path)):
            os.makedirs(os.path.dirname(self.sys_path))
        if not os.path.exists(self.sys_path):
            with open(self.sys_path, "w") as f:
                f.write(json.dumps({c: {} for c in self.col_keys}, ensure_ascii=False, indent=4))

    def write_to_json(self, table_name, key, value):
        with self.lock:
            with open(self.sys_path, "r+", encoding="UTF-8") as f:
                origin = json.load(f)
                _is_dict = table_name.endswith("{}")
                table_name = table_name[:-2]
                origin.setdefault(table_name, {})
                if _is_dict:
                    origin[table_name].setdefault(key, {})
                    origin[table_name][key] = value
                else:
                    origin[table_name].setdefault(key, [])
                    if isinstance(value, list):
                        origin[table_name][key] = value
                    else:
                        origin[table_name][key].append(value)
                    origin[table_name][key] = list(set(origin[table_name][key]))[:10]

                f.seek(0)
                json.dump(origin, f, ensure_ascii=False, indent=4)
                f.truncate()

    def read_from_json(self, table_name, key):
        with self.lock:
            with open(self.sys_path, "r", encoding="UTF-8") as f:
                data = json.load(f)
                return data.get(table_name, {}).get(key, None)

    def del_from_json(self, table_name, key, value):
        with self.lock:
            with open(self.sys_path, "r+", encoding="UTF-8") as f:
                origin = json.load(f)
                if value in origin[table_name].get(key, []):
                    origin[table_name][key].remove(value)
                f.seek(0)
                json.dump(origin, f, ensure_ascii=False, indent=4)
                f.truncate()

    def save_path_sql(self, path, source="origin"):
        key, value = None, None
        if os.path.exists(path):
            if os.path.isdir(path):
                if source == "origin":
                    key = "directory"
                    value = path
                elif source == "select_stock":
                    key = "select_stock_directory"
                    value = path
            elif os.path.isfile(path):
                if path.endswith(".db"):
                    key = "db"
                    value = path
                elif path.endswith(".xlsx"):
                    key = "file"
                    value = path
            else:
                print("it's an invalid path")
        if key and value:
            self.write_to_json("path[]", key, value)

    def get_latest_path_sql(self, category):
        result = self.read_from_json("path", category)
        return result[-1] if result else ""

    def del_path_sql(self, category, path):
        self.del_from_json("path", category, path)

    def save_select_stock_condition_to_sql(self, combination):
        for com in zip(*combination):
            cond_dic = {
                "cond_name": com[0],
                "activate": com[1],
                "cond_content": com[2],
                "operator": com[3],
                "cond_value": com[4]
            }
            self.write_to_json("condition{}", cond_dic["cond_name"], cond_dic)

    def get_select_stock_condition_to_sql(self, condition):
        result = self.read_from_json("condition", condition)
        return result if result is not None else {
            "cond_name": condition,
            "activate": False,
            "cond_content": "",
            "operator": "",
            "cond_value": ""
        }

    # 把列出資料夾的程式碼寫成一個函式
    @classmethod
    def show_folder_content(cls, folder_path, prefix=None, postfix=None):
        files_list = []
        folder_content = os.listdir(folder_path)
        for item in folder_content:
            fullpath = os.path.join(folder_path, item)
            if os.path.isdir(fullpath):
                files_list += cls.show_folder_content(fullpath, prefix=prefix, postfix=postfix)
            elif os.path.isfile(fullpath):
                if prefix:
                    if item.startswith(prefix):
                        files_list.append(os.path.join(folder_path, item))
                elif postfix:
                    if item.endswith(postfix):
                        files_list.append(os.path.join(folder_path, item))
                else:
                    files_list.append(os.path.join(folder_path, item))
        return files_list


class CrawlerProcessor(Crawler):
    def __init__(self, conn, msg_queue):
        super().__init__(conn, msg_queue)
        self.conn = conn
        self.msg_queue = msg_queue

    async def exec_func(self, table, from_date, to_date, force=False, multi_threads=True):
        additional_arg = {}
        if table == "price":
            date = self.date_range(from_date, to_date)
            function = self.crawl_price
        elif table == "monthly_revenue":
            date = self.month_range(from_date, to_date)
            function = self.crawl_monthly_report
        elif table == "finance_statement":
            date = self.season_range(from_date, to_date)
            function = self.determine_crawl_finance_statement_func_by_date
            additional_arg = {
                "force": force,
                "base_directory": "",
            }
        if multi_threads:
            await self.update_table_multi_thread(table, function, date, **additional_arg)
        else:
            await self.update_table(table, function, date, **additional_arg)

    def date_func(self, table, pattern):
        if table == "finance_statement":
            table = "balance_sheet"
        if pattern == "from":
            latest_date = self.table_latest_date(table)
            date_list = latest_date
            if latest_date.date() < datetime.datetime.now().date() - datetime.timedelta(days=1):
                date_list += datetime.timedelta(days=1)
            elif latest_date.date() == datetime.datetime.now().date():
                date_list -= datetime.timedelta(days=1)
            date_list = date_list.strftime('%Y-%m-%d')
        else:
            date_list = datetime.datetime.now().strftime('%Y-%m-%d')
        return [date_list]


class SelectStock(RetrieveDataModule):
    def __init__(self, conn, msg_queue):
        super().__init__(conn, msg_queue)
        self.msg_queue = msg_queue

    async def my_strategy(self, date, cond_content, activate):
        share_capital = self.get_data(name='股本合計', n=1, start=date)
        price = self.get_data(name='收盤價', n=120, start=date)
        price_today = price[:share_capital.index[-1]].iloc[-1]
        capital_today = share_capital.iloc[-1]
        market_value = capital_today * price_today / 10 * 1000

        df1 = self.data_process(self.get_data(name='投資活動之淨現金流入（流出）', n=15, start=date))
        df2 = self.data_process(self.get_data(name='營業活動之淨現金流入（流出）', n=15, start=date))
        three_yrs_cash_flow = (df1 + df2).iloc[-12:].mean()

        net_profit_after_tax = self.get_data(name='本期淨利（淨損）', n=9, start=date)
        # 股東權益，有兩個名稱，有些公司叫做權益總計，有些叫做權益總額，所以得把它們抓出來
        total_stockholders_equity = self.get_data(name='權益總計', n=1, start=date)
        total_equity = self.get_data(name='權益總額', n=1, start=date)
        # 並且把它們合併起來
        total_stockholders_equity.fillna(total_equity, inplace=True)

        return_on_equity = ((net_profit_after_tax.iloc[-4:].sum()) / total_stockholders_equity.iloc[-1]) * 100

        operating_profit = self.get_data(name='營業利益（損失）', n=9, start=date)
        revenue_season = self.get_data(name='營業收入合計', n=9, start=date)
        operating_profit_margin = operating_profit / revenue_season
        prev_season_opm = operating_profit.shift(1) / revenue_season.shift(1)
        yr_growth_of_opm = (operating_profit_margin.iloc[-1] / operating_profit_margin.iloc[-5] - 1) * 100
        eight_seasons_opm = (operating_profit_margin / prev_season_opm - 1) * 100
        eight_seasons_opm = eight_seasons_opm.dropna(axis=1, how="all").dropna(how="all")

        revenue_this_month = self.get_data(name='當月營收', n=12, start=date) * 1000
        revenue_this_year = revenue_this_month.iloc[-12:].sum()
        market_value_per_revenue = market_value / revenue_this_year

        mr_year_growth = self.get_data(name='去年同月增減(%)', n=12, start=date)
        rolling_3_months_mr_year_growth = mr_year_growth.rolling(3).mean().reindex(index=mr_year_growth.index).iloc[-1]
        rolling_12_months_mr_year_growth = mr_year_growth.rolling(12).mean().reindex(index=mr_year_growth.index).iloc[-1]

        net_profit_margin_after_tax = net_profit_after_tax / revenue_season
        last_yr_net_profit_margin_after_tax = net_profit_margin_after_tax.shift(4)
        yr_growth_npm = (net_profit_margin_after_tax - last_yr_net_profit_margin_after_tax) / last_yr_net_profit_margin_after_tax * 100
        yr_growth_npm = yr_growth_npm
        short_term_yr_growth_npm = yr_growth_npm.iloc[-1]
        long_term_yr_growth_npm = yr_growth_npm[-4:].mean()

        inventory = self.get_data(name="存貨", n=3, start=date)
        operating_cost = self.get_data(name="營業成本合計", n=2, start=date)
        inventory_turnover = operating_cost.iloc[-1] / ((inventory.iloc[-1] + inventory.iloc[-2]) / 2) * 4
        prev_season_inventory_turnover = operating_cost.iloc[-2] / ((inventory.iloc[-2] + inventory.iloc[-3]) / 2) * 4
        inventory_turnover_ratio = (inventory_turnover - prev_season_inventory_turnover) / prev_season_inventory_turnover * 100

        rsv = (price.iloc[-1] - price.iloc[-60:].min()) / (price.iloc[-60:].max() - price.iloc[-60:].min())

        mapper = {
            "市值": market_value,
            "三年自由現金流": three_yrs_cash_flow,
            "股東權益報酬率": return_on_equity,
            "營業利益年成長率": yr_growth_of_opm,
            "八季營益率變化": eight_seasons_opm,
            "市值營收比": market_value_per_revenue,
            "短期營收年增": rolling_3_months_mr_year_growth,
            "短期營收年增2": rolling_3_months_mr_year_growth,
            "長期營收年增": rolling_12_months_mr_year_growth,
            "短期淨利年增": short_term_yr_growth_npm,
            "長期淨利年增": long_term_yr_growth_npm,
            "存貨周轉變化率": inventory_turnover_ratio,
            "rsv": rsv
        }
        ops = {
            "<": operator.lt,
            "<=": operator.le,
            ">": operator.gt,
            ">=": operator.ge,
            "=": operator.eq,
        }
        def _operator_func(var, op, con):
            a = mapper[var]
            if con in mapper:
                value = mapper[con]
            else:
                value = float(con)
            return ops[op](a, value)

        condition_list = []
        for b, e in zip(activate, cond_content):
            if len(e.split()) >= 3 and b is True:
                operators = _operator_func(*(e.split()))
                if isinstance(operators, pd.DataFrame):
                    operators = mapper[e.split()[0]][operators].isnull().sum() <= 0
                condition_list.append(operators)
        select_stock = condition_list[0]
        for cond in condition_list:
            select_stock = select_stock & cond
        return select_stock[select_stock]

    async def backtest(self, start_date, end_date, hold_days, cond_content, activate, weight='average', benchmark=None,
                 stop_loss=None,
                 stop_profit=None):
        # portfolio check
        if weight != 'average' and weight != 'price':
            print(f'Backtest stop, weight should be "average" or "price", find {weight} instead')
            self.msg_queue.put(f'Backtest stop, weight should be "average" or "price", find {weight} instead')

        # get price data in order backtest
        self.date = end_date
        price = self.get_data('收盤價', (end_date - start_date).days)
        # start from 1 TWD at start_date,
        end = 1

        # record some history
        equality = pd.Series()
        n_stock = {}
        comparison = []
        transactions = pd.DataFrame()
        max_return = -10000
        min_return = 10000

        if isinstance(hold_days, int):
            dates = self._date_iter_periodicity(start_date, end_date, hold_days)
        elif isinstance(hold_days, list):
            dates = self._date_iter_specify_dates(start_date, end_date, hold_days)
        else:
            print('the type of hold_dates should be list or int.')
            self.msg_queue.put('the type of hold_dates should be list or int.')
            return None

        figure, ax = plt.subplots(2, 1, sharex=True, sharey=False)

        keep_list = []
        keep_idx = pd.Index(keep_list)
        for sdate, edate in dates:
            # select stocks at date
            self.date = sdate
            stocks = await self.my_strategy(sdate, cond_content, activate)

            idx = stocks.index.append([keep_idx]).drop_duplicates()
            print(f"回測的股票為: {idx.tolist()}")
            self.msg_queue.put(f"回測的股票為: {idx.tolist()}")
            selected_columns = price[idx.tolist()]

            result = selected_columns[sdate:edate].iloc[1:]
            s = price[idx][sdate:edate].iloc[1:]

            if s.empty:
                s = pd.Series(1, index=pd.date_range(sdate + datetime.timedelta(days=1), edate))
            else:
                if stop_loss is not None:
                    below_stop = ((s / s.bfill().iloc[0]) - 1) * 100 < -np.abs(stop_loss)
                    below_stop = (below_stop.cumsum() > 0).shift(2).fillna(False)
                    s[below_stop] = np.nan
                if stop_profit is not None:
                    above_stop = ((s / s.bfill().iloc[0]) - 1) * 100 > np.abs(stop_profit)
                    above_stop = (above_stop.cumsum() > 0).shift(2).fillna(False)
                    s[above_stop] = np.nan

                s.dropna(axis=1, how='all', inplace=True)
                keep_list = s.dropna(axis=1)
                keep_idx = pd.Index(keep_list.columns)

                # record transactions
                buy_price = s.bfill().iloc[0]
                sell_price = s.apply(lambda s: s.dropna().iloc[-1])
                append_tran = pd.DataFrame({
                    'buy_price': buy_price,
                    'sell_price': sell_price,
                    'lowest_price': s.min(),
                    'highest_price': s.max(),
                    'buy_date': pd.Series(s.index[0], index=s.columns),
                    'sell_date': s.apply(lambda s: s.dropna().index[-1]),
                    'profit(%)': (sell_price / buy_price - 1) * 100
                })
                transactions = pd.concat([transactions, append_tran]).sort_index(ascending=True)

                s.ffill(inplace=True)
                s = s.sum(axis=1)

                if weight == 'average':
                    s = s / s.bfill().iloc[0]
                else:
                    s = s / s.bfill()[0]

            # print some log
            start_time = sdate.strftime("%Y-%m-%d")
            end_time = edate.strftime("%Y-%m-%d")

            profit_str = "{} - {} 報酬率: {:.2f}% nstock {}".format(start_time, end_time,
                                                                    (s.iloc[-1] / s.iloc[0] * 100 - 100), len(idx))
            comparison.append(profit_str)
            benchmark1 = price['0050'][sdate:edate].iloc[1:]
            p0050_str = "{} - {} 的0050報酬率: {:.2f}% ".format(start_time, end_time,
                                                                (benchmark1.iloc[-1] / benchmark1.iloc[0] * 100 - 100))
            comparison.append(p0050_str)
            print(f"{profit_str}\n{p0050_str}")
            self.msg_queue.put(f"{profit_str}\n{p0050_str}")

            max_return = max(max_return, s.iloc[-1] / s.iloc[0] * 100 - 100)
            min_return = min(min_return, s.iloc[-1] / s.iloc[0] * 100 - 100)

            # plot backtest result
            ((s * end - 1) * 100).plot(ax=ax[0])
            equality = pd.concat([equality, s * end])
            end = (s / s[0] * end).iloc[-1]

            if math.isnan(end):
                end = 1

            # add nstock history
            n_stock[sdate] = len(stocks)

            print('每次換手最大報酬 : %.2f ％' % max_return)
            print('每次換手最少報酬 : %.2f ％' % min_return)
            self.msg_queue.put('每次換手最大報酬 : %.2f ％' % max_return)
            self.msg_queue.put('每次換手最少報酬 : %.2f ％' % min_return)

        if benchmark is None:
            benchmark = price['0050'][start_date:end_date].iloc[1:]

        # bechmark (thanks to Markk1227)
        ((benchmark / benchmark[0] - 1) * 100).plot(ax=ax[0], legend=True, color=(0.8, 0.8, 0.8), grid=True)

        ax[0].set_ylabel('Return On Investment (%)')
        ax[0].grid(linestyle='-.')

        ((benchmark / benchmark.cummax() - 1) * 100).plot(ax=ax[1], legend=True, color=(0.8, 0.8, 0.8))
        ((equality / equality.cummax() - 1) * 100).plot(ax=ax[1], legend=True)
        plt.ylabel('Dropdown (%)')
        plt.grid(linestyle='-.')

        # pd.Series(nstock).plot.bar(ax=ax[2])
        # plt.ylabel('Number of stocks held')
        plt.show()

        return equality, transactions, max_return, min_return, comparison

    @staticmethod
    def _date_iter_periodicity(start_date, end_date, hold_days):
        date = start_date
        while date < end_date:
            yield date, (date + datetime.timedelta(hold_days))
            date += datetime.timedelta(hold_days)

    @staticmethod
    def _date_iter_specify_dates(start_date, end_date, hold_days):
        date_list = [start_date] + hold_days + [end_date]
        if date_list[0] == date_list[1]:
            date_list = date_list[1:]
        if date_list[-1] == date_list[-2]:
            date_list = date_list[:-1]
        for s_date, e_date in zip(date_list, date_list[1:]):
            yield s_date, e_date


class TWStockRetrieveModule(RetrieveDataModule):
    price = pd.DataFrame()
    month_df = pd.DataFrame()
    season_df = pd.DataFrame()
    cash_df = pd.DataFrame()
    estimation_df = pd.DataFrame()
    per_df = pd.DataFrame()
    mapper = {}
    db_path = None

    @classmethod
    def retrieve_data_from_db(cls, stock_id, season_num=16):
        if not cls.db_path:
            raise ValueError("assign a db path first.")
        conn = sqlite3.connect(cls.db_path)
        retriever = RetrieveDataModule(conn)

        mapper, dfs = retriever.get_bundle_data(['收盤價'], season_num*65, stock_id)
        cls.price = pd.concat([cls.price, dfs[0]])

        target_cols = ['當月營收', '上月比較增減(%)', '去年同月增減(%)']
        mapper, dfs = retriever.get_bundle_data(target_cols, season_num*4, stock_id)
        cls.month_df = pd.concat([cls.month_df, cls.parse_month_df(dfs, cls.price.copy())])

        target_cols = [
            '營業收入合計',
            '營業利益（損失）',
            '營業毛利（毛損）',
            "股本合計",
            "繼續營業單位稅前淨利（淨損）",
            "本期淨利（淨損）",
            "營業成本合計",
            "應收帳款淨額",
            "存貨",
            "資產總計",
            "負債總計",
            "應付帳款",
            "無形資產",
            "折舊費用",
            '權益總計',
            '權益總額',
        ]
        mapper, dfs = retriever.get_bundle_data(target_cols, season_num, stock_id, assign_table={"折舊費用": "cash_flows"})
        cls.season_df = pd.concat([cls.season_df, cls.parse_season_df(dfs, cls.price.copy())])

        target_cols = [
            "投資活動之淨現金流入（流出）",
            "營業活動之淨現金流入（流出）",
            "籌資活動之淨現金流入（流出）",
            "期初現金及約當現金餘額",
            "期末現金及約當現金餘額"
        ]
        mapper, dfs = retriever.get_bundle_data(target_cols, int(season_num/4*10), stock_id)
        cls.cash_df = pd.concat([cls.cash_df, cls.parse_cash_df(dfs)])

        est_month_df = cls.month_df.loc[[stock_id], [('營收', '月營收(億)'), ('營收', '月營收年增率')]]
        est_month_df.columns = est_month_df.columns.droplevel()
        est_season_df = cls.season_df.loc[[stock_id], [('獲利能力', '每股稅後盈餘'), ('獲利能力', '稅後淨利率'), ('資產負債表', '股本合計')]]
        est_season_df.columns = est_season_df.columns.droplevel()
        est_price_df = cls.price.copy()
        est_df, per_df = cls.parse_price_estimation(est_month_df, est_season_df, est_price_df)
        cls.estimation_df = pd.concat([cls.estimation_df, est_df]).drop_duplicates(keep="last")
        cls.per_df = pd.concat([cls.per_df, per_df])

        cls.mapper = {
            "股價": cls.price["收盤價"],
            "月營收": cls.month_df[('營收', '月營收(億)')],
            "月營收月增率": cls.month_df[('營收', '月營收月增率')],
            "月營收年增率": cls.month_df[('營收', '月營收年增率')],
        }

    @classmethod
    def retrieve_month_data(cls, stock_id):
        if cls.month_df.empty or stock_id not in cls.month_df.index.get_level_values("stock_id"):
            try:
                cls.retrieve_data_from_db(stock_id)
            except Exception as e:
                print(e)
                return pd.DataFrame(index=["分類", "內容"])
        return cls.month_df.loc[stock_id].loc[::-1].T.rename_axis(["分類", "內容"])

    @classmethod
    def retrieve_season_data(cls, stock_id):
        if cls.season_df.empty or stock_id not in cls.season_df.index.get_level_values("stock_id"):
            try:
                cls.retrieve_data_from_db(stock_id)
            except Exception as e:
                print(e)
                return pd.DataFrame(index=["分類", "內容"])
        return cls.season_df.loc[stock_id].loc[::-1].T.rename_axis(["分類", "內容"])

    @classmethod
    def retrieve_cash_data(cls, stock_id):
        if cls.cash_df.empty or stock_id not in cls.cash_df.index.get_level_values("stock_id"):
            try:
                cls.retrieve_data_from_db(stock_id)
            except Exception as e:
                print(e)
                return pd.DataFrame(index=["分類", "內容"])
        return cls.cash_df.loc[stock_id].loc[::-1].T.rename_axis(["分類", "內容"])

    @classmethod
    def retrieve_price_estimation(cls, stock_id):
        if cls.estimation_df.empty or stock_id not in cls.estimation_df.index.get_level_values("stock_id"):
            try:
                cls.retrieve_data_from_db(stock_id)
            except Exception as e:
                print(e)
                return pd.DataFrame(index=["分類", "內容"])
        return cls.estimation_df.loc[stock_id].loc[::-1].T.rename_axis(["分類", "內容"])

    @classmethod
    def parse_month_df(cls, dfs, price_df):
        # 輸入數字並存在變數中，可以透過該變數(字串)，呼叫特定股票
        df = pd.concat(dfs, axis=1)
        df['當月營收'] = df['當月營收'].multiply(0.00001)  # 單位: 億

        price_df = price_df.groupby(['stock_id', pd.Grouper(level='date', freq='M')]).aggregate(['min', 'mean', 'max'])
        price_df.columns = price_df.columns.get_level_values(1)
        price_df = price_df.rename(columns={
            "min": "最低股價",
            "mean": "平均股價",
            "max": "最高股價"
        })
        price_df.index = pd.MultiIndex.from_tuples(price_df.index, names=['stock_id', 'date'])
        price_df.index = price_df.index.map(lambda s: (s[0], pd.to_datetime(s[1].strftime("%Y-%m-10"), format="%Y-%m-%d")))

        # 計算3個月以及12個月的移動平均數
        df["3個月移動平均年增率"] = df.groupby('stock_id')["去年同月增減(%)"].transform(lambda s: s.rolling(3).mean())
        df["12個月移動平均年增率"] = df.groupby('stock_id')["去年同月增減(%)"].transform(lambda s: s.rolling(12).mean())

        df = df.reset_index().set_index(["stock_id", "date"])

        df = pd.concat([df, price_df], join="inner", axis=1).sort_index().round(2)
        df.index = df.index.map(lambda s: (s[0], s[1].strftime("%b-%y")))
        df = df[
            ["當月營收", "上月比較增減(%)", "去年同月增減(%)", "3個月移動平均年增率", "12個月移動平均年增率", "最低股價", "平均股價", "最高股價"]
        ]
        rename_cols = [
            ('營收', '月營收(億)'),
            ('營收', '月營收月增率'),
            ('營收', '月營收年增率'),

            ('營收變化', '3個月移動平均年增率'),
            ('營收變化', '12個月移動平均年增率'),

            ('股價', '最低股價'),
            ('股價', '平均股價'),
            ('股價', '最高股價'),
        ]
        df.columns = pd.MultiIndex.from_tuples(rename_cols)
        return df

    @classmethod
    def parse_season_df(cls, dfs, price_df):
        df = pd.concat(dfs, axis=1)

        df = df.multiply(0.00001)  # 單位: 億
        df["權益總計"].fillna(df["權益總額"], inplace=True)
        df.drop("權益總額", axis=1, inplace=True)

        price_df.index = price_df.index.map(lambda s: (s[0], cls._convert_price_to_season_date(s[1])))
        season_report_price = price_df.groupby(["stock_id", "date"])["收盤價"].last()
        '''        拆解數據處理        '''
        df["遞減折舊費用"] = df.groupby("stock_id")["折舊費用"].transform(lambda s: cls.data_process_multi_index(s, cum=False))
        '''        累積數據處理        '''
        df["累積股東權益報酬率(季)"] = df["本期淨利（淨損）"] / df["權益總計"] * 100
        df["累積股東權益報酬率(季)"] = df.groupby("stock_id")["累積股東權益報酬率(季)"].transform(lambda s: cls.data_process_multi_index(s, cum=True))
        df["累積季營收"] = df.groupby("stock_id")['營業收入合計'].transform(lambda s: cls.data_process_multi_index(s, cum=True))
        df["累積稅後淨利"] = df.groupby("stock_id")["本期淨利（淨損）"].transform(lambda s: cls.data_process_multi_index(s, cum=True))
        df["累積稅後淨利率"] = df["累積稅後淨利"] / df["累積季營收"] * 100
        df["累積營收淨值比"] = (df["本期淨利（淨損）"] / df["累積季營收"]) * 100
        df["累積股東權益資產轉換率"] = (df["權益總計"] / df["資產總計"]) * 100
        df["累積資產變化"] = (df["資產總計"] + df["資產總計"].shift(1)) / 2
        df["累積資產變化"] = df.groupby("stock_id")["累積資產變化"].transform(lambda s: cls.data_process_multi_index(s, cum=True))
        df["總資產週轉率(次/年)"] = df["累積季營收"] / df["累積資產變化"] * 4

        '''        處理需要放到excel的資料        '''
        df["季營收年增率"] = df.groupby("stock_id")["營業收入合計"].transform(lambda s: 100 * (s / s.shift(4)) - 100)
        df["營業毛利率"] = 100 * (df["營業毛利（毛損）"] / df["營業收入合計"])
        df["營業利益率"] = 100 * (df["營業利益（損失）"] / df["營業收入合計"])
        df["營業利益成長率"] = 100 * (df["營業利益率"] / df["營業利益率"].shift(1)) - 100
        df["股本季增率"] = df.groupby("stock_id")["股本合計"].transform(lambda s: 100 * (s / s.shift(1)) - 100)
        df["市值"] = season_report_price * df["股本合計"] / 10  # 市值 = 股價 * 總股數 (股本合計單位為 k元)
        df["營收市值比"] = df.groupby("stock_id")["營業收入合計"].transform(lambda s: s.rolling(4).sum())
        df["營收市值比"] = df["營收市值比"] / df["市值"] * 100
        df["稅前淨利率"] = 100 * (df["繼續營業單位稅前淨利（淨損）"] / df["營業收入合計"])
        df["本業收入比率"] = 100 * (df["營業利益（損失）"] / df["繼續營業單位稅前淨利（淨損）"])
        df["稅後淨利率"] = 100 * (df["本期淨利（淨損）"] / df["營業收入合計"])
        df["稅後淨利年增率"] = 100 * (df["稅後淨利率"] / df["稅後淨利率"].shift(4)) - 100
        df["每股稅後盈餘"] = df["本期淨利（淨損）"] / (df["股本合計"] / 10)
        df["每股稅後盈餘年成長率"] = df.groupby("stock_id")["每股稅後盈餘"].transform(lambda s: 100 * (s / s.shift(4)) - 100)
        df["應收帳款週轉率"] = df.groupby("stock_id")["應收帳款淨額"].transform(lambda s: (s + s.shift(1)) / 2)
        df["應收帳款週轉率"] = df["營業收入合計"] / df["應收帳款週轉率"] * 4
        df["存貨"] = df.groupby("stock_id")["存貨"].transform(lambda s: (s + s.shift(1)) / 2)
        df["存貨周轉率"] = df["營業成本合計"] / df["存貨"] * 4
        df["存貨占營收比"] = 100 * (df["存貨"] / df["營業收入合計"])
        df["折舊負擔比率"] = df["遞減折舊費用"] / df["營業收入合計"]
        df["供應商應付帳款總資產占比"] = 100 * (df["應付帳款"] / df["資產總計"])
        df["負債總資產占比"] = 100 * (df["負債總計"] / df["資產總計"])
        df["無形資產占比"] = 100 * (df["無形資產"] / df["資產總計"])
        df["股東權益報酬率(年預估)"] = df.loc[:, ["累積股東權益報酬率(季)"]].apply(cls._estimate_roe, axis=1)
        df["權益係數"] = 100 / df["累積股東權益資產轉換率"]

        df = df.rename(columns={"營業收入合計": "當季營收"})

        remain_cols = [
            ('獲利能力', '當季營收'),
            ('獲利能力', '季營收年增率'),
            ('獲利能力', '市值'),
            ('獲利能力', '營業毛利率'),
            ('獲利能力', '營業利益率'),
            ('獲利能力', '營業利益成長率'),
            ('獲利能力', '稅前淨利率'),
            ('獲利能力', '本業收入比率'),
            ('獲利能力', '稅後淨利率'),
            ('獲利能力', '稅後淨利年增率'),
            ('獲利能力', '每股稅後盈餘'),
            ('獲利能力', '每股稅後盈餘年成長率'),

            ('經營能力', '應收帳款週轉率'),
            ('經營能力', '存貨周轉率'),
            ('經營能力', '存貨占營收比'),
            ('經營能力', '營收市值比'),

            ('資產負債表', '股本合計'),
            ('資產負債表', '股本季增率'),
            ('資產負債表', '供應商應付帳款總資產占比'),
            ('資產負債表', '負債總資產占比'),
            ('資產負債表', '無形資產占比'),

            ('現金流量表', '遞減折舊費用'),
            ('現金流量表', '折舊負擔比率'),

            ('杜邦分析(累季)', '累積股東權益報酬率(季)'),
            ('杜邦分析(累季)', '股東權益報酬率(年預估)'),
            ('杜邦分析(累季)', '累積稅後淨利率'),
            ('杜邦分析(累季)', '總資產週轉率(次/年)'),
            ('杜邦分析(累季)', '權益係數'),
            ('杜邦分析(累季)', '累積股東權益資產轉換率')
        ]

        df.index = df.index.map(lambda s: (s[0], cls.report_season_determination(s[1])))
        df = df[[col for category, col in remain_cols]].round(2)
        df.columns = pd.MultiIndex.from_tuples(remain_cols)
        return df

    @classmethod
    def parse_cash_df(cls, dfs):
        df = pd.concat(dfs, axis=1).multiply(0.00001)  # 單位:億
        df = df.groupby("stock_id").apply(cls._pick_cash_flow)
        df["自由現金流量"] = df["投資活動之淨現金流入（流出）"] + df["營業活動之淨現金流入（流出）"]
        df = df[
            ["投資活動之淨現金流入（流出）", "營業活動之淨現金流入（流出）", "籌資活動之淨現金流入（流出）", "自由現金流量", "期初現金及約當現金餘額", "期末現金及約當現金餘額"]
        ].round(3)

        rename_cols = [
            ('金流', '理財活動現金'),
            ('金流', '營業活動現金'),
            ('金流', '籌資活動現金'),
            ('金流', '自由現金流量'),

            ('統計', '期初現金及約當現金餘額'),
            ('統計', '期末現金及約當現金餘額'),
        ]
        df.columns = pd.MultiIndex.from_tuples(rename_cols)
        return df

    @classmethod
    def parse_per_df(cls, season_df, price_df):
        price_df.index = price_df.index.map(lambda s: (s[0], s[1].to_period('Q').strftime("%YQ%q")))
        price_df = price_df.groupby(['stock_id', 'date']).last()
        if season_df.index[-1] != price_df.index[-1]:
            add_row = season_df.iloc[-1:].copy()
            add_row.index = pd.MultiIndex.from_tuples([price_df.index[-1]], names=season_df.index.names)
            season_df = pd.concat([season_df, add_row])
        price_df = price_df[price_df.index.isin(season_df.index)]
        season_df["每股稅後盈餘四季總和"] = season_df.groupby("stock_id")["每股稅後盈餘"].transform(
            lambda s: s.rolling(4).sum())
        per_df = (price_df["收盤價"] / season_df["每股稅後盈餘四季總和"]).to_frame("本益比").dropna()
        return per_df, per_df.groupby("stock_id")["本益比"].aggregate(['min', 'mean', 'max'])

    @classmethod
    def parse_price_estimation(cls, month_df, season_df, price_df):
        per_df, agg_per_df = cls.parse_per_df(season_df, price_df)

        est_df = pd.DataFrame()
        all_ids = month_df.index.get_level_values("stock_id").tolist()
        for stock_id in list(set(all_ids)):
            df = pd.DataFrame(dtype=float, index=(["月營收", "營收年增", "既有營收"]))

            ms_df = month_df.loc[stock_id]
            df[(stock_id, "短期")] = [ms_df.iloc[:3]["月營收(億)"].sum(), ms_df.iloc[:3]["月營收年增率"].mean(), ms_df.iloc[:9]["月營收(億)"].sum()]
            df[(stock_id, "中期")] = [ms_df.iloc[:6]["月營收(億)"].sum(), ms_df.iloc[:6]["月營收年增率"].mean(), ms_df.iloc[:6]["月營收(億)"].sum()]
            df[(stock_id, "長期")] = [ms_df.iloc[:12]["月營收(億)"].sum(), ms_df.iloc[:12]["月營收年增率"].mean(), ms_df.iloc[1]["月營收(億)"].sum()]

            avg_4s_pat = season_df.loc[stock_id, "稅後淨利率"].iloc[-4:].mean() / 100

            df = pd.concat(
                [
                    df,
                    (df.loc[["月營收"]] * (1 + df.loc["營收年增"].max() / 100) * avg_4s_pat).rename(index={"月營收": "樂觀推估稅後淨利"}),
                    (df.loc[["月營收"]] * (1 + df.loc["營收年增"].min() / 100) * avg_4s_pat).rename(index={"月營收": "悲觀推估稅後淨利"}),
                    (df.loc[["既有營收"]] * avg_4s_pat).rename(index={"既有營收": "既有稅後淨利"}),
                ],
            )
            equity = season_df.loc[stock_id, "股本合計"].iloc[-1]

            optimistic_eps = df.loc[["樂觀推估稅後淨利", "既有稅後淨利"]].sum() / equity * 10
            pessimistic_eps = df.loc[["悲觀推估稅後淨利", "既有稅後淨利"]].sum() / equity * 10

            df = pd.concat([
                df,
                optimistic_eps.to_frame(name="樂觀推估EPS").T,
                pessimistic_eps.to_frame(name="悲觀推估EPS").T,
            ])

            price_df = pd.DataFrame({
                "樂觀推估價位": optimistic_eps * agg_per_df.loc[stock_id, "mean"],
                "極端樂觀推估價位": optimistic_eps * agg_per_df.loc[stock_id, "max"],
                "悲觀推估價位": pessimistic_eps * agg_per_df.loc[stock_id, "mean"],
                "極端悲觀推估價位": pessimistic_eps * agg_per_df.loc[stock_id, "min"],
            }).rename(index={"樂觀推估價位": "樂觀推估價位", "極端樂觀推估價位": "極端樂觀推估價位",
                             "悲觀推估價位": "悲觀推估價位", "極端悲觀推估價位": "極端悲觀推估價位"})

            df = pd.concat([df, price_df.T])
            est_df = pd.concat([df, est_df], axis=1)

        est_df = est_df.T.round(2)
        est_df.index = pd.MultiIndex.from_tuples(est_df.index, names=['stock_id', '時間長度'])

        est_df = est_df[
            [
                "樂觀推估稅後淨利", "悲觀推估稅後淨利", "樂觀推估EPS", "悲觀推估EPS",
                "極端樂觀推估價位", "樂觀推估價位", "悲觀推估價位", "極端悲觀推估價位",
            ]
        ]
        pattern = re.compile(r"(\w+觀)(\w+)")
        est_df.columns = pd.MultiIndex.from_tuples(
            [(pattern.match(col).group(2), pattern.match(col).group(1)) for col in est_df.columns]
        )
        return est_df, per_df

    @classmethod
    def prepare_df_to_draw(cls, stock_id, setting):
        if not cls.mapper or stock_id not in cls.month_df.index.get_level_values("stock_id"):
            try:
                cls.retrieve_data_from_db(stock_id)
            except Exception as e:
                print(e)
                return pd.DataFrame()

        df_list = []
        for m in setting.get("main"):
            df = cls.mapper.get(m).loc[stock_id]
            if m == "股價":
                df.index = df.index.strftime("%b-%y")
                df = df.groupby(df.index).mean()
            df_list.append(df.rename(f"m*{m}"))

        for s in setting.get("sub", []):
            is_ma = re.match(r"([\u4e00-\u9fa5]+)(\d+)\w+移動平均", s)
            if is_ma:
                col_name = is_ma.group(1)
                is_month = int(is_ma.group(2))
            else:
                col_name = s
                is_month = None
            df = cls.mapper.get(col_name).loc[stock_id]
            if is_month:
                df = df.rolling(is_month).mean().reindex(index=df.index)
            df_list.append(df.rename(f"s*{s}"))

        final = pd.concat(df_list, join="inner", axis=1).rename_axis(index="日期").round(2)
        final.sort_index(inplace=True, ascending=False, key=lambda s: pd.to_datetime(s, format="%b-%y"))
        return final, setting

    @staticmethod
    def _pick_cash_flow(raw_data):
        date_idx = raw_data.index.get_level_values(1)
        q4_data = raw_data[date_idx.month == 3]
        df_data = pd.concat([q4_data, raw_data.iloc[-1:]]).drop_duplicates()
        df_data.index = [
            datetime.datetime(idxs[1].year - 1, idxs[1].month, 1)
            if idxs[1].month == 3 else idxs[1]
            for idxs in df_data.index
        ]
        df_data.index = df_data.index.strftime("%Y")
        return df_data

    @staticmethod
    def _estimate_roe(c_roe):
        if c_roe.name[1].month == 5:
            return c_roe * 4
        elif c_roe.name[1].month == 8:
            return c_roe * 2
        elif c_roe.name[1].month == 11:
            return c_roe * 4 / 3
        return c_roe

    @staticmethod
    def report_season_determination(date):
        year = date.year
        if date.month <= 3:
            season = 4
            year = year - 1
        elif date.month <= 5:
            season = 1
        elif date.month <= 8:
            season = 2
        elif date.month <= 11:
            season = 3
        elif date.month <= 12:
            season = 4
        else:
            print("Wrong month to determine")
        return f"{year}Q{season}"

    @staticmethod
    def _convert_price_to_season_date(date):
        q = date.to_period('Q')
        y = q.year
        if q.quarter == 1:
            date_str = '-05-15'
        elif q.quarter == 2:
            date_str = '-08-14'
        elif q.quarter == 3:
            date_str = '-11-14'
        elif q.quarter == 4:
            y += 1
            date_str = '-03-31'
        return pd.to_datetime(f"{y}{date_str}", format="%Y-%m-%d")


class FinancialAnalysis(TWStockRetrieveModule, CrawlerConnection):
    def __init__(self, db_path, msg_queue, file_path):
        self.file_path = file_path
        self.wb = load_workbook(self.file_path)
        self.ws0 = self.wb["月財報"]
        self.ws1 = self.wb["季財報"]
        self.ws2 = self.wb["現金流量"]
        self.ws3 = self.wb["進出場參考"]
        self.ws4 = self.wb["合理價推估"]
        self.msg_queue = msg_queue
        conn = sqlite3.connect(db_path)
        super().__init__(conn, msg_queue)

    @staticmethod
    def _warning_func(use_cond, sheet=None, rows=None, cols=None, threat=None):
        if use_cond:
            if threat:
                sheet.cell(row=rows, column=cols).font = Font(color='FF0000', bold=True)  # 紅色
                sheet.cell(row=rows, column=cols).fill = PatternFill(fill_type="solid", fgColor="FFFFBB")
                side_style = Side(style="thin", color="FF0000")
                sheet.cell(row=rows, column=cols).border = Border(left=side_style, right=side_style, top=side_style,
                                                                  bottom=side_style)
                sheet.cell(row=rows, column=1).fill = PatternFill(fill_type="solid", fgColor="AA0000")  # 深紅色
            else:
                sheet.cell(row=rows, column=cols).font = Font(color='FF0000', bold=False)  # 紅色
                sheet.cell(row=rows, column=cols).fill = PatternFill(fill_type="solid", fgColor="FFFFBB")
                sheet.cell(row=rows, column=1).fill = PatternFill(fill_type="solid", fgColor="FFAA33")  # 橘色
        else:
            sheet.cell(row=rows, column=cols).font = Font(color='000000')  # 黑色
            sheet.cell(row=rows, column=1).fill = PatternFill(fill_type="solid", fgColor="FFFFFF")  # 白色

    def _write_to_excel(self, data, round_num=None, sheet=None, rows=None, cols=None, string="", date=None):
        if round_num:
            data = round(data, round_num)
        sheet.cell(row=rows, column=cols).value = data
        sheet.cell(row=rows, column=cols).alignment = Alignment(horizontal="center", vertical="center",
                                                                wrap_text=True)
        if string:
            msg = f"新增{date}的{string}: {data}"
            self.msg_queue.put(msg)
            print(msg)

    async def update_monthly_report(self, stock_id, path=None):
        '''    從資料庫獲取月營收最新日期    '''
        revenue_month = self.get_data('當月營收', 2)

        '''    時間判斷    '''
        # 改成用資料庫的最新時間尤佳
        latest_date = revenue_month[stock_id].dropna().index[-1]
        latest_date_str = datetime.datetime.strftime(latest_date, '%Y-%m')
        table_month = datetime.datetime.strftime(self.ws0["A5"].value, '%Y-%m')

        if table_month == latest_date_str:
            self.msg_queue.put("No month data need to update.")
            print("No month data need to update.")
        else:
            year1, month1 = latest_date_str[:4],  latest_date_str[5:7]
            year2, month2 = table_month[:4], table_month[5:7]
            add_row_num = (int(year1) - int(year2)) * 12 + (int(month1) - int(month2))

            '''        根據相差月份取相對應數量的資料        '''
            add_revenue = add_row_num + 24
            target_cols = ['當月營收', '上月比較增減(%)', '去年同月增減(%)']
            mapper, dfs = self.get_bundle_data(target_cols, add_revenue, stock_id)
            mapper, price_dfs = self.get_bundle_data(['收盤價'], add_row_num*40, stock_id)
            df = self.parse_month_df(dfs, price_dfs[0])
            df.columns = df.columns.droplevel()
            df = df.loc[stock_id]

            target_pos = [
                (5, 2, "月營收(億)"),
                (5, 3, "月營收月增率"),
                (5, 4, "月營收年增率"),
                (5, 19, "3個月移動平均年增率"),
                (5, 20, "12個月移動平均年增率"),
                (5, 8, "最低股價"),
                (5, 7, "平均股價"),
                (5, 6, "最高股價"),
            ]

            add_row_num -= 1
            for add_row in range(add_row_num, -1, -1):
                self.ws0.insert_rows(5, amount=1)

                '''  新增月份  '''
                update_month = (latest_date - relativedelta(months=add_row))
                self._write_to_excel(
                    update_month, sheet=self.ws0, rows=5, cols=1, string="月份標籤", date=f"{update_month}"
                )
                self.ws0.cell(row=5, column=1).number_format = "mmm-yy"

                '''        更新營收        '''
                for pos in target_pos:
                    target_month = update_month.strftime("%b-%y")
                    value = df.loc[target_month, pos[2]]
                    self._write_to_excel(
                        value,
                        sheet=self.ws0,
                        rows=pos[0],
                        cols=pos[1],
                        string=pos[2],
                        date=f"{update_month}"
                    )

                    if pos[2] in ['月營收月增率', '月營收年增率']:
                        if value >= 0:
                            self.ws0.cell(row=pos[0], column=pos[1]).font = Font(color='FF0000')  # 紅色
                        else:
                            self.ws0.cell(row=pos[0], column=pos[1]).font = Font(color='00FF00')  # 綠色

            self.wb.save(path or self.file_path)
            print("完成更新 {} 的 月報".format(stock_id))
            self.msg_queue.put("完成更新 {} 的 月報".format(stock_id))

    async def update_directors_and_supervisors(self, stock_id, path=None):
        url = "https://goodinfo.tw/StockInfo/StockDirectorSharehold.asp?STOCK_ID={}".format(str(stock_id))
        r = await self.requests_get(url)

        dfs = pd.read_html(StringIO(r.text))
        df = pd.concat([df for df in dfs if df.shape[1] > 15 and df.shape[0] > 30])
        idx = pd.IndexSlice
        df = df.loc[idx[:], idx[["月別", "全體董監持股"], :]]
        df.columns = df.columns.get_level_values(1)
        df = df.set_index(["月別"])
        df.columns = df.columns.str.replace(' ', '')

        df["持股(%)"] = pd.to_numeric(df["持股(%)"], errors="coerce")
        df = df[~ df["持股(%)"].isnull()].dropna()["持股(%)"]

        df = df.rename(index=lambda s: s.replace("/", "-"))
        data, index = [], []
        for cells in list(self.ws0.columns):
            data.append(cells[9].value)
            index.append(cells[0].value)
        data = data[4:]
        index = index[4:]

        data_now = pd.DataFrame({'date': index, 'Data': data})
        data_now = data_now[data_now['date'].notnull()].rename(index=lambda s: s + 5)

        # 確認爬蟲到的最新資料是否與excel的資料時間點相同，沒有就刪除excel資料點
        while datetime.datetime.strftime(data_now['date'].iloc[0], "%Y-%m") != df.index[0]:
            data_now = data_now.drop(data_now.index[0])
        update_data = data_now[data_now['Data'].isnull()]

        pd.options.mode.chained_assignment = None

        for n in range(len(update_data)):
            date = update_data['date'].iloc[n]
            date_str = datetime.datetime.strftime(date, "%Y-%m")
            try:
                update_data['Data'].iloc[n] = df.loc[date_str]
                r = update_data.index[n]
                if self.ws0.cell(row=r, column=1).value == date:
                    self._write_to_excel(
                        update_data['Data'].iloc[n],
                        sheet=self.ws0,
                        rows=r,
                        cols=10,
                        string="股東占比",
                        date=date_str
                    )
            except Exception as e:
                print(f"Doesn't get {date_str} Data")
                self.msg_queue.put(f"Doesn't get {date_str} Data")

        self.wb.save(path or self.file_path)
        print("完成更新 {} 的 股東占比".format(stock_id))
        self.msg_queue.put("完成更新 {} 的 股東占比".format(stock_id))

    async def update_season_report(self, stock_id, path=None):
        '''    從資料庫獲取季報最新日期    '''
        revenue_season = self.get_data_assign_table('營業收入合計', 5)
        revenue_season = revenue_season[stock_id]

        '''    時間判斷    '''
        # 改成用資料庫的最新時間尤佳
        latest_date = revenue_season.dropna().index[-1]
        latest_date_str = self.report_season_determination(latest_date)
        table_month = self.ws1["E1"].value

        add_column_num = 4 * (int(latest_date_str[0:4]) - int(table_month[0:4])) + (
                int(latest_date_str[-1]) - int(table_month[-1]))

        if add_column_num <= 0:
            print("No data need to update.")
            self.msg_queue.put("No data need to update.")
        else:
            '''        根據相差月份取相對應數量的資料        '''
            get_data_num = add_column_num + 6
            target_cols = [
                '營業收入合計',
                '營業利益（損失）',
                '營業毛利（毛損）',
                "股本合計",
                "繼續營業單位稅前淨利（淨損）",
                "本期淨利（淨損）",
                "營業成本合計",
                "應收帳款淨額",
                "存貨",
                "資產總計",
                "負債總計",
                "應付帳款",
                "無形資產",
                "折舊費用",
                '權益總計',
                '權益總額',
            ]
            mapper, dfs = self.get_bundle_data(target_cols, get_data_num, stock_id, assign_table={"折舊費用": "cash_flows"})
            mapper, price_dfs = self.get_bundle_data(["收盤價"], add_column_num*65, stock_id)
            df = self.parse_season_df(dfs, price_dfs[0])
            df = df.loc[stock_id]
            df.columns = df.columns.droplevel()

            condition_df = pd.DataFrame()
            condition_df["季營收年增率"] = df["季營收年增率"] <= 0
            condition_df["營業利益成長率t"] = df["營業利益成長率"] < -30
            condition_df["營業利益成長率"] = df["營業利益成長率"].between(-30, -20)
            condition_df["營收市值比"] = df["營收市值比"] < 20
            condition_df["每股稅後盈餘年成長率"] = df["每股稅後盈餘年成長率"] < 0
            condition_df["負債總資產占比"] = df["負債總資產占比"] > 40
            condition_df["無形資產占比"] = df["無形資產占比"] > 10
            condition_df["無形資產占比t"] = df["無形資產占比"] > 30
            condition_df["折舊負擔比率"] = df["折舊負擔比率"] > df["營業利益率"]

            warning_on_excel = {
                "季營收年增率": [df["季營收年增率"] <= 0],
                "營業利益成長率": [df["營業利益成長率"] < -30, df["營業利益成長率"].between(-30, -20)],
                "營收市值比": [df["營收市值比"] < 20],
                "每股稅後盈餘年成長率": [df["每股稅後盈餘年成長率"] < 0],
                "負債總資產占比": [df["負債總資產占比"] > 40],
                "無形資產占比": [df["無形資產占比"] > 10, df["無形資產占比"] > 30],
                "折舊負擔比率": [df["折舊負擔比率"] > df["營業利益率"]],
            }
            write_df_to_excel = [
                (3, 5, "當季營收",),
                (4, 5, "季營收年增率",),
                (5, 5, "市值",),
                (6, 5, "營業毛利率",),
                (7, 5, "營業利益率",),
                (8, 5, "營業利益成長率",),
                (9, 5, "稅前淨利率",),
                (10, 5, "本業收入比率",),
                (11, 5, "稅後淨利率",),
                (12, 5, "稅後淨利年增率",),
                (13, 5, "每股稅後盈餘",),
                (14, 5, "每股稅後盈餘年成長率",),

                (16, 5, "應收帳款週轉率",),
                (17, 5, "存貨周轉率",),
                (18, 5, "存貨占營收比",),
                (19, 5, "營收市值比",),

                (21, 5, "股本合計",),
                (22, 5, "股本季增率",),
                (23, 5, "供應商應付帳款總資產占比",),
                (24, 5, "負債總資產占比",),
                (25, 5, "無形資產占比",),

                (27, 5, "遞減折舊費用",),
                (28, 5, "折舊負擔比率",),

                (30, 5, "累積股東權益報酬率(季)",),
                (31, 5, "股東權益報酬率(年預估)",),
                (32, 5, "累積稅後淨利率",),
                (33, 5, "總資產週轉率(次/年)",),
                (34, 5, "權益係數",),
                (35, 5, "累積股東權益資產轉換率",),
            ]

            add_column_num *= -1
            for add_row in range(add_column_num, 0, 1):
                self.ws1.insert_cols(5, amount=1)
                update_season = df["當季營收"].index[add_row]

                '''  新增季度標籤  '''
                self._write_to_excel(update_season, sheet=self.ws1, rows=1, cols=5, string="季度標籤",
                                     date=f"{update_season}")
                self.ws1.cell(row=1, column=5).fill = PatternFill(fill_type="solid", fgColor="DDDDDD")

                for data in write_df_to_excel:
                    self._write_to_excel(
                        df.loc[update_season, data[2]],
                        sheet=self.ws1,
                        round_num=2,
                        rows=data[0],
                        cols=data[1],
                        string=data[2],
                        date=f"{update_season}"
                    )
                    if warning_on_excel.get(data[2]):
                        warning_cond = warning_on_excel[data[2]]
                        for i, cond in enumerate(warning_cond):
                            if cond[update_season] and i == 0:
                                self._warning_func(
                                    True,
                                    sheet=self.ws1,
                                    rows=data[0],
                                    cols=data[1],
                                    threat=True
                                )
                                break
                            elif cond[update_season] and i == 1:
                                self._warning_func(
                                    True,
                                    sheet=self.ws1,
                                    rows=data[0],
                                    cols=data[1],
                                    threat=False
                                )

            self.wb.save(path or self.file_path)
            print("完成更新 {} 的 季報".format(stock_id))
            self.msg_queue.put("完成更新 {} 的 季報".format(stock_id))

    async def update_cash_flow(self, stock_id, path=None):
        '''    從資料庫獲取季報最新日期    '''
        cash_flow_for_investing = self.get_data_assign_table("投資活動之淨現金流入（流出）", 5)
        cash_flow_for_investing = cash_flow_for_investing[stock_id]

        '''    時間判斷    '''
        latest_date = cash_flow_for_investing.dropna().index[-1]
        if latest_date.month == 3:
            year = latest_date.year - 1
        else:
            year = latest_date.year
        table_year = self.ws2["D1"].value
        add_column_num = year - int(table_year)

        '''    確認當年資料是否需要更新    '''
        if self.ws2["D4"].value != cash_flow_for_investing[-1]:
            self.ws2.delete_cols(4, 1)
            add_column_num += 1
            print("當年度資料更新")
            self.msg_queue.put("當年度資料更新")

        if add_column_num <= 0:
            print("No data need to update.")
            self.msg_queue.put("No data need to update.")
        else:
            '''        根據相差月份取相對應數量的資料        '''
            get_data_num = add_column_num * 4
            target_cols = [
                "投資活動之淨現金流入（流出）",
                "營業活動之淨現金流入（流出）",
                "籌資活動之淨現金流入（流出）",
                "期初現金及約當現金餘額",
                "期末現金及約當現金餘額"
            ]
            mapper, dfs = self.get_bundle_data(target_cols, get_data_num, stock_id)

            df = self.parse_cash_df(dfs)
            df.columns = df.columns.droplevel()
            df = df.loc[stock_id]

            target_pos = [
                (4, 4, "理財活動現金"),
                (3, 4, "營業活動現金"),
                (6, 4, "籌資活動現金"),
                (7, 4, "期初現金及約當現金餘額"),
                (8, 4, "期末現金及約當現金餘額"),
                (5, 4, "自由現金流量"),
            ]
            add_column_num *= -1
            for add_row in range(add_column_num, 0, 1):
                self.ws2.insert_cols(4, amount=1)
                update_year = df.index[add_row]

                '''  新增年度標籤  '''
                self._write_to_excel(update_year, sheet=self.ws2, rows=1, cols=4, string="現金流量標籤", date=update_year)
                self.ws2.cell(row=1, column=4).fill = PatternFill(fill_type="solid", fgColor="DDDDDD")

                '''  新增營業活動現金、理財活動現金、自由現金流量、籌資活動現金 新增期初現金及約當現金餘額、期末現金及約當現金餘額'''
                for pos in target_pos:
                    value = df.loc[update_year, pos[2]]
                    self._write_to_excel(
                        value,
                        round_num=1,
                        sheet=self.ws2,
                        rows=pos[0],
                        cols=pos[1],
                        string=pos[2],
                        date=update_year
                    )
                    if pos[2] in ['營業活動現金', '自由現金流量']:
                        self._warning_func(
                            value < 0,
                            sheet=self.ws2,
                            rows=pos[0],
                            cols=pos[1],
                            threat=True
                        )

            self.wb.save(path or self.file_path)
            print("完成更新 {} 的 現金流量表".format(stock_id))
            self.msg_queue.put("完成更新 {} 的 現金流量表".format(stock_id))

    async def update_per(self, stock_id, path=None):
        '''    使用現在的時間當作最新的更新時間點    '''
        season_now = pd.Period(datetime.datetime.now(), freq="Q").strftime("%YQ%q")

        # 與table最新資料比對時間，決定需要增加的數據量
        table_month = self.ws4["A16"].value
        # 更新當下是第幾季度，從Q1到當下都是更新目標
        update_row_num = int(table_month[-1])
        diff_year, diff_season = (int(season_now[0:4]) - int(table_month[0:4])), (int(season_now[-1]) - int(update_row_num))
        add_row_num = 4 * diff_year + diff_season

        total_num = update_row_num + add_row_num

        if add_row_num <= 0:
            print("Update PER this year.")
            self.msg_queue.put("Update PER this year.")
        else:
            print("Increase PER this season and update PER this year.")
            self.msg_queue.put("Increase PER this season and update PER this year.")

        # 根據需要跟新以及新增的數量，去從sqlite3抓取相對應的數據量
        get_data_num = total_num + 4

        mapper, dfs = self.get_bundle_data(["股本合計", "本期淨利（淨損）"], get_data_num, stock_id)
        season_df = pd.concat(dfs, axis=1)
        season_df["每股稅後盈餘"] = season_df["本期淨利（淨損）"] / (season_df["股本合計"] / 10)
        season_df.index = season_df.index.map(lambda s: (s[0], self.report_season_determination(s[1])))
        mapper, price_dfs = self.get_bundle_data(['收盤價'], get_data_num * 65, stock_id)
        per_df, agg_per_df = self.parse_per_df(season_df, price_dfs[0])
        per_df = per_df.loc[stock_id]

        '''  檢查公布財報的EPS時間與實際時間的差別，如果尚未公布財報，則填入現在的時間，新增最新時間資料  '''
        per_df = per_df.iloc[-1*total_num:]

        if len(per_df) < total_num-update_row_num:
            update_row_num = 0
            total_num = add_row_num = len(per_df)
        elif len(per_df) < total_num:
            update_row_num = total_num - len(per_df)
            total_num = len(per_df)

        # 新增PER資料
        for add_row in range(-1*total_num, 0, 1):
            # total_num = update_row_num + add_row_num
            # 要插入的數量是 total_num - update_row_num
            if add_row < -1*add_row_num:
                row = 15 + -1*(add_row+add_row_num)
            else:
                self.ws4.insert_rows(16, amount=1)
                row = 16

            update_season = per_df.iloc[add_row].name
            '''  新增季度標籤  '''
            self._write_to_excel(
                update_season, sheet=self.ws4, rows=row, cols=1, string="PER季度標籤", date=f"{update_season}"
            )
            self.ws4.cell(row=row, column=1).fill = PatternFill(fill_type="solid", fgColor="FFEE99")
            '''  新增本益比  '''
            self._write_to_excel(per_df.loc[update_season, "本益比"], round_num=2, sheet=self.ws4, rows=row, cols=2, string="新增PER", date=update_season)

        self.wb.save(path or self.file_path)
        print("完成更新 {} 的 本益比".format(stock_id))
        self.msg_queue.put("完成更新 {} 的 本益比".format(stock_id))

    async def update_price_today(self, stock_id, path=None):
        total_cols = ['最高價', '最低價', '開盤價', '收盤價']
        excel_pos = [(12, 3), (13, 3), (12, 5), (13, 5)]

        mapper, dfs = self.get_bundle_data(total_cols, 1, stock_id)
        df = pd.concat(dfs, axis=1)
        df = df.loc[stock_id]
        date_str = df.index[0].strftime("%Y/%m/%d")

        self._write_to_excel(
            date_str, sheet=self.ws4, rows=13, cols=1, string=f"新增{date_str}"
        )
        for col, pos in zip(total_cols, excel_pos):
            self._write_to_excel(
                df[col].iloc[0], round_num=1, sheet=self.ws4, rows=pos[0], cols=pos[1], string=f"新增{col}", date=date_str
            )

        self.wb.save(path or self.file_path)
        print("完成更新 {} 的 價位".format(stock_id))
        self.msg_queue.put("完成更新 {} 的 價位".format(stock_id))
