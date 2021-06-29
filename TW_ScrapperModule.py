import pandas as pd
import datetime
import time
from dateutil.relativedelta import relativedelta
import os
from io import StringIO
import sqlite3
import requests

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Side

from finlab.data import Data
from finlab.crawler import (
    crawl_price,
    crawl_monthly_report,
    crawl_finance_statement_by_date,
    update_table,

    table_exist,
    table_latest_date,

    date_range, month_range, season_range
)

from matplotlib import pyplot as plt
import numpy as np
import math
import operator

# 檢查下載檔案機制有問題
conn = sqlite3.connect(os.path.join("data", "data.db"))
data = Data()

class TW_scrapper():
    def __init__(self, path=None):
        self.File_path = path
        self.wb = load_workbook(self.File_path)
        self.ws0 = self.wb["月財報"]
        self.ws1 = self.wb["季財報"]
        self.ws2 = self.wb["現金流量"]
        self.ws3 = self.wb["進出場參考"]
        self.ws4 = self.wb["合理價推估"]

    def Season_determination(self, date):
        year = date.year
        if date.month <= 3:
            season = 4
            year = year - 1
        elif date.month <= 5:
            season = 1
        elif date.month <= 8:
            season = 2
        elif date.month <= 11:
            season = 3
        elif date.month <= 12:
            season = 4
        else:
            print("Wrong month to determine")

        Q_str = str(year) + "Q" + str(season)

        return Q_str
    def Season2Month(self, str):

        season = int(str[-1])
        year = int(str[0:4])
        Months = 1
        day = 1
        if season == 4:
            Months = 3
            day = 31
            year += 1
        elif season == 3:
            Months = 11
            day = 14
        elif season == 2:
            Months = 8
            day = 14
        elif season == 1:
            Months = 5
            day = 15
        else:
            print("Wrong season")

        result = datetime.datetime(year, Months, day).strftime("%Y-%m")

        return result
    def months(self, str1, str2):
        year1 = datetime.datetime.strptime(str1[0:10], "%Y-%m").year
        year2 = datetime.datetime.strptime(str2[0:10], "%Y-%m").year
        month1 = datetime.datetime.strptime(str1[0:10], "%Y-%m").month
        month2 = datetime.datetime.strptime(str2[0:10], "%Y-%m").month
        num = (year1 - year2) * 12 + (month1 - month2)
        return num
    def deltaSeasons(self, date, delta):
        str1 = self.Season_determination(date)
        year = int(str1[0:4])
        s = int(str1[-1])

        season = s - delta
        while season <= 0:
            season += 4
            year -= 1
        while season >= 5:
            season -= 4
            year += 1

        if season == 4:
            month = 3
            day = 31
            year += 1
        elif season == 3:
            month = 11
            day = 14
        elif season == 2:
            month = 8
            day = 14
        elif season == 1:
            month = 5
            day = 15
        else:
            print("Wrong season")

        r = datetime.datetime(year, month, day)

        return r
    def DataProcess(self, df, cum=None):

            season4 = df[df.index.month == 3]
            season1 = df[df.index.month == 5]
            season2 = df[df.index.month == 8]
            season3 = df[df.index.month == 11]

            season1.index = season1.index.year
            season2.index = season2.index.year
            season3.index = season3.index.year
            season4.index = season4.index.year - 1

            if cum:
                newseason1 = season1
                newseason2 = season2 + newseason1.reindex_like(season2)
                newseason3 = season3 + newseason2.reindex_like(season3)
                newseason4 = season4 + newseason3.reindex_like(season4)
            else:
                newseason1 = season1
                newseason2 = season2 - season1.reindex_like(season2)
                newseason3 = season3 - season2.reindex_like(season3)
                newseason4 = season4 - season3.reindex_like(season4)

            newseason1.index = pd.to_datetime(newseason1.index.astype(str) + '-05-15')
            newseason2.index = pd.to_datetime(newseason2.index.astype(str) + '-08-14')
            newseason3.index = pd.to_datetime(newseason3.index.astype(str) + '-11-14')
            newseason4.index = pd.to_datetime((newseason4.index + 1).astype(str) + '-03-31')

            return newseason1.append(newseason2).append(newseason3).append(newseason4).sort_index()
    def WarningFunc(self, con, sheet=None, rows=None, cols=None, threat=None):
        if con:
            if threat:
                sheet.cell(row=rows, column=cols).font = Font(color='FF0000', bold=True)  # 紅色
                sheet.cell(row=rows, column=cols).fill = PatternFill(fill_type="solid", fgColor="FFFFBB")
                side_style = Side(style="thin", color="FF0000")
                sheet.cell(row=rows, column=cols).border = Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
                sheet.cell(row=rows, column=1).fill = PatternFill(fill_type="solid", fgColor="AA0000") # 深紅色
            else:
                sheet.cell(row=rows, column=cols).font = Font(color='FF0000', bold=False)  # 紅色
                sheet.cell(row=rows, column=cols).fill = PatternFill(fill_type="solid", fgColor="FFFFBB")
                sheet.cell(row=rows, column=1).fill = PatternFill(fill_type="solid", fgColor="FFAA33")  # 橘色
        else:
            sheet.cell(row=rows, column=cols).font = Font(color='000000')  # 黑色
            sheet.cell(row=rows, column=1).fill = PatternFill(fill_type="solid", fgColor="FFFFFF")  # 白色

    def Write2Excel(self, data, rounds=None, sheet=None, rows=None, cols=None, string=None, date=None):
        data = round(data, rounds)
        sheet.cell(row=rows, column=cols).value = data
        sheet.cell(row=rows, column=cols).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if string:
            print("新增", date, "的"+string+":", data)

    def CashFlowGet(self, rawData):

        rawData = rawData.fillna(0)
        idx = rawData.index[-1]
        # 抓當年度最新一筆資料
        rawData_1 = pd.Series(rawData[-1], index=[idx])
        # Q4
        if idx.month == 3:
            rawData_year = idx.year - 1
        else:
            rawData_year = idx.year

        # 抓每年的Q4
        newData = rawData[rawData.index.month == 3]
        newData.index = newData.index.year - 1
        newData.index = pd.to_datetime((newData.index).astype(str))

        if newData.empty:
            newData = rawData_1
        elif newData.index[-1].year != rawData_year:
            newData = pd.concat([newData, rawData_1], ignore_index=False)

        return newData
    def PNdetermination(self, data, sheet=None, rows=None, cols=None):
        if data >= 0:
            sheet.cell(row=rows, column=cols).font = Font(color='000000')  # 黑色
        else:
            sheet.cell(row=rows, column=cols).font = Font(color='FF0000')  # 紅色
        return {}
    def Update_Monthly_report(self, Stock_ID, path):

        '''    從資料庫獲取月營收最新日期    '''
        Revenue_Month = data.get('當月營收', 2)

        '''    時間判斷    '''
        # 改成用資料庫的最新時間尤佳
        latest_date = Revenue_Month[Stock_ID].dropna().index[-1]
        latest_date_str = datetime.datetime.strftime(latest_date, '%Y-%m')
        table_month = datetime.datetime.strftime(self.ws0["A5"].value, '%Y-%m')

        if table_month == latest_date_str:
            print("No data need to update.")
        else:
            add_row_num = self.months(latest_date_str, table_month)

            '''        根據相差月份取相對應數量的資料        '''
            add_revenue = add_row_num + 24
            Revenue_Month = data.get('當月營收', add_revenue) * 0.00001
            add_price = add_row_num * 40
            price = data.get('收盤價', add_price)
            MR_MonthGrowth = data.get('上月比較增減(%)', add_revenue)
            MR_YearGrowth = data.get('去年同月增減(%)', add_revenue)

            # 輸入數字並存在變數中，可以透過該變數(字串)，呼叫特定股票
            Month_Revenue = Revenue_Month[Stock_ID]
            price = price[Stock_ID]
            MR_MonthGrowth = MR_MonthGrowth[Stock_ID]
            MR_YearGrowth = MR_YearGrowth[Stock_ID]
            # print("****", Revenue_Month)

            MAG_3M = MR_YearGrowth.rolling(3).mean().reindex(index=MR_YearGrowth.index)
            MAG_3M = round(MAG_3M, 2)
            MAG_12M = MR_YearGrowth.rolling(12).mean().reindex(index=MR_YearGrowth.index)
            MAG_12M = round(MAG_12M, 2)

            add_row_num -= 1

            for add_row in range(add_row_num, -1, -1):

                self.ws0.insert_rows(5, amount=1)

                '''  新增月份  '''
                Update_Month = latest_date - relativedelta(months=add_row)
                self.ws0.cell(row=5, column=1).value = Update_Month
                self.ws0.cell(row=5, column=1).number_format = "mmm-yy"
                self.ws0.cell(row=5, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                print("新增:", self.ws0.cell(row=5, column=1).value)

                '''        更新營收        '''
                MR = round(Month_Revenue.loc[Update_Month], 2)
                self.ws0.cell(row=5, column=2).value = MR
                self.ws0.cell(row=5, column=2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                print("新增月份:", Update_Month, "的月營收:", MR)
                '''        更新月增率        '''
                MR_MG = round(MR_MonthGrowth.loc[Update_Month], 2)
                self.ws0.cell(row=5, column=3).value = MR_MG
                self.ws0.cell(row=5, column=3).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if MR_MG >= 0:
                    self.ws0.cell(row=5, column=3).font = Font(color='FF0000')  # 紅色
                else:
                    self.ws0.cell(row=5, column=3).font = Font(color='00FF00')  # 綠色
                print("新增", Update_Month, "的月增率:", MR_MG)
                '''        更新年增率        '''
                MR_YG = round(MR_YearGrowth.loc[Update_Month], 2)
                self.ws0.cell(row=5, column=4).value = MR_YG
                self.ws0.cell(row=5, column=4).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if MR_YG >= 0:
                    self.ws0.cell(row=5, column=4).font = Font(color='FF0000')  # 紅色
                else:
                    self.ws0.cell(row=5, column=4).font = Font(color='00FF00')  # 綠色
                print("新增", Update_Month, "的年增率:", round(MR_YG, 2))

                '''        更新當月最高、最低、平均收盤價        '''
                Update_Month_str = Update_Month.strftime('%Y-%m')
                self.ws0.cell(row=5, column=6).value = round(price.loc[Update_Month_str].max(), 2)
                self.ws0.cell(row=5, column=6).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                self.ws0.cell(row=5, column=7).value = round(price.loc[Update_Month_str].mean(), 2)
                self.ws0.cell(row=5, column=7).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                self.ws0.cell(row=5, column=8).value = round(price.loc[Update_Month_str].min(), 2)
                self.ws0.cell(row=5, column=8).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                '''        更新長、短期年增        '''
                self.ws0.cell(row=5, column=19).value = MAG_3M.loc[Update_Month]
                self.ws0.cell(row=5, column=19).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                self.ws0.cell(row=5, column=20).value = MAG_12M.loc[Update_Month]
                self.ws0.cell(row=5, column=20).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.wb.save(path)
        print("Month Report end")
    def Update_Directors_and_supervisors(self, Stock_ID, path):
        # 設定headers
        headers = {
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36'
        }

        url = "https://goodinfo.tw/StockInfo/StockDirectorSharehold.asp?STOCK_ID=" + str(Stock_ID)
        r = requests.get(url, headers=headers)
        r.encoding = "utf-8"

        dfs = pd.read_html(StringIO(r.text))
        df = pd.concat([df for df in dfs if df.shape[1] > 15 and df.shape[0] > 30])
        df.columns = [df.iloc[0], df.iloc[1]]

        idx = pd.IndexSlice
        df = df.loc[idx[:], idx[["月別", "全體董監持股"], :]]
        df.columns = df.columns.get_level_values(1)
        df = df.set_index(["月別"])

        df["持股(%)"] = pd.to_numeric(df["持股(%)"], errors="coerce")
        df = df[~ df["持股(%)"].isnull()].dropna()["持股(%)"]

        def ChangeName(str):
            dt_obj = datetime.datetime.strptime(str, '%Y/%m')
            dt_str = datetime.datetime.strftime(dt_obj, '%Y-%m')
            return dt_str

        df = df.rename(index=lambda s: ChangeName(s))
        data = []
        index = []
        for cell in list(self.ws0.columns)[9]:
            data.append(cell.value)
        data = data[4:]
        for cell in list(self.ws0.columns)[0]:
            index.append(cell.value)
        index = index[4:]

        DataNow = pd.DataFrame({'date': index, 'Data': data})
        DataNow = DataNow[DataNow['date'].notnull()].rename(index=lambda s: s + 5)

        # 確認爬蟲到的最新資料是否與excel的資料時間點相同，沒有就刪除excel資料點
        while datetime.datetime.strftime(DataNow['date'].iloc[0], "%Y-%m") != df.index[0]:
            DataNow = DataNow.drop(DataNow.index[0])
        UpdateData = DataNow[DataNow['Data'].isnull()]

        pd.options.mode.chained_assignment = None

        for n in range(len(UpdateData)):
            date = UpdateData['date'].iloc[n]
            date_str = datetime.datetime.strftime(date, "%Y-%m")
            try:
                UpdateData['Data'].iloc[n] = df.loc[date_str]
                r = UpdateData.index[n]
                if self.ws0.cell(row=r, column=1).value == date:
                    self.ws0.cell(row=r, column=10).value = UpdateData['Data'].iloc[n]
                    self.ws0.cell(row=r, column=10).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    print("更新月份: " + date_str + " 的股東占比: " + str(self.ws0.cell(row=r, column=10).value))
            except:
                print("Doesn't get " + date_str + " Data")

        self.wb.save(path)
        time.sleep(20)
        print("Directors and supervisors end")
    def Update_Season_report(self, Stock_ID, path):
        '''    從資料庫獲取季報最新日期    '''
        Revenue_Season = data.get2('營業收入合計', 5)
        # print(Revenue_Season)
        Revenue_Season = Revenue_Season[Stock_ID]

        '''    時間判斷    '''
        # 改成用資料庫的最新時間尤佳
        latest_date = Revenue_Season.dropna().index[-1]
        latest_date_str = self.Season_determination(latest_date)
        table_month = self.ws1["E1"].value
        add_column_num = 4 * (int(latest_date_str[0:4]) - int(table_month[0:4])) + (
                int(latest_date_str[-1]) - int(table_month[-1]))

        if add_column_num <= 0:
            print("No data need to update.")
        else:

            '''        根據相差月份取相對應數量的資料        '''
            get_data_num = add_column_num + 6
            Revenue_Season = data.get2('營業收入合計', get_data_num) * 0.00001  # 單位: 億
            # 營業利益率，也可以簡稱營益率，英文Operating Margin或Operating profit Margin
            OPM_raw = data.get2('營業利益（損失）', get_data_num) * 0.00001  # 單位: 億
            gross_profit = data.get2('營業毛利（毛損）', get_data_num) * 0.00001  # 單位: 億
            Equity = data.get2("股本合計", get_data_num) * 0.00001  # 單位: 億
            profit_before_tax = data.get2("繼續營業單位稅前淨利（淨損）", get_data_num) * 0.00001  # 單位: 億  本期稅前淨利（淨損）
            profit_after_tax = data.get2("本期淨利（淨損）", get_data_num) * 0.00001  # 單位: 億
            Operating_costs = data.get2("營業成本合計", get_data_num) * 0.00001  # 單位: 億
            Account_receivable = data.get2("應收帳款淨額", get_data_num) * 0.00001  # 單位: 億
            inventory = data.get2("存貨", get_data_num) * 0.00001  # 單位: 億
            Assets = data.get2("資產總計", get_data_num) * 0.00001  # 單位: 億
            Liabilities = data.get2("負債總計", get_data_num) * 0.00001  # 單位: 億
            Accounts_payable = data.get2("應付帳款", get_data_num) * 0.00001  # 單位: 億
            Intangible_Assets = data.get2("無形資產", get_data_num) * 0.00001  # 單位: 億
            Depreciation = data.get2("折舊費用", get_data_num, table="Cash_flows") * 0.00001  # 單位: 億
            Net_Income = data.get2('本期淨利（淨損）', get_data_num) * 0.00001  # 單位: 億
            # 修正：因為有些股東權益的名稱叫作「權益總計」有些叫作「權益總額」，所以要先將這兩個dataframe合併起來喔！
            權益總計 = data.get2('權益總計', get_data_num)
            權益總額 = data.get2('權益總額', get_data_num)
            # 把它們合併起來（將「權益總計」為NaN的部分填上「權益總額」）
            Shareholders_equity = 權益總計.fillna(權益總額, inplace=False) * 0.00001  # 單位: 億

            price_num = add_column_num * 65
            price = data.get2("收盤價", price_num)

            '''        輸入數字並存在變數中，可以透過該變數(字串)，呼叫特定股票        '''
            Revenue_Season = Revenue_Season[Stock_ID]
            OPM_raw = OPM_raw[Stock_ID]
            gross_profit = gross_profit[Stock_ID]
            Equity = Equity[Stock_ID]
            price = price[Stock_ID]
            profit_before_tax = profit_before_tax[Stock_ID]
            profit_after_tax = profit_after_tax[Stock_ID]
            Operating_costs = Operating_costs[Stock_ID]
            Account_receivable = Account_receivable[Stock_ID]
            inventory = inventory[Stock_ID]
            Assets = Assets[Stock_ID]
            Liabilities = Liabilities[Stock_ID]
            Accounts_payable = Accounts_payable[Stock_ID]
            Intangible_Assets = Intangible_Assets[Stock_ID]
            Depreciation = Depreciation[Stock_ID]
            Net_Income = Net_Income[Stock_ID]
            Shareholders_equity = Shareholders_equity[Stock_ID]

            '''        拆解數據處理        '''
            D_Depreciation = self.DataProcess(Depreciation, cum=False)
            '''        累積數據處理        '''
            C_Return_On_Equity = Net_Income / Shareholders_equity * 100
            C_Return_On_Equity = self.DataProcess(C_Return_On_Equity, cum=True)

            C_profit_after_tax = self.DataProcess(profit_after_tax, cum=True)
            C_Revenue_Season = self.DataProcess(Revenue_Season, cum=True)
            C_profit_after_tax = C_profit_after_tax / C_Revenue_Season * 100

            C_Shareholders_equity = Shareholders_equity / Assets * 100

            newAssets = []
            for idx in range(len(Assets)):
                newAssets.append((Assets[idx] + Assets[idx-1]) / 2)
            newAssets = pd.Series(newAssets, index=Assets.index)
            newAssets = newAssets.drop(labels=[Assets.index[0]])
            C_newAssets = self.DataProcess(newAssets, cum=True)
            C_Total_Assets_Turnover = C_Revenue_Season / C_newAssets * 4

            add_column_num *= -1

            for add_row in range(add_column_num, 0, 1):

                self.ws1.insert_cols(5, amount=1)

                Update_Season_date = Revenue_Season.index[add_row]
                Update_Season_str = Update_Season_date.strftime('%Y-%m-%d')
                Season_lastyear = self.deltaSeasons(Update_Season_date, 4)
                Season_prev4Season = self.deltaSeasons(Update_Season_date, 3)
                Season_prevSeason = self.deltaSeasons(Update_Season_date, 1)

                '''  新增季度標籤  '''
                Update_Season = self.Season_determination(Update_Season_date)

                self.ws1.cell(row=1, column=5).value = Update_Season
                self.ws1.cell(row=1, column=5).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                self.ws1.cell(row=1, column=5).fill = PatternFill(fill_type="solid", fgColor="DDDDDD")
                print("新增標籤:", self.ws1.cell(row=1, column=5).value)

                '''  新增當期營收、當期營收年成長率  '''
                SR = Revenue_Season.loc[Update_Season_date]
                SR_4 = Revenue_Season.loc[Season_lastyear]
                SRevenue_YG = (SR - SR_4) / SR_4 * 100

                self.Write2Excel(SR, rounds=2, sheet=self.ws1, rows=3, cols=5, string="當季營收", date=Update_Season_str)
                self.Write2Excel(SRevenue_YG, rounds=2, sheet=self.ws1, rows=4, cols=5, string="年增率", date=Update_Season_str)

                '''   營業毛利率   '''
                GP = gross_profit.loc[Update_Season_date] / SR * 100

                self.Write2Excel(GP, rounds=2, sheet=self.ws1, rows=6, cols=5, string="營業毛利率", date=Update_Season_str)

                '''   營業利益率、營業利益成長率   '''
                OPM = OPM_raw.loc[Update_Season_date] / SR * 100
                OPM_1 = OPM_raw.loc[Season_prevSeason] / Revenue_Season.loc[Season_prevSeason] * 100
                OPM_SG = (OPM - OPM_1) / OPM_1 * 100

                self.Write2Excel(OPM, rounds=2, sheet=self.ws1, rows=7, cols=5, string="營業利益率", date=Update_Season_str)
                self.Write2Excel(OPM_SG, rounds=2, sheet=self.ws1, rows=8, cols=5, string="營業利益成長率", date=Update_Season_str)

                '''   新增股本、股本季增率、當期市值與市值營收比   '''
                price_Eq = price.loc[:Update_Season_date].iloc[-1]  # 確認股本公布當天是否為交易日
                Equity_Eq = Equity.loc[Update_Season_date]  # 取得最新一筆的股本
                Equity_Eq_1 = Equity.loc[Season_prevSeason]

                Equity_Eq_SG = (Equity_Eq - Equity_Eq_1) / Equity_Eq_1 * 100
                Market_value = price_Eq * Equity_Eq / 10  # 市值 = 股價 * 總股數 (股本合計單位為 k元)
                PSR = Revenue_Season.loc[Season_prev4Season: Update_Season_date].sum() / Market_value * 100

                self.Write2Excel(Equity_Eq, rounds=0, sheet=self.ws1, rows=21, cols=5, string="股本", date=Update_Season_str)
                self.Write2Excel(Equity_Eq_SG, rounds=0, sheet=self.ws1, rows=22, cols=5, string="股本季增率", date=Update_Season_str)
                self.Write2Excel(Market_value, rounds=0, sheet=self.ws1, rows=5, cols=5, string="市值", date=Update_Season_str)
                self.Write2Excel(PSR, rounds=2, sheet=self.ws1, rows=19, cols=5, string="營收市值比", date=Update_Season_str)

                '''   新增稅前淨利率、本業收入比率、稅後淨利率、稅後淨利年增率  '''
                PBT = profit_before_tax.loc[Update_Season_date] / SR * 100
                RevenueSource = OPM / PBT
                PAT = profit_after_tax.loc[Update_Season_date] / SR * 100
                PAT_4 = profit_after_tax.loc[Season_lastyear]
                PAT_YG = (profit_after_tax.loc[Update_Season_date] - PAT_4) / PAT_4 * 100

                self.Write2Excel(PBT, rounds=2, sheet=self.ws1, rows=9, cols=5, string="稅前淨利率", date=Update_Season_str)
                self.Write2Excel(RevenueSource, rounds=2, sheet=self.ws1, rows=10, cols=5, string="本業收入比率", date=Update_Season_str)
                self.Write2Excel(PAT, rounds=2, sheet=self.ws1, rows=11, cols=5, string="稅後淨利率", date=Update_Season_str)
                self.Write2Excel(PAT_YG, rounds=2, sheet=self.ws1, rows=12, cols=5, string="稅後淨利年增率", date=Update_Season_str)

                '''   新增EPS、EPS年成長率   '''
                EPS = profit_after_tax.loc[Update_Season_date] / (Equity_Eq / 10)
                EPS_4 = PAT_4 / (Equity.loc[Season_lastyear] / 10)
                EPS_YG = (EPS - EPS_4) / EPS_4 * 100

                self.Write2Excel(EPS, rounds=2, sheet=self.ws1, rows=13, cols=5, string="每股稅後盈餘", date=Update_Season_str)
                self.Write2Excel(EPS_YG, rounds=2, sheet=self.ws1, rows=14, cols=5, string="每股稅後盈餘年成長率", date=Update_Season_str)

                '''   新增應收帳款週轉率、存貨周轉率、存貨營收比   '''
                AR = Account_receivable.loc[Update_Season_date]
                AR_1 = Account_receivable.loc[Season_prevSeason]
                # receivables turnover
                RT = SR / ((AR + AR_1) / 2) * 4

                OC = Operating_costs.loc[Update_Season_date]
                INV = inventory.loc[Update_Season_date]
                INV_1 = inventory.loc[Season_prevSeason]
                # inventory turnover
                IT = OC / ((INV + INV_1) / 2) * 4
                # inventory revenue ratio
                IR = INV / SR * 100

                self.Write2Excel(RT, rounds=2, sheet=self.ws1, rows=16, cols=5, string="應收帳款週轉率", date=Update_Season_str)
                self.Write2Excel(IT, rounds=2, sheet=self.ws1, rows=17, cols=5, string="存貨周轉率", date=Update_Season_str)
                self.Write2Excel(IR, rounds=2, sheet=self.ws1, rows=18, cols=5, string="存貨占營收比", date=Update_Season_str)

                '''   新增應付帳款總資產占比、負債總資產占比、無形資產占比'''
                Ass = Assets.loc[Update_Season_date]
                Lia = Liabilities.loc[Update_Season_date]
                AP = Accounts_payable.loc[Update_Season_date]
                IntA = Intangible_Assets.loc[Update_Season_date]

                Lia_ratio = Lia / Ass * 100
                AP_ratio = AP / Ass * 100
                IntA_ratio = IntA / Ass * 100

                self.Write2Excel(AP_ratio, rounds=2, sheet=self.ws1, rows=23, cols=5, string="供應商應付帳款總資產占比", date=Update_Season_str)
                self.Write2Excel(Lia_ratio, rounds=2, sheet=self.ws1, rows=24, cols=5, string="負債總資產占比", date=Update_Season_str)
                self.Write2Excel(IntA_ratio, rounds=2, sheet=self.ws1, rows=25, cols=5, string="無形資產占比", date=Update_Season_str)

                '''   新增折舊、折舊負擔比率'''
                Dep = D_Depreciation.loc[Update_Season_date]
                # Debt Asset ratio
                DAR = Dep / SR

                self.Write2Excel(Dep, rounds=2, sheet=self.ws1, rows=27, cols=5, string="折舊", date=Update_Season_str)
                self.Write2Excel(DAR, rounds=2, sheet=self.ws1, rows=28, cols=5, string="折舊負擔比率", date=Update_Season_str)

                '''   杜邦分析   '''
                C_ROE = C_Return_On_Equity.loc[Update_Season_date]
                if Update_Season_date.month == 5:
                    CE_ROE = C_ROE * 4
                elif Update_Season_date.month == 8:
                    CE_ROE = C_ROE * 2
                elif Update_Season_date.month == 11:
                    CE_ROE = C_ROE * 4 /3
                else:
                    CE_ROE = C_ROE
                C_TAT = C_Total_Assets_Turnover.loc[Update_Season_date]
                C_PAT = C_profit_after_tax.loc[Update_Season_date]
                C_SE = C_Shareholders_equity.loc[Update_Season_date]
                # Equity Multiplier
                C_EM = 1 / C_SE * 100

                self.Write2Excel(C_ROE, rounds=2, sheet=self.ws1, rows=30, cols=5, string="股東權益報酬率(季)", date=Update_Season_str)
                self.Write2Excel(CE_ROE, rounds=2, sheet=self.ws1, rows=31, cols=5, string="股東權益報酬率(年預估)", date=Update_Season_str)
                self.Write2Excel(C_PAT, rounds=2, sheet=self.ws1, rows=32, cols=5, string="稅後淨利率(累計)", date=Update_Season_str)
                self.Write2Excel(C_TAT, rounds=2, sheet=self.ws1, rows=33, cols=5, string="總資產週轉率(次/年)", date=Update_Season_str)
                self.Write2Excel(C_EM, rounds=2, sheet=self.ws1, rows=34, cols=5, string="權益係數", date=Update_Season_str)
                self.Write2Excel(C_SE, rounds=2, sheet=self.ws1, rows=35, cols=5, string="股東權益總額(%)", date=Update_Season_str)

            self.wb.save(path)

        # 營收年成長率
        condition_SG = pd.Series([], dtype=pd.StringDtype())
        for date, data1 in zip(self.ws1['E1':'L1'], self.ws1['E4':'L4']):
            for e, a in zip(date, data1):
                condition_SG[e.value] = a.value
        condition_SG = condition_SG.fillna(0) < 0

        # 營收利益成長率
        condition_OPMSG2 = pd.Series([], dtype=pd.StringDtype())
        for date, data1 in zip(self.ws1['E1':'L1'], self.ws1['E8':'L8']):
            for e, a in zip(date, data1):
                condition_OPMSG2[e.value] = a.value
        condition_OPMSG2 = condition_OPMSG2 < -30

        # 營收利益成長率
        condition_OPMSG = pd.Series([], dtype=pd.StringDtype())
        for date, data1 in zip(self.ws1['E1':'L1'], self.ws1['E8':'L8']):
            for e, a in zip(date, data1):
                condition_OPMSG[e.value] = a.value
        condition_OPMSG = condition_OPMSG.between(-30, -20)

        # 營收市值比
        condition_PSR = pd.Series([], dtype=pd.StringDtype())
        for date, data1 in zip(self.ws1['E1':'L1'], self.ws1['E19':'L19']):
            for e, a in zip(date, data1):
                condition_PSR[e.value] = a.value
        condition_PSR = condition_PSR.fillna(0) < 20

        # EPS年成長率
        condition_EPS_YG = pd.Series([], dtype=pd.StringDtype())
        for date, data1 in zip(self.ws1['E1':'L1'], self.ws1['E17':'L17']):
            for e, a in zip(date, data1):
                condition_EPS_YG[e.value] = a.value
        condition_EPS_YG = condition_EPS_YG.fillna(0) < 0

        # 負債總額
        condition_Lia_ratio = pd.Series([], dtype=pd.StringDtype())
        for date, data1 in zip(self.ws1['E1':'L1'], self.ws1['E24':'L24']):
            for e, a in zip(date, data1):
                condition_Lia_ratio[e.value] = a.value
        condition_Lia_ratio = condition_Lia_ratio.fillna(0) > 40

        # 無形資產占比
        condition_IntA_ratio1 = pd.Series([], dtype=pd.StringDtype())
        for date, data1 in zip(self.ws1['E1':'L1'], self.ws1['E25':'L25']):
            for e, a in zip(date, data1):
                condition_IntA_ratio1[e.value] = a.value
        condition_IntA_ratio1 = condition_IntA_ratio1.fillna(0) > 10

        # 無形資產占比
        condition_IntA_ratio = pd.Series([], dtype=pd.StringDtype())
        for date, data1 in zip(self.ws1['E1':'L1'], self.ws1['E25':'L25']):
            for e, a in zip(date, data1):
                condition_IntA_ratio[e.value] = a.value
        condition_IntA_ratio = condition_IntA_ratio.fillna(0) > 30

        # 折舊負擔比率
        condition_DAR = pd.DataFrame()
        for date, data1, data2 in zip(self.ws1['E1':'L1'], self.ws1['E28':'L28'], self.ws1['E6':'L6']):
            for e, a1, a2 in zip(date, data1, data2):
                condition_DAR[e.value] = [a1.value, a2.value]
        condition_DAR = condition_DAR.fillna(0).iloc[0] > condition_DAR.fillna(0).iloc[1]

        '''   判斷條件   '''
        for c in range(5, 13):
            n = c - 5
            # 營收年成長率
            self.WarningFunc(condition_SG[n], sheet=self.ws1, rows=4, cols=c, threat='False')
            # 營收利益成長率
            self.WarningFunc(condition_OPMSG[n], sheet=self.ws1, rows=8, cols=c, threat='False')
            # 營收利益成長率
            self.WarningFunc(condition_OPMSG2[n], sheet=self.ws1, rows=8, cols=c, threat='True')
            # 營收市值比
            self.WarningFunc(condition_PSR[n], sheet=self.ws1, rows=19, cols=c, threat='False')
            # EPS年成長率
            self.WarningFunc(condition_EPS_YG[n], sheet=self.ws1, rows=17, cols=c, threat='False')
            # 負債總額
            self.WarningFunc(condition_Lia_ratio[n], sheet=self.ws1, rows=24, cols=c, threat='False')
            # 無形資產占比
            self.WarningFunc(condition_IntA_ratio1[n], sheet=self.ws1, rows=25, cols=c, threat='False')
            self.WarningFunc(condition_IntA_ratio[n], sheet=self.ws1, rows=25, cols=c, threat='True')
            # 折舊負擔比率
            self.WarningFunc(condition_DAR[n], sheet=self.ws1, rows=28, cols=c, threat='False')
        self.wb.save(path)
    def Update_CashFlow(self, Stock_ID, path):

        '''    從資料庫獲取季報最新日期    '''
        Cash_Flow_for_investing = data.get2("投資活動之淨現金流入（流出）", 5)
        Cash_Flow_for_investing = Cash_Flow_for_investing[Stock_ID]

        '''    時間判斷    '''
        latest_date = Cash_Flow_for_investing.dropna().index[-1]
        if latest_date.month == 3:
            year = latest_date.year - 1
        else:
            year = latest_date.year
        table_year = self.ws2["D1"].value
        add_column_num = year - int(table_year)

        '''    確認當年資料是否需要更新    '''
        if self.ws2["D4"].value != Cash_Flow_for_investing[-1]:
            self.ws2.delete_cols(4, 1)
            print("當年度資料更新")
            add_column_num += 1

        if add_column_num <= 0:
            print("No data need to update.")
        else:
            '''        根據相差月份取相對應數量的資料        '''
            get_data_num = add_column_num * 4
            # Cash Flow for investing
            Cash_Flow_for_investing = data.get2("投資活動之淨現金流入（流出）", get_data_num)
            # Operating Cash Flow
            Operating_Cash_Flow = data.get2("營業活動之淨現金流入（流出）", get_data_num)
            # Cash Flows Provided from Financing Activities
            Cash_Flow_for_Financing = data.get2("籌資活動之淨現金流入（流出）", get_data_num)
            # Cash Balances - Beginning of Period
            Cash_Balances_Beginning = data.get2("期初現金及約當現金餘額", get_data_num)
            # Cash Balances - End of Period
            Cash_Balances_End = data.get2("期末現金及約當現金餘額", get_data_num)


            '''        輸入數字並存在變數中，可以透過該變數(字串)，呼叫特定股票        '''
            Cash_Flow_for_investing = Cash_Flow_for_investing[Stock_ID] * 0.00001 # 單位:億
            Operating_Cash_Flow = Operating_Cash_Flow[Stock_ID] * 0.00001 # 單位:億
            # Free cash flow(FCF)
            Free_cash_flow = (Cash_Flow_for_investing + Operating_Cash_Flow)
            Cash_Flow_for_Financing = Cash_Flow_for_Financing[Stock_ID] * 0.00001 # 單位:億
            Cash_Balances_Beginning = Cash_Balances_Beginning[Stock_ID] * 0.00001 # 單位:億
            Cash_Balances_End = Cash_Balances_End[Stock_ID] * 0.00001 # 單位:億

            Cash_Flow_for_investing = self.CashFlowGet(Cash_Flow_for_investing)
            Operating_Cash_Flow = self.CashFlowGet(Operating_Cash_Flow)
            Free_cash_flow = self.CashFlowGet(Free_cash_flow)
            Cash_Flow_for_Financing = self.CashFlowGet(Cash_Flow_for_Financing)
            Cash_Balances_Beginning = self.CashFlowGet(Cash_Balances_Beginning)
            Cash_Balances_End = self.CashFlowGet(Cash_Balances_End)

            add_column_num *= -1

            for add_row in range(add_column_num, 0, 1):

                self.ws2.insert_cols(4, amount=1)

                Update_year = Cash_Flow_for_investing.index[add_row]
                Update_year_str = Update_year.strftime('%Y')

                '''  新增年度標籤  '''
                self.ws2.cell(row=1, column=4).value = Update_year_str
                self.ws2.cell(row=1, column=4).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                self.ws2.cell(row=1, column=4).fill = PatternFill(fill_type="solid", fgColor="DDDDDD")
                print("新增標籤:", self.ws2.cell(row=1, column=4).value)

                '''  新增營業活動現金、理財活動現金、自由現金流量、籌資活動現金'''
                ICF = Cash_Flow_for_investing.loc[Update_year]
                OCF = Operating_Cash_Flow.loc[Update_year]
                FCF = Free_cash_flow.loc[Update_year]
                CFPFA = Cash_Flow_for_Financing.loc[Update_year]

                self.Write2Excel(OCF, rounds=1, sheet=self.ws2, rows=3, cols=4, string="營業活動現金", date=Update_year_str)
                self.Write2Excel(ICF, rounds=1, sheet=self.ws2, rows=4, cols=4, string="理財活動現金", date=Update_year_str)
                self.Write2Excel(FCF, rounds=1, sheet=self.ws2, rows=5, cols=4, string="自由現金流量", date=Update_year_str)
                self.Write2Excel(CFPFA, rounds=1, sheet=self.ws2, rows=6, cols=4, string="籌資活動現金", date=Update_year_str)

                self.PNdetermination(OCF, sheet=self.ws2, rows=3, cols=4)
                self.PNdetermination(ICF, sheet=self.ws2, rows=4, cols=4)
                self.PNdetermination(FCF, sheet=self.ws2, rows=5, cols=4)
                self.PNdetermination(CFPFA, sheet=self.ws2, rows=6, cols=4)

                '''  新增期初現金及約當現金餘額、期末現金及約當現金餘額'''
                CBBP = Cash_Balances_Beginning.loc[Update_year]
                CBEP = Cash_Balances_End.loc[Update_year]

                self.Write2Excel(CBBP, rounds=1, sheet=self.ws2, rows=7, cols=4, string="期初現金及約當現金餘額", date=Update_year_str)
                self.Write2Excel(CBEP, rounds=1, sheet=self.ws2, rows=8, cols=4, string="期末現金及約當現金餘額", date=Update_year_str)
        try:
            '''   判斷條件   '''
            for c in range(4, 9):
                # 營業活動現金
                condition_OCF = int(self.ws2.cell(row=3, column=c).value) < 0
                self.WarningFunc(condition_OCF, sheet=self.ws2, rows=3, cols=c, threat='True')
                # 自由現金
                condition_FCF = int(self.ws2.cell(row=5, column=c).value) < 0
                self.WarningFunc(condition_FCF, sheet=self.ws2, rows=5, cols=c, threat='True')
        except:
            print(Stock_ID + " 警告上色錯誤")

        self.wb.save(path)
    def Update_PER(self, Stock_ID,path):

        '''    從資料庫獲取季報最新日期    '''
        # *未結束年度之EPS預估值, 以最近四季之合計EPS取代之, 例如: 某股票EPS僅公布至今年第三季, 則
        # 今年之預估EPS = 去年第四季至今年第三季之合計EPS。
        # https://goodinfo.tw/StockInfo/ShowK_ChartFlow.asp?RPT_CAT=PER&STOCK_ID=2330&CHT_CAT=QUAR

        '''    使用現在的時間當作最新的更新時間點    '''
        now = datetime.datetime.now()
        date = pd.Series(now)
        df = pd.DataFrame()
        df['Quarter'] = pd.to_datetime(date)
        df['Quarter'] = df['Quarter'].dt.to_period('Q').dt.strftime("%YQ%q")
        latest_date_str = df['Quarter'].iloc[-1]

        table_month = self.ws4["A16"].value
        add_row_num = 4 * (int(latest_date_str[0:4]) - int(table_month[0:4])) + (
                    int(latest_date_str[-1]) - int(table_month[-1]))

        if add_row_num <= 0:
            print("Update PER this year.")
        else:
            print("Increase PER this season and update PER this year.")

        PER_data = [self.ws4.cell(row=n, column=1).value[0:4] for n in range(16, 20) if self.ws4.cell(row=n, column=1).value]
        Update_row = 0
        for n in range(len(PER_data)):
            if PER_data[n] == now.strftime("%Y"):
                Update_row += 1

        total_num = Update_row + add_row_num

        get_data_num = total_num + 4
        Equity = data.get2("股本合計", get_data_num) * 0.00001  # 單位: 億
        profit_after_tax = data.get2("本期淨利（淨損）", get_data_num) * 0.00001  # 單位: 億

        price_num = (total_num) * 100
        price = data.get2("收盤價", price_num)

        Equity = Equity[Stock_ID].dropna()
        profit_after_tax = profit_after_tax[Stock_ID].dropna()
        price = price[Stock_ID].dropna()

        price_Q1 = price[price.index.month == 1].append(price[price.index.month == 2]).append(price[price.index.month == 3]).sort_index()
        price_Q2 = price[price.index.month == 4].append(price[price.index.month == 5]).append(price[price.index.month == 6]).sort_index()
        price_Q3 = price[price.index.month == 7].append(price[price.index.month == 8]).append(price[price.index.month == 9]).sort_index()
        price_Q4 = price[price.index.month == 10].append(price[price.index.month == 11]).append(price[price.index.month == 12]).sort_index()

        EPS = profit_after_tax / (Equity / 10)
        Estimated_EPS = EPS.rolling(4).sum()

        '''  檢查公布財報的EPS時間與實際時間的差別，如果尚未公布財報則填入現在的時間，新增最新時間資料  '''
        FR_date = self.Season_determination(Estimated_EPS.index[-1])
        num = 4 * (int(latest_date_str[0:4]) - int(FR_date[0:4])) + (int(latest_date_str[-1]) - int(FR_date[-1]))
        for n in range(num):
            date = self.deltaSeasons(Estimated_EPS.index[-1], -1)
            Estimated_EPS[date] = Estimated_EPS[-1]

        Start = 16
        End = 16 + Update_row
        for add_row in range(Start, End):

            Update_date = self.Season2Month(self.ws4.cell(row=add_row, column=1).value)
            Update_Season = self.ws4.cell(row=add_row, column=1).value

            if Update_Season[-1] == "1":
                PRICE = price_Q1.loc[Update_Season[0:4]][-1]
            elif Update_Season[-1] == "2":
                PRICE = price_Q2.loc[Update_Season[0:4]][-1]
            elif Update_Season[-1] == "3":
                PRICE = price_Q3.loc[Update_Season[0:4]][-1]
            else:
                PRICE = price_Q4.loc[Update_Season[0:4]][-1]
            E_EPS = Estimated_EPS.loc[Update_date][-1]
            PER = PRICE / E_EPS

            print("更新 ", self.ws4.cell(row=add_row, column=1).value," 的EPS: ", round(E_EPS, 2))
            self.Write2Excel(PER, rounds=2, sheet=self.ws4, rows=add_row, cols=2, string="更新PER", date=Update_Season)

        add_row_num *= -1

        for add_row in range(add_row_num, 0, 1):

            self.ws4.insert_rows(16, amount=1)

            Update_Season_date = Estimated_EPS.index[add_row]
            Update_Season_str = Update_Season_date.strftime('%Y-%m')

            '''  新增季度標籤  '''
            Update_Season = self.Season_determination(Update_Season_date)

            self.ws4.cell(row=16, column=1).value = Update_Season
            self.ws4.cell(row=16, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            self.ws4.cell(row=16, column=1).fill = PatternFill(fill_type="solid", fgColor="FFEE99")
            print("新增標籤:", self.ws4.cell(row=16, column=1).value)

            '''  新增本益比  '''
            if Update_Season[-1] == "1":
                PRICE = price_Q1.loc[Update_Season[0:4]][-1]
            elif Update_Season[-1] == "2":
                PRICE = price_Q2.loc[Update_Season[0:4]][-1]
            elif Update_Season[-1] == "3":
                PRICE = price_Q3.loc[Update_Season[0:4]][-1]
            else:
                PRICE = price_Q4.loc[Update_Season[0:4]][-1]
            E_EPS = Estimated_EPS.loc[Update_Season_str][-1]
            PER = PRICE / E_EPS

            print("使用季度: ", Update_Season, " 所得到的EPS: ", round(E_EPS, 2))
            self.Write2Excel(PER, rounds=2, sheet=self.ws4, rows=16, cols=2, string="新增PER", date=Update_Season)

        self.wb.save(path)
    def Update_PRICEToday(self, Stock_ID, path):

        Highest = data.get('最高價', 1)
        Lowest = data.get('最低價', 1)
        Opening = data.get('開盤價', 1)
        Closing = data.get('收盤價', 1)

        Highest = Highest[Stock_ID]
        Lowest = Lowest[Stock_ID]
        Opening = Opening[Stock_ID]
        Closing = Closing[Stock_ID]

        DATES = Highest.index[0]

        DATES_str = DATES.strftime("%Y/%m/%d")

        self.ws4.cell(row=13, column=1).value = DATES_str
        self.ws4.cell(row=13, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.Write2Excel(Highest.iloc[0], rounds=1, sheet=self.ws4, rows=12, cols=3, string="新增最高價", date=DATES_str)
        self.Write2Excel(Lowest.iloc[0], rounds=1, sheet=self.ws4, rows=13, cols=3, string="新增最低價", date=DATES_str)
        self.Write2Excel(Opening.iloc[0], rounds=1, sheet=self.ws4, rows=12, cols=5, string="新增開盤價", date=DATES_str)
        self.Write2Excel(Closing.iloc[0], rounds=1, sheet=self.ws4, rows=13, cols=5, string="新增收盤價", date=DATES_str)

        self.wb.save(path)

class SelectStock():
    def __init__(self):
        pass

    def toSeasonal(self, df):
        season4 = df[df.index.month == 3]
        season1 = df[df.index.month == 5]
        season2 = df[df.index.month == 8]
        season3 = df[df.index.month == 11]

        season1.index = season1.index.year
        season2.index = season2.index.year
        season3.index = season3.index.year
        season4.index = season4.index.year - 1

        # 第一季即自己的資料
        newseason1 = season1
        # 第二季為 Q2 扣掉累積至Q1的資料
        # reindex_like: https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.reindex_like.html
        # reindex_like: 將season1的資料依照season2的index重新排列
        newseason2 = season2 - season1.reindex_like(season2)
        # 第三季為 Q3 扣掉累積至Q2的資料
        newseason3 = season3 - season2.reindex_like(season3)
        newseason4 = season4 - season3.reindex_like(season4)

        newseason1.index = pd.to_datetime(newseason1.index.astype(str) + '-05-15')
        newseason2.index = pd.to_datetime(newseason2.index.astype(str) + '-08-14')
        newseason3.index = pd.to_datetime(newseason3.index.astype(str) + '-11-14')
        newseason4.index = pd.to_datetime((newseason4.index + 1).astype(str) + '-03-31')

        return newseason1.append(newseason2).append(newseason3).append(newseason4).sort_index()

    def mystrategy(self, date, exec, bool):

        股本 = data.get3(name='股本合計', n=1, start=date)
        price = data.get3(name='收盤價', n=120, start=date)
        當天股價 = price[:股本.index[-1]].iloc[-1]
        當天股本 = 股本.iloc[-1]
        市值 = 當天股本 * 當天股價 / 10 * 1000

        df1 = self.toSeasonal(data.get3(name='投資活動之淨現金流入（流出）', n=15, start=date))
        df2 = self.toSeasonal(data.get3(name='營業活動之淨現金流入（流出）', n=15, start=date))
        三年自由現金流 = (df1 + df2).iloc[-12:].mean()

        稅後淨利 = data.get3(name='本期淨利（淨損）', n=9, start=date)
        # 股東權益，有兩個名稱，有些公司叫做權益總計，有些叫做權益總額
        # 所以得把它們抓出來
        權益總計 = data.get3(name='權益總計', n=1, start=date)
        權益總額 = data.get3(name='權益總額', n=1, start=date)

        # 並且把它們合併起來
        權益總計.fillna(權益總額, inplace=True)

        股東權益報酬率 = ((稅後淨利.iloc[-4:].sum()) / 權益總計.iloc[-1]) * 100

        營業利益 = data.get3(name='營業利益（損失）', n=9, start=date)
        Revenue_Season = data.get3(name='營業收入合計', n=9, start=date)
        營業利益率 = 營業利益 / Revenue_Season
        前季營業利益率 = 營業利益.shift(1) / Revenue_Season.shift(1)
        營業利益年成長率 = (營業利益率.iloc[-1] / 營業利益率.iloc[-5] - 1) * 100
        八季營益率變化 = (營業利益率 / 前季營業利益率 - 1) * 100
        八季營益率變化 = 八季營益率變化.dropna(axis=1, how="all").dropna(how="all")

        當月營收 = data.get3(name='當月營收', n=12, start=date) * 1000
        年營收 = 當月營收.iloc[-12:].sum()
        市值營收比 = 市值 / 年營收

        MR_YearGrowth = data.get3(name='去年同月增減(%)', n=12, start=date)
        短期營收年增 = MR_YearGrowth.rolling(3).mean().reindex(index=MR_YearGrowth.index).iloc[-1]
        長期營收年增 = MR_YearGrowth.rolling(12).mean().reindex(index=MR_YearGrowth.index).iloc[-1]

        稅後淨利率 = 稅後淨利 / Revenue_Season
        去年稅後淨利率 = 稅後淨利率.shift(4)
        稅後淨利年增 = (稅後淨利率 - 去年稅後淨利率) / 去年稅後淨利率 * 100
        稅後淨利年增 = 稅後淨利年增
        短期淨利年增 = 稅後淨利年增.iloc[-1]
        長期淨利年增 = 稅後淨利年增[-4:].mean()

        INV = data.get3(name="存貨", n=3, start=date)
        OC = data.get3(name="營業成本合計", n=2, start=date)
        存貨周轉率 = OC.iloc[-1] / ((INV.iloc[-1] + INV.iloc[-2]) / 2) * 4
        前季存貨周轉率 = OC.iloc[-2] / ((INV.iloc[-2] + INV.iloc[-3]) / 2) * 4
        存貨周轉變化率 = (存貨周轉率 - 前季存貨周轉率) / 前季存貨周轉率 * 100

        rsv = (price.iloc[-1] - price.iloc[-60:].min()) / (price.iloc[-60:].max() - price.iloc[-60:].min())

        dict = {
            "市值": 市值,
            "三年自由現金流": 三年自由現金流,
            "股東權益報酬率": 股東權益報酬率,
            "營業利益年成長率": 營業利益年成長率,
            "八季營益率變化": 八季營益率變化,
            "市值營收比": 市值營收比,
            "短期營收年增": 短期營收年增,
            "長期營收年增": 長期營收年增,
            "短期淨利年增": 短期淨利年增,
            "長期淨利年增": 長期淨利年增,
            "存貨周轉變化率": 存貨周轉變化率,
            "rsv": rsv
        }

        ops = {
            "<": operator.lt,
            "<=": operator.le,
            ">": operator.gt,
            ">=": operator.ge,
            "=": operator.eq,
        }

        def operator_func(var, op, con):
            a = dict[var]
            if con in dict:
                b = dict[con]
            else:
                b = float(con)
            return ops[op](a, b)

        condition_list = []
        for b, e in zip(bool, exec):
            if len(e.split()) >= 3 and b==True:
                oper = operator_func(*(e.split()))
                if isinstance(oper, pd.DataFrame):
                    oper = dict[e.split()[0]][oper].isnull().sum() <= 0
                condition_list.append(oper)


        select_stock = condition_list[0]
        for con in condition_list:
            select_stock = select_stock & con

        return select_stock[select_stock]

    def backtest(self, root, start_date, end_date, hold_days, cond, bool, weight='average', benchmark=None, stop_loss=None,
                  stop_profit=None):
        # portfolio check
        if weight != 'average' and weight != 'price':
            print('Backtest stop, weight should be "average" or "price", find', weight, 'instead')

        # get price data in order backtest
        data.date = end_date
        price = data.get('收盤價', (end_date - start_date).days)
        # start from 1 TWD at start_date,
        end = 1
        date = start_date

        # record some history
        equality = pd.Series()
        nstock = {}
        comparison = []
        transactions = pd.DataFrame()
        maxreturn = -10000
        minreturn = 10000

        def trading_day(date):
            if date not in price.index:
                temp = price.loc[date:]
                if temp.empty:
                    return price.index[-1]
                else:
                    return temp.index[0]
            else:
                return date

        def date_iter_periodicity(start_date, end_date, hold_days):
            date = start_date
            while date < end_date:
                yield (date), (date + datetime.timedelta(hold_days))
                date += datetime.timedelta(hold_days)

        def date_iter_specify_dates(start_date, end_date, hold_days):
            dlist = [start_date] + hold_days + [end_date]
            if dlist[0] == dlist[1]:
                dlist = dlist[1:]
            if dlist[-1] == dlist[-2]:
                dlist = dlist[:-1]
            for sdate, edate in zip(dlist, dlist[1:]):
                yield (sdate), (edate)

        if isinstance(hold_days, int):
            dates = date_iter_periodicity(start_date, end_date, hold_days)
        elif isinstance(hold_days, list):
            dates = date_iter_specify_dates(start_date, end_date, hold_days)
        else:
            print('the type of hold_dates should be list or int.')
            return None

        figure, ax = plt.subplots(2, 1, sharex=True, sharey=False)

        keep_list = []
        keep_idx = pd.Index(keep_list)
        for sdate, edate in dates:

            stock_list = []
            # select stocks at date
            data.date = sdate
            # https://stackoverflow.com/questions/39137506/map-to-list-error-series-object-not-callable
            stocks = self.mystrategy(sdate, cond, bool)
            # Idx = stocks.index
            # Idx = stocks.index.append([keep_idx])
            Idx = stocks.index.append([keep_idx]).drop_duplicates()
            print("回測的股票為: ", Idx)
            # hold the stocks for hold_days day
            s = price[Idx & price.columns][sdate:edate].iloc[1:]

            if s.empty:
                s = pd.Series(1, index=pd.date_range(sdate + datetime.timedelta(days=1), edate))
            else:
                if stop_loss != None:
                    below_stop = ((s / s.bfill().iloc[0]) - 1) * 100 < -np.abs(stop_loss)
                    below_stop = (below_stop.cumsum() > 0).shift(2).fillna(False)
                    s[below_stop] = np.nan
                if stop_profit != None:
                    above_stop = ((s / s.bfill().iloc[0]) - 1) * 100 > np.abs(stop_profit)
                    above_stop = (above_stop.cumsum() > 0).shift(2).fillna(False)
                    s[above_stop] = np.nan

                s.dropna(axis=1, how='all', inplace=True)
                keep_list = s.dropna(axis=1)
                keep_idx = pd.Index(keep_list.columns)

                # record transactions
                bprice = s.bfill().iloc[0]
                sprice = s.apply(lambda s: s.dropna().iloc[-1])
                transactions = transactions.append(pd.DataFrame({
                    'buy_price': bprice,
                    'sell_price': sprice,
                    'lowest_price': s.min(),
                    'highest_price': s.max(),
                    'buy_date': pd.Series(s.index[0], index=s.columns),
                    'sell_date': s.apply(lambda s: s.dropna().index[-1]),
                    'profit(%)': (sprice / bprice - 1) * 100
                })).sort_index(ascending=True)

                s.ffill(inplace=True)
                s = s.sum(axis=1)
                # calculate equality
                # normalize and average the price of each stocks
                if weight == 'average':
                    s = s / s.bfill().iloc[0]
                else:
                    s = s / s.bfill()[0]

            # print some log
            start_time = sdate.strftime("%Y-%m-%d")
            end_time = edate.strftime("%Y-%m-%d")

            profit_str = "{} - {} 報酬率: {:.2f}% nstock {}".format(start_time, end_time, (s.iloc[-1] / s.iloc[0] * 100 - 100), len(Idx))
            comparison.append(profit_str)
            print(profit_str)

            benchmark1 = price['0050'][sdate:edate].iloc[1:]
            p0050_str = "{} - {} 的0050報酬率: {:.2f}% ".format(start_time, end_time, (benchmark1.iloc[-1] / benchmark1.iloc[0] * 100 - 100))
            comparison.append(p0050_str)
            print(p0050_str)

            maxreturn = max(maxreturn, s.iloc[-1] / s.iloc[0] * 100 - 100)
            minreturn = min(minreturn, s.iloc[-1] / s.iloc[0] * 100 - 100)

            # plot backtest result
            ((s * end - 1) * 100).plot(ax=ax[0])
            equality = equality.append(s * end)
            end = (s / s[0] * end).iloc[-1]

            if math.isnan(end):
                end = 1

            # add nstock history
            nstock[sdate] = len(stocks)

        print('每次換手最大報酬 : %.2f ％' % maxreturn)
        print('每次換手最少報酬 : %.2f ％' % minreturn)

        if benchmark is None:
            benchmark = price['0050'][start_date:end_date].iloc[1:]

        # bechmark (thanks to Markk1227)
        ((benchmark / benchmark[0] - 1) * 100).plot(ax=ax[0], legend=True, color=(0.8, 0.8, 0.8), grid=True)

        ax[0].set_ylabel('Return On Investment (%)')
        ax[0].grid(linestyle='-.')

        ((benchmark / benchmark.cummax() - 1) * 100).plot(ax=ax[1], legend=True, color=(0.8, 0.8, 0.8))
        ((equality / equality.cummax() - 1) * 100).plot(ax=ax[1], legend=True)
        plt.ylabel('Dropdown (%)')
        plt.grid(linestyle='-.')

        # pd.Series(nstock).plot.bar(ax=ax[2])
        # plt.ylabel('Number of stocks held')
        plt.show()

        return equality, transactions, maxreturn, minreturn, comparison

def save_path_sql(path, source="OR"):
    table_name = "Path"
    exist = table_exist(conn, table_name)
    df = pd.DataFrame()
    if os.path.exists(path):
        if os.path.isdir(path):
            if source == "OR":
                df["category"] = ["directory"]
                df["path"] = [path]
            elif source == "SS":
                df["category"] = ["SSdirectory"]
                df["path"] = [path]
        elif os.path.isfile(path):
            a, b = os.path.splitext(path)
            if b == ".db":
                df["category"] = ["db"]
                df["path"] = [path]
            elif b == ".xlsx":
                df["category"] = ["file"]
                df["path"] = [path]
        else:
            print("it's an invalid path")
    if exist:
        origin = pd.read_sql("SELECT * FROM {}".format(table_name), conn)
        df = origin.append(df)
        df = df.drop_duplicates(subset=["category", "path"], keep='last')
        df.to_sql(table_name, conn, index=False, if_exists='replace')
    else:
        df.to_sql(table_name, conn, index=False, if_exists='append')
def get_path_sql(type):
    table_name = "Path"
    exist = table_exist(conn, table_name)
    if exist:
        df = pd.read_sql("SELECT path FROM {} WHERE category = '{}'".format(table_name, type), conn)
        return [' '.join(map(str, p)) for p in df.values]
def del_path_sql(type, path):
    table_name = "Path"
    conn.execute("DELETE FROM {} WHERE category = '{}' AND path = '{}'".format(table_name, type, path))
    conn.commit()

def date_func(table, type):
    if type == "F":
        latest_date = table_latest_date(conn, table)
        date_list = latest_date + datetime.timedelta(days=1)
        date_list = date_list.strftime('%Y-%m-%d')
    elif type == "T":
        date_list = datetime.datetime.now().strftime('%Y-%m-%d')
    return [date_list]

def exec_func(type, from_Date, to_Date):

    if type == "P":
        date = date_range(from_Date, to_Date)
        table = 'price'
        function = crawl_price
    elif type == "M":
        date = month_range(from_Date, to_Date)
        table = 'monthly_revenue'
        function = crawl_monthly_report
    elif type == "S":
        date = season_range(from_Date, to_Date)
        table = 'finance_statement'
        function = crawl_finance_statement_by_date

    update_table(conn, table, function, date)

# 把列出資料夾的程式碼寫成一個函式
def show_folder_content(folder_path, prefix=None, postfix=None):
    # print(folder_path + '，的資料夾內容：')

    files_list = []
    folder_content = os.listdir(folder_path)
    for item in folder_content:

        fullpath = os.path.join(folder_path, item)

        if os.path.isdir(fullpath):
            # print('資料夾：' + item)
            # 呼叫自己處理這個子資料夾
            files_list += show_folder_content(fullpath, prefix=prefix, postfix=postfix)

        elif os.path.isfile(fullpath):
            # print('檔案：' + item)
            if prefix:
                if item.startswith(prefix):
                    files_list.append(os.path.join(folder_path, item))
            elif postfix:
                if item.endswith(postfix):
                    files_list.append(os.path.join(folder_path, item))
            else:
                files_list.append(os.path.join(folder_path, item))
        # else:
            # print('無法辨識：' + item)
    return files_list

def save_cache_sql(list):
    dic={
        "type": list[0],
        "boolean": list[1],
        "content": list[2],
        "combo": list[3],
        "entry": list[4]
    }
    df = pd.DataFrame(dic)

    table_name = "Cache"
    exist = table_exist(conn, table_name)
    if exist:
        df.to_sql(table_name, conn, index=False, if_exists='replace')
    else:
        df.to_sql(table_name, conn, index=False, if_exists='append')

def get_cache_sql(type, bool=False):
    table_name = "Cache"
    exist = table_exist(conn, table_name)
    if exist:
        if not bool:
            df = pd.read_sql("SELECT * FROM {} WHERE type = '{}'".format(table_name, type), conn)
            if df.size != 0:
                return df
            else:
                return pd.DataFrame()
        else:
            df = pd.read_sql("SELECT boolean FROM {} WHERE type = '{}'".format(table_name, type), conn)
            if df.size != 0:
                return str(df.values[0][0])
            else:
                return False
