import datetime
import os
import sqlite3
import time
from io import StringIO

import openpyxl
import pandas as pd
import requests
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Alignment
from openpyxl.styles import Font

from finlab.data import Data

# 一個工作簿(workbook)在建立的時候同時至少也新建了一張工作表(worksheet)
# wb = Workbook()

conn = sqlite3.connect(os.path.join("data", "data.db"))
data = Data()


# 把列出資料夾的程式碼寫成一個函式
def show_folder_content(folder_path, prefix=None, postfix=None):
    # print(folder_path + '，的資料夾內容：')

    files_list = []
    folder_content = os.listdir(folder_path)
    for item in folder_content:

        fullpath = os.path.join(folder_path, item)

        if os.path.isdir(fullpath):
            # print('資料夾：' + item)
            # 呼叫自己處理這個子資料夾
            files_list += show_folder_content(fullpath, prefix=prefix, postfix=postfix)

        elif os.path.isfile(fullpath):
            # print('檔案：' + item)
            if prefix:
                if item.startswith(prefix):
                    files_list.append(os.path.join(folder_path, item))
            elif postfix:
                if item.endswith(postfix):
                    files_list.append(os.path.join(folder_path, item))
            else:
                files_list.append(os.path.join(folder_path, item))
        # else:
        # print('無法辨識：' + item)
    return files_list


def months(str1, str2):
    year1 = datetime.datetime.strptime(str1[0:10], "%Y-%m").year
    year2 = datetime.datetime.strptime(str2[0:10], "%Y-%m").year
    month1 = datetime.datetime.strptime(str1[0:10], "%Y-%m").month
    month2 = datetime.datetime.strptime(str2[0:10], "%Y-%m").month
    num = (year1 - year2) * 12 + (month1 - month2)
    return num


def Update_Monthly_report(path, Stock_ID):
    '''    從資料庫獲取月營收最新日期    '''
    Revenue_Month = data.get('當月營收', 2)

    '''    時間判斷    '''
    # 改成用資料庫的最新時間尤佳
    latest_date = Revenue_Month[Stock_ID].dropna().index[-1]
    latest_date_str = datetime.datetime.strftime(latest_date, '%Y-%m')
    table_month = datetime.datetime.strftime(ws0["A5"].value, '%Y-%m')

    if table_month == latest_date_str:
        print("No data need to update.")
    else:
        add_row_num = months(latest_date_str, table_month)

        '''        根據相差月份取相對應數量的資料        '''
        add_revenue = add_row_num + 24
        Revenue_Month = data.get('當月營收', add_revenue) * 0.00001
        add_price = add_row_num * 40
        price = data.get('收盤價', add_price)
        MR_MonthGrowth = data.get('上月比較增減(%)', add_revenue)
        MR_YearGrowth = data.get('去年同月增減(%)', add_revenue)

        # 輸入數字並存在變數中，可以透過該變數(字串)，呼叫特定股票
        Month_Revenue = Revenue_Month[Stock_ID]
        price = price[Stock_ID]
        MR_MonthGrowth = MR_MonthGrowth[Stock_ID]
        MR_YearGrowth = MR_YearGrowth[Stock_ID]
        # print("****", Revenue_Month)

        MAG_3M = MR_YearGrowth.rolling(3).mean().reindex(index=MR_YearGrowth.index)
        MAG_3M = round(MAG_3M, 2)
        MAG_12M = MR_YearGrowth.rolling(12).mean().reindex(index=MR_YearGrowth.index)
        MAG_12M = round(MAG_12M, 2)

        add_row_num -= 1

        for add_row in range(add_row_num, -1, -1):

            ws0.insert_rows(5, amount=1)

            '''  新增月份  '''
            Update_Month = latest_date - relativedelta(months=add_row)
            ws0.cell(row=5, column=1).value = Update_Month
            ws0.cell(row=5, column=1).number_format = "mmm-yy"
            ws0.cell(row=5, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            print("新增:", ws0.cell(row=5, column=1).value)

            '''        更新營收        '''
            MR = round(Month_Revenue.loc[Update_Month], 2)
            ws0.cell(row=5, column=2).value = MR
            ws0.cell(row=5, column=2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            print("新增月份:", Update_Month, "的月營收:", MR)
            '''        更新月增率        '''
            MR_MG = round(MR_MonthGrowth.loc[Update_Month], 2)
            ws0.cell(row=5, column=3).value = MR_MG
            ws0.cell(row=5, column=3).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if MR_MG >= 0:
                ws0.cell(row=5, column=3).font = Font(color='FF0000')  # 紅色
            else:
                ws0.cell(row=5, column=3).font = Font(color='00FF00')  # 綠色
            print("新增", Update_Month, "的月增率:", MR_MG)
            '''        更新年增率        '''
            MR_YG = round(MR_YearGrowth.loc[Update_Month], 2)
            ws0.cell(row=5, column=4).value = MR_YG
            ws0.cell(row=5, column=4).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if MR_YG >= 0:
                ws0.cell(row=5, column=4).font = Font(color='FF0000')  # 紅色
            else:
                ws0.cell(row=5, column=4).font = Font(color='00FF00')  # 綠色
            print("新增", Update_Month, "的年增率:", round(MR_YG, 2))

            '''        更新當月最高、最低、平均收盤價        '''
            Update_Month_str = Update_Month.strftime('%Y-%m')
            ws0.cell(row=5, column=6).value = round(price.loc[Update_Month_str].max(), 2)
            ws0.cell(row=5, column=6).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws0.cell(row=5, column=7).value = round(price.loc[Update_Month_str].mean(), 2)
            ws0.cell(row=5, column=7).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws0.cell(row=5, column=8).value = round(price.loc[Update_Month_str].min(), 2)
            ws0.cell(row=5, column=8).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            '''        更新長、短期年增        '''
            ws0.cell(row=5, column=19).value = MAG_3M.loc[Update_Month]
            ws0.cell(row=5, column=19).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws0.cell(row=5, column=20).value = MAG_12M.loc[Update_Month]
            ws0.cell(row=5, column=20).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wb.save(path)
    print("Month Report end")


def Update_Directors_and_supervisors(path, Stock_ID):
    # 設定headers
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36'
    }

    url = "https://goodinfo.tw/StockInfo/StockDirectorSharehold.asp?STOCK_ID=" + str(Stock_ID)
    r = requests.get(url, headers=headers)
    r.encoding = "utf-8"

    dfs = pd.read_html(StringIO(r.text))
    df = pd.concat([df for df in dfs if df.shape[1] > 15 and df.shape[0] > 30])
    idx = pd.IndexSlice
    df = df.loc[idx[:], idx[['月別', '全體董監持股'], :]]
    df.columns = df.columns.get_level_values(1)
    df = df.set_index(["月別"])

    df["持股(%)"] = pd.to_numeric(df["持股(%)"], errors="coerce")
    df = df[~ df["持股(%)"].isnull()].dropna()["持股(%)"]

    def ChangeName(str):
        dt_obj = datetime.datetime.strptime(str, '%Y/%m')
        dt_str = datetime.datetime.strftime(dt_obj, '%Y-%m')
        return dt_str

    df = df.rename(index=lambda s: ChangeName(s))

    data = []
    index = []
    for cell in list(ws0.columns)[9]:
        data.append(cell.value)
    data = data[4:]
    for cell in list(ws0.columns)[0]:
        index.append(cell.value)
    index = index[4:]

    DataNow = pd.DataFrame({'date': index, 'Data': data})
    DataNow = DataNow[DataNow['date'].notnull()].rename(index=lambda s: s + 5)
    while datetime.datetime.strftime(DataNow['date'].iloc[0], "%Y-%m") != df.index[0]:
        DataNow = DataNow.drop(DataNow.index[0])
        print("drop one row")
    UpdateData = DataNow[DataNow['Data'].isnull()]

    pd.options.mode.chained_assignment = None

    for n in range(len(UpdateData)):
        date = UpdateData['date'].iloc[n]
        date_str = datetime.datetime.strftime(date, "%Y-%m")
        UpdateData['Data'].iloc[n] = df.loc[date_str]
        r = UpdateData.index[n]
        if ws0.cell(row=r, column=1).value == date:
            ws0.cell(row=r, column=10).value = UpdateData['Data'].iloc[n]
            ws0.cell(row=r, column=10).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            print("更新月份: " + date_str + " 的資料: " + str(ws0.cell(row=r, column=10).value))

    wb.save(path)
    time.sleep(20)
    print("Directors and supervisors end")


# def Update_PER(Stock_ID,path):
#
#     '''    從資料庫獲取季報最新日期    '''
#     # *未結束年度之EPS預估值, 以最近四季之合計EPS取代之, 例如: 某股票EPS僅公布至今年第三季, 則
#     # 今年之預估EPS = 去年第四季至今年第三季之合計EPS。
#     # https://goodinfo.tw/StockInfo/ShowK_ChartFlow.asp?RPT_CAT=PER&STOCK_ID=2330&CHT_CAT=QUAR
#
#     '''    從資料庫獲取季報最新日期    '''
#     Equity_Season = data.get("股本合計", 1)
#     Equity_Season = Equity_Season[Stock_ID]
#
#     '''    時間判斷    '''
#     # 改成用資料庫的最新時間尤佳
#     latest_date = Equity_Season.dropna().index[-1]
#     latest_date_str = Season_determination(latest_date)
#
#     # now = datetime.datetime.now()
#     now = datetime.datetime(2020,12,1)
#     Season_now = Season_determination(now)
#
#     table_month = ws4["A16"].value
#     add_row_num = 4 * (int(latest_date_str[0:4]) - int(table_month[0:4])) + (
#                 int(latest_date_str[-1]) - int(table_month[-1]))
#
#     print(latest_date_str)
#     print(table_month)
#     print(add_row_num)
#
#     if add_row_num <= 0:
#         print("Update PER this year.")
#     else:
#         print("Increase PER this season and update PER this year.")
#
#     PER_data = [ws4.cell(row=n, column=1).value[0:4] for n in range(16, 20) if ws4.cell(row=n, column=1).value]
#     Update_row = 0
#     for n in range(len(PER_data)):
#         if PER_data[n] == now.strftime("%Y"):
#             Update_row += 1
#
#     print(Update_row)
#
#     get_data_num = Update_row + add_row_num + 4
#     Equity = data.get("股本合計", get_data_num) * 0.00001  # 單位: 億
#     profit_after_tax = data.get("本期淨利（淨損）", get_data_num) * 0.00001  # 單位: 億
#
#     price_num = (Update_row + add_row_num) * 100
#     price = data.get("收盤價", price_num)
#
#     Equity = Equity[Stock_ID]
#     profit_after_tax = profit_after_tax[Stock_ID]
#     price = price[Stock_ID]
#     price_Q1 = price[price.index.month == 1].append(price[price.index.month == 2]).append(price[price.index.month == 3]).sort_index()
#     price_Q2 = price[price.index.month == 4].append(price[price.index.month == 5]).append(price[price.index.month == 6]).sort_index()
#     price_Q3 = price[price.index.month == 7].append(price[price.index.month == 8]).append(price[price.index.month == 9]).sort_index()
#     price_Q4 = price[price.index.month == 10].append(price[price.index.month == 11]).append(price[price.index.month == 12]).sort_index()
#
#     EPS = profit_after_tax / (Equity / 10)
#     Estimated_EPS = EPS.rolling(4).sum()
#
#     Start = 16
#     End = 16 + Update_row
#     for add_row in range(Start, End):
#
#         Update_date = Season2Month(ws4.cell(row=add_row, column=1).value)
#         Update_Season = ws4.cell(row=add_row, column=1).value
#         if Update_Season[-1] == "1":
#             PRICE = price_Q1.loc[Update_Season[0:4]][-1]
#         elif Update_Season[-1] == "2":
#             PRICE = price_Q2.loc[Update_Season[0:4]][-1]
#         elif Update_Season[-1] == "3":
#             PRICE = price_Q3.loc[Update_Season[0:4]][-1]
#         else:
#             PRICE = price_Q4.loc[Update_Season[0:4]][-1]
#         E_EPS = Estimated_EPS.loc[Update_date][-1]
#         PER = PRICE / E_EPS
#
#         Write2Excel(PER, rounds=2, sheet=ws4, rows=add_row, cols=2, string="更新PER", date=Update_Season)
#
#     add_row_num *= -1
#
#     for add_row in range(add_row_num, 0, 1):
#
#         ws4.insert_rows(16, amount=1)
#
#         Update_Season_date = Equity.index[add_row]
#         Update_Season_str = Update_Season_date.strftime('%Y-%m')
#
#         '''  新增季度標籤  '''
#         Update_Season = Season_determination(Update_Season_date)
#
#         ws4.cell(row=16, column=1).value = Update_Season
#         ws4.cell(row=16, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#         ws4.cell(row=16, column=1).fill = PatternFill(fill_type="solid", fgColor="FFEE99")
#         print("新增標籤:", ws4.cell(row=16, column=1).value)
#
#         '''  新增本益比  '''
#         if Update_Season[-1] == "1":
#             PRICE = price_Q1.loc[Update_Season[0:4]][-1]
#         elif Update_Season[-1] == "2":
#             PRICE = price_Q2.loc[Update_Season[0:4]][-1]
#         elif Update_Season[-1] == "3":
#             PRICE = price_Q3.loc[Update_Season[0:4]][-1]
#         else:
#             PRICE = price_Q4.loc[Update_Season[0:4]][-1]
#         E_EPS = Estimated_EPS.loc[Update_Season_str][-1]
#         PER = PRICE / E_EPS
#
#         Write2Excel(PER, rounds=2, sheet=ws4, rows=16, cols=2, string="新增PER", date=Update_Season)
#
#     if Season_now != ws4.cell(row=16, column=1).value:
#         ws4.insert_rows(16, amount=1)
#
#         ws4.cell(row=16, column=1).value = Season_now
#         ws4.cell(row=16, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#         ws4.cell(row=16, column=1).fill = PatternFill(fill_type="solid", fgColor="FFEE99")
#         print("新增標籤:", ws4.cell(row=16, column=1).value)
#
#         if Season_now[-1] == "1":
#             PRICE = price_Q1.loc[Season_now[0:4]][-1]
#         elif Season_now[-1] == "2":
#             PRICE = price_Q2.loc[Season_now[0:4]][-1]
#         elif Season_now[-1] == "3":
#             PRICE = price_Q3.loc[Season_now[0:4]][-1]
#         else:
#             PRICE = price_Q4.loc[Season_now[0:4]][-1]
#         E_EPS = Estimated_EPS.iloc[-1]
#         PER = PRICE / E_EPS
#
#         Write2Excel(PER, rounds=2, sheet=ws4, rows=16, cols=2, string="新增PER", date=Season_now)
#
#     wb.save(path)


# 顯示指定資料夾的內容
target_folder = 'D:\GOOGLE 雲端硬碟\Google 雲端硬碟\個人計畫追蹤\財報分析\台股'
file = show_folder_content(target_folder, prefix="O_", postfix=".xlsx")
index = []
dictionary = {}
for num in file[0:]:
    idx = ''.join([x for x in num if x.isdigit()])
    dictionary[idx] = num
    index.append(idx)
print("合格的股票代碼有: ", index)

Stock_ID = input("輸入欲更新股票之代碼: ", )
Stock_ID = str(Stock_ID)

if Stock_ID == "all":
    for id in index[0:]:
        try:
            File_path = dictionary[id]
            print("正在更新: ", id)

            wb = openpyxl.load_workbook(File_path)

            # 可以試著用for迴圈將 wb.sheetnames 寫進新的list之中，之後可以嘗試
            ws0 = wb["月財報"]
            ws1 = wb["季財報"]
            ws2 = wb["現金流量"]
            ws3 = wb["進出場參考"]
            ws4 = wb["合理價推估"]

            Update_Monthly_report(File_path, id)
            Update_Directors_and_supervisors(File_path, id)
            # Update_PER(id, path=File_path)
        except:
            print("資料庫無此id資訊")
else:
    try:
        File_path = dictionary[Stock_ID]

        wb = openpyxl.load_workbook(File_path)

        # 可以試著用for迴圈將 wb.sheetnames 寫進新的list之中，之後可以嘗試
        ws0 = wb["月財報"]
        ws1 = wb["季財報"]
        ws2 = wb["現金流量"]
        ws3 = wb["進出場參考"]
        ws4 = wb["合理價推估"]

        Update_Monthly_report(File_path, Stock_ID)
        Update_Directors_and_supervisors(File_path, Stock_ID)
        # Update_PER(id, path=File_path)
    except:
        print("ID 輸入有誤")

'''
手動計算月增率、年增率以及3/12個月平均:

    MAG = []
    for idx in range(add_row_num+12):
        MAG_idx = -1 * idx - 1
        YGratio = (Revenue_Month.iloc[MAG_idx] - Revenue_Month.iloc[(MAG_idx-12)]) / Revenue_Month.iloc[(MAG_idx-12)] * 100
        MAG.append(round(YGratio, 2))
    MAG = pd.Series(MAG)
    MAG_3M = MAG.rolling(3).mean().dropna().reset_index(drop=True)
    MAG_3M = round(MAG_3M, 2)
    MAG_12M = MAG.rolling(12).mean().dropna().reset_index(drop=True)
    MAG_12M = round(MAG_12M, 2)

    MGrowth_index = Revenue_index - 1
    Revenue_Month_MGrowth = (Revenue_Month.iloc[Revenue_index] - Revenue_Month.iloc[MGrowth_index]) / Revenue_Month.iloc[MGrowth_index]
    Revenue_Month_MGrowth = Revenue_Month_MGrowth * 100

    YGrowth_index = Revenue_index - 12
    Revenue_Month_YGrowth = (Revenue_Month.iloc[Revenue_index] - Revenue_Month.iloc[YGrowth_index]) / Revenue_Month.iloc[YGrowth_index]
    Revenue_Month_YGrowth = Revenue_Month_YGrowth * 100
'''
